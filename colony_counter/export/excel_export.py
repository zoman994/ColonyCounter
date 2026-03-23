"""Excel export — generates .xlsx with summary + detail sheets."""
import datetime

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from colony_counter.core.constants import C


def export_excel(filepath, rows, version=C.VERSION):
    """Export to Excel. rows = list of dicts from format_result_row()."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Результаты"

    hf = PatternFill("solid", fgColor="2E4DA0")
    hfn = Font(color="FFFFFF", bold=True, size=11)
    tf = PatternFill("solid", fgColor="D6E4BC")
    af = PatternFill("solid", fgColor="EEF2FB")
    thin = Side(style='thin', color="AAAAAA")
    brd = Border(left=thin, right=thin, top=thin, bottom=thin)
    ca = Alignment(horizontal='center', vertical='center')
    la = Alignment(horizontal='left', vertical='center')

    ws.merge_cells("A1:H1")
    ws["A1"].value = f"Colony Counter v{version}"
    ws["A1"].font = Font(bold=True, size=14, color="1A1A6E")
    ws["A1"].alignment = ca
    ws["A2"] = f"Дата: {datetime.datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws["A2"].font = Font(italic=True, size=9, color="555555")

    headers = ["#", "Файл", "Авто", "Искл.", "Ручн.", "ИТОГО",
               "Одиноч.", "Кластер.", "Ср.площ.px", "Ср.площ.мм\u00b2",
               "CFU/мл", "Путь"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=ci, value=h)
        c.fill = hf
        c.font = hfn
        c.alignment = ca
        c.border = brd
    widths = [4, 28, 10, 10, 10, 12, 14, 12, 14, 14, 14, 50]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row_n = 5
    total_sum = 0
    for idx, r in enumerate(rows, 1):
        fl = af if idx % 2 == 0 else PatternFill()
        vals = [idx, r['name'], r['auto'],
                f"-{r['excluded']}" if r['excluded'] else 0,
                r['manual'], r['total'], r['singles'], r['clusters'],
                r['avg_area_px'], r['avg_area_mm2'] or "",
                r['cfu'] or "", r['path']]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row_n, column=ci, value=v)
            c.border = brd
            c.fill = fl
            c.alignment = ca if ci not in (2, 10) else la
            if ci == 6:
                c.font = Font(bold=True)
        total_sum += r['total']
        row_n += 1

    ws.cell(row=row_n, column=2, value="ИТОГО").font = Font(bold=True)
    ws.cell(row=row_n, column=6, value=total_sum).font = Font(bold=True, size=12)
    for ci in range(1, 13):
        c = ws.cell(row=row_n, column=ci)
        c.fill = tf
        c.border = brd
        c.alignment = ca

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:L{row_n - 1}"
    wb.save(filepath)
