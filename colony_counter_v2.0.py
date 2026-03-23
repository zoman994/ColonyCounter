#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Colony Counter v2.0
Автоматический подсчет колоний бактерий на чашках Петри.
Dark UI inspired by CloneTracker (zinc-950 + emerald accent).
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog
import cv2
import numpy as np
from PIL import Image, ImageTk, ImageDraw, ImageFont
import openpyxl
from openpyxl.styles import Font as XlFont, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os
import sys
import json
import datetime
import threading
import tempfile
from pathlib import Path

try:
    import matplotlib
    matplotlib.use('TkAgg')
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    from matplotlib.figure import Figure
    HAS_MATPLOTLIB = True
except ImportError:
    HAS_MATPLOTLIB = False

try:
    from skimage.feature import peak_local_max
    HAS_SKIMAGE = True
except ImportError:
    HAS_SKIMAGE = False


# ═══════════════════════════ THEME ═══════════════════════════════════════════

_THEMES = {
    'dark': dict(
        BG='#09090b', BG1='#18181b', BG2='#27272a', BG3='#3f3f46',
        BORDER='#27272a', BORDER_HI='#059669',
        FG='#d4d4d8', FG2='#a1a1aa', FG3='#71717a', FG4='#52525b',
        ACCENT='#10b981', ACCENT_DIM='#0d3024',
        ACCENT_BTN='#059669', ACCENT_HOV='#047857',
        RED='#ef4444', RED_DIM='#7f1d1d',
        ORANGE='#f59e0b', YELLOW='#eab308',
        CANVAS_BG='#0c0c0e',
    ),
    'light': dict(
        BG='#f4f4f5',    # zinc-100
        BG1='#ffffff',   # white (panels)
        BG2='#e4e4e7',   # zinc-200 (inputs)
        BG3='#d4d4d8',   # zinc-300 (buttons secondary)
        BORDER='#d4d4d8', BORDER_HI='#059669',
        FG='#18181b',    # zinc-900 (primary text)
        FG2='#3f3f46',   # zinc-700
        FG3='#52525b',   # zinc-600 (secondary)
        FG4='#71717a',   # zinc-500 (muted)
        ACCENT='#059669',  # emerald-600 (darker for light bg)
        ACCENT_DIM='#d1fae5',  # emerald-100
        ACCENT_BTN='#059669', ACCENT_HOV='#047857',
        RED='#dc2626', RED_DIM='#fecaca',
        ORANGE='#d97706', YELLOW='#ca8a04',
        CANVAS_BG='#e8e8ec',
    ),
}

class T:
    """Dynamic theme — call T.apply('dark') or T.apply('light') to switch."""
    # Fonts (bigger than before)
    FONT        = ('Consolas', 11)
    FONT_SM     = ('Consolas', 10)
    FONT_XS     = ('Consolas', 9)
    FONT_LG     = ('Consolas', 13)
    FONT_TITLE  = ('Consolas', 15, 'bold')
    FONT_HDR    = ('Consolas', 12, 'bold')

    # Colors — set by apply()
    BG = BG1 = BG2 = BG3 = ''
    BORDER = BORDER_HI = ''
    FG = FG2 = FG3 = FG4 = ''
    ACCENT = ACCENT_DIM = ACCENT_BTN = ACCENT_HOV = ''
    RED = RED_DIM = ORANGE = YELLOW = CANVAS_BG = ''

    _current = 'dark'
    _listeners = []   # list of callables to invoke after theme switch

    @classmethod
    def apply(cls, name='dark'):
        cls._current = name
        colors = _THEMES.get(name, _THEMES['dark'])
        for k, v in colors.items():
            setattr(cls, k, v)

    @classmethod
    def toggle(cls):
        cls.apply('light' if cls._current == 'dark' else 'dark')
        for cb in cls._listeners:
            cb()

    @classmethod
    def is_dark(cls):
        return cls._current == 'dark'

# Apply default theme immediately
T.apply('dark')


# ═══════════════════════════ UNICODE-SAFE I/O ════════════════════════════════

def cv_imread(path):
    """cv2.imread that works with unicode/cyrillic paths on Windows."""
    data = np.fromfile(path, dtype=np.uint8)
    return cv2.imdecode(data, cv2.IMREAD_COLOR)

def cv_imwrite(path, img, params=None):
    """cv2.imwrite that works with unicode/cyrillic paths on Windows."""
    ext = os.path.splitext(path)[1]
    ok, buf = cv2.imencode(ext, img, params or [])
    if ok:
        buf.tofile(path)
    return ok


# ═══════════════════════════ CONSTANTS ═══════════════════════════════════════

class C:
    VERSION = "2.0"
    MAX_IMAGE_DIM = 2000
    HOUGH_BLUR_KERNEL = (21, 21)
    HOUGH_MIN_R_RATIO = 0.28
    HOUGH_MAX_R_RATIO = 0.56
    HOUGH_FALLBACK_R_RATIO = 0.44
    HOUGH_PARAM1 = 60
    HOUGH_PARAM2 = 35
    DISH_MASK_RATIO = 0.96
    BG_MORPH_KERNEL = 71
    CLAHE_CLIP = 3.0
    CLAHE_TILE = (8, 8)
    MORPH_KERNEL = 3
    LABEL_DARK_THRESH = 60
    LABEL_MIN_AREA = 0.02
    LABEL_MAX_AREA = 0.40
    LABEL_MIN_ASPECT = 2.5
    LABEL_MIN_FILL = 0.4
    LABEL_DILATE_K = 30
    LABEL_LIGHT_THRESH = 200
    LABEL_LIGHT_STD_THRESH = 15
    LABEL_LIGHT_MIN_ASPECT = 2.0
    LABEL_LIGHT_MIN_FILL = 0.45
    BUBBLE_CIRC = 0.90
    BUBBLE_AREA_MULT = 1.8
    MAX_ASPECT = 6.0
    MIN_ASPECT = 0.16
    ELONGATION_THRESH = 3.5
    SOLIDITY_THRESH = 0.45
    CLUSTER_AREA_MULT = 1.8
    LOG_HIST_MAX_BINS = 30
    LOG_HIST_MIN_BINS = 10
    WS_MIN_DIST_FACTOR = 0.5
    WS_THRESH_FACTOR = 0.2
    WS_SANITY_LO = 0.5
    WS_SANITY_HI = 1.5
    LEARN_ALPHA = 0.30
    LEARN_MIN_AUTO = 5
    LEARN_MIN_RATIO = 0.04
    LEARN_MAX_DELTA = 5
    LEARN_DELTA_K = 12
    ZOOM_MIN = 0.25
    ZOOM_MAX = 10.0
    ZOOM_FACTOR = 1.15
    HSV_S_LO = 30


# ═══════════════════════════ CUSTOM WIDGETS ══════════════════════════════════

class DarkButton(tk.Frame):
    """Flat dark button matching CloneTracker style."""

    def __init__(self, parent, text, command=None, variant='primary', small=False, **kw):
        super().__init__(parent, **kw)
        if T.is_dark():
            colors = {
                'primary':   (T.ACCENT_BTN, T.ACCENT_HOV, '#ffffff'),
                'secondary': (T.BG3,        '#52525b',     T.FG2),
                'danger':    (T.RED_DIM,    '#991b1b',     '#fca5a5'),
                'ghost':     (T.BG1,        T.BG2,         T.FG3),
            }
        else:
            colors = {
                'primary':   (T.ACCENT_BTN, T.ACCENT_HOV, '#ffffff'),
                'secondary': (T.BG2,        T.BG3,         T.FG2),
                'danger':    (T.RED_DIM,    '#fca5a5',     T.RED),
                'ghost':     (T.BG1,        T.BG2,         T.FG3),
            }
        self._bg, self._hover, self._fg = colors.get(variant, colors['primary'])
        self._cmd = command
        pad_x = 6 if small else 10
        pad_y = 2 if small else 5
        font = T.FONT_XS if small else T.FONT_SM

        self.config(bg=self._bg, cursor='hand2')
        self.lbl = tk.Label(self, text=text, bg=self._bg, fg=self._fg,
                            font=font, padx=pad_x, pady=pad_y)
        self.lbl.pack()

        for w in (self, self.lbl):
            w.bind('<Enter>', self._on_enter)
            w.bind('<Leave>', self._on_leave)
            w.bind('<Button-1>', self._on_click)

    def _on_enter(self, _e):
        self.config(bg=self._hover)
        self.lbl.config(bg=self._hover)

    def _on_leave(self, _e):
        self.config(bg=self._bg)
        self.lbl.config(bg=self._bg)

    def _on_click(self, _e):
        if self._cmd:
            self._cmd()

    def set_text(self, text):
        self.lbl.config(text=text)


class DarkCheck(tk.Frame):
    """Minimal dark checkbox."""

    def __init__(self, parent, text, variable, **kw):
        super().__init__(parent, bg=T.BG1, **kw)
        self._var = variable
        self._box = tk.Canvas(self, width=14, height=14, bg=T.BG2,
                              highlightthickness=1, highlightbackground=T.BORDER,
                              cursor='hand2')
        self._box.pack(side=tk.LEFT, padx=(0, 6))
        self._lbl = tk.Label(self, text=text, bg=T.BG1, fg=T.FG3,
                             font=T.FONT_SM, cursor='hand2')
        self._lbl.pack(side=tk.LEFT)
        self._box.bind('<Button-1>', self._toggle)
        self._lbl.bind('<Button-1>', self._toggle)
        self._var.trace_add('write', lambda *_: self._draw())
        self._draw()

    def _toggle(self, _e=None):
        self._var.set(not self._var.get())

    def _draw(self):
        self._box.delete('check')
        if self._var.get():
            self._box.config(highlightbackground=T.ACCENT)
            self._box.create_line(3, 7, 6, 11, fill=T.ACCENT, width=2, tags='check')
            self._box.create_line(6, 11, 12, 3, fill=T.ACCENT, width=2, tags='check')
        else:
            self._box.config(highlightbackground=T.BORDER)


class DarkSlider(tk.Frame):
    """Label + Scale + Value display, dark themed."""

    def __init__(self, parent, label, variable, from_, to, resolution=1, **kw):
        super().__init__(parent, bg=T.BG1, **kw)
        self._var = variable
        self._res = resolution
        tk.Label(self, text=label, bg=T.BG1, fg=T.FG3,
                 font=T.FONT_XS, anchor='w').pack(fill=tk.X)
        row = tk.Frame(self, bg=T.BG1)
        row.pack(fill=tk.X, pady=(2, 0))
        self._scale = tk.Scale(row, from_=from_, to=to, orient=tk.HORIZONTAL,
                               variable=variable, resolution=resolution,
                               bg=T.BG2, fg=T.FG, troughcolor=T.BG,
                               highlightthickness=0, sliderrelief='flat',
                               activebackground=T.ACCENT, font=T.FONT_XS,
                               showvalue=False, bd=0, length=160)
        self._scale.pack(side=tk.LEFT, fill=tk.X, expand=True)
        self._val_lbl = tk.Label(row, text='', bg=T.BG1, fg=T.ACCENT,
                                 font=T.FONT_SM, width=6, anchor='e')
        self._val_lbl.pack(side=tk.RIGHT, padx=(4, 0))
        variable.trace_add('write', self._update_val)
        self._update_val()

    def _update_val(self, *_):
        self._val_lbl.config(text=str(int(self._var.get())))


class DarkSection(tk.Frame):
    """Collapsible section with header label and border."""

    def __init__(self, parent, title, **kw):
        super().__init__(parent, bg=T.BG1, highlightthickness=1,
                         highlightbackground=T.BORDER, **kw)
        hdr = tk.Frame(self, bg=T.BG1)
        hdr.pack(fill=tk.X, padx=8, pady=(6, 2))
        tk.Label(hdr, text=title.upper(), bg=T.BG1, fg=T.FG4,
                 font=T.FONT_XS).pack(anchor=tk.W)
        self.body = tk.Frame(self, bg=T.BG1)
        self.body.pack(fill=tk.X, padx=8, pady=(0, 8))


# ═══════════════════════════ IMAGE PROCESSING ════════════════════════════════
# (identical to previous v2.0 — all algorithms preserved)

class ImageProcessor:
    @staticmethod
    def _hough_circles(gray):
        h, w = gray.shape
        blurred = cv2.GaussianBlur(gray, C.HOUGH_BLUR_KERNEL, 0)
        min_r = int(min(h, w) * C.HOUGH_MIN_R_RATIO)
        max_r = int(min(h, w) * C.HOUGH_MAX_R_RATIO)
        circles = cv2.HoughCircles(
            blurred, cv2.HOUGH_GRADIENT, dp=1, minDist=min_r,
            param1=C.HOUGH_PARAM1, param2=C.HOUGH_PARAM2,
            minRadius=min_r, maxRadius=max_r)
        return circles, min_r, max_r

    @staticmethod
    def detect_dish(gray):
        circles, _, _ = ImageProcessor._hough_circles(gray)
        h, w = gray.shape
        if circles is not None:
            circles = sorted(np.round(circles[0]).astype(int).tolist(),
                             key=lambda c: c[2], reverse=True)
            return tuple(circles[0])
        return (w // 2, h // 2, int(min(h, w) * C.HOUGH_FALLBACK_R_RATIO))

    @staticmethod
    def detect_dishes(gray, max_dishes=2):
        circles, _, _ = ImageProcessor._hough_circles(gray)
        h, w = gray.shape
        if circles is None:
            return [(w // 2, h // 2, int(min(h, w) * C.HOUGH_FALLBACK_R_RATIO))]
        circles = sorted(np.round(circles[0]).astype(int).tolist(),
                         key=lambda c: c[2], reverse=True)
        result = []
        for cx, cy, r in circles:
            if all(((cx - rx)**2 + (cy - ry)**2)**0.5 >= max(r, rr)*0.7
                   for rx, ry, rr in result):
                result.append((int(cx), int(cy), int(r)))
            if len(result) >= max_dishes:
                break
        return result or [(w // 2, h // 2, int(min(h, w) * C.HOUGH_FALLBACK_R_RATIO))]

    @staticmethod
    def normalize_background(gray, mask):
        k = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (C.BG_MORPH_KERNEL, C.BG_MORPH_KERNEL))
        bg = cv2.morphologyEx(gray, cv2.MORPH_DILATE, k)
        bg = cv2.morphologyEx(bg, cv2.MORPH_ERODE, k)
        diff = cv2.subtract(bg, gray)
        return cv2.bitwise_and(diff, diff, mask=mask)

    @staticmethod
    def detect_label_mask_dark(gray, dish_mask):
        masked = cv2.bitwise_and(gray, gray, mask=dish_mask)
        _, dark = cv2.threshold(masked, C.LABEL_DARK_THRESH, 255, cv2.THRESH_BINARY_INV)
        dark = cv2.bitwise_and(dark, dark, mask=dish_mask)
        contours, _ = cv2.findContours(dark, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        label_mask = np.zeros(gray.shape, dtype=np.uint8)
        dish_area = cv2.countNonZero(dish_mask)
        for cnt in contours:
            area = cv2.contourArea(cnt)
            if area < dish_area * C.LABEL_MIN_AREA or area > dish_area * C.LABEL_MAX_AREA:
                continue
            rect = cv2.minAreaRect(cnt)
            rw, rh = rect[1]
            if rw * rh == 0: continue
            aspect = max(rw, rh) / (min(rw, rh) + 1e-5)
            fill = area / (rw * rh)
            if aspect > C.LABEL_MIN_ASPECT and fill > C.LABEL_MIN_FILL:
                cv2.fillConvexPoly(label_mask, np.intp(cv2.boxPoints(rect)), 255)
        if cv2.countNonZero(label_mask) > 0:
            k = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (C.LABEL_DILATE_K, C.LABEL_DILATE_K))
            label_mask = cv2.dilate(label_mask, k, iterations=2)
        return label_mask

    @staticmethod
    def detect_label_mask_light(gray, dish_mask):
        masked = cv2.bitwise_and(gray, gray, mask=dish_mask)
        _, bright = cv2.threshold(masked, C.LABEL_LIGHT_THRESH, 255, cv2.THRESH_BINARY)
        bright = cv2.bitwise_and(bright, bright, mask=dish_mask)
        blur = cv2.GaussianBlur(gray.astype(np.float32), (31, 31), 0)
        blur2 = cv2.GaussianBlur((gray.astype(np.float32))**2, (31, 31), 0)
        local_std = np.sqrt(np.maximum(blur2 - blur**2, 0))
        uniform = (local_std < C.LABEL_LIGHT_STD_THRESH).astype(np.uint8) * 255
        combined = cv2.bitwise_and(bright, uniform)
        combined = cv2.bitwise_and(combined, combined, mask=dish_mask)
        contours, _ = cv2.findContours(combined, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        label_mask = np.zeros(gray.shape, dtype=np.uint8)
        dish_area = cv2.countNonZero(dish_mask)
        for cnt in contours:
            area = cv2.contourArea(cnt)
            if area < dish_area * C.LABEL_MIN_AREA or area > dish_area * C.LABEL_MAX_AREA:
                continue
            rect = cv2.minAreaRect(cnt)
            rw, rh = rect[1]
            if rw * rh == 0: continue
            aspect = max(rw, rh) / (min(rw, rh) + 1e-5)
            fill = area / (rw * rh)
            if aspect > C.LABEL_LIGHT_MIN_ASPECT and fill > C.LABEL_LIGHT_MIN_FILL:
                cv2.fillConvexPoly(label_mask, np.intp(cv2.boxPoints(rect)), 255)
        if cv2.countNonZero(label_mask) > 0:
            k = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (C.LABEL_DILATE_K, C.LABEL_DILATE_K))
            label_mask = cv2.dilate(label_mask, k, iterations=2)
        return label_mask

    @staticmethod
    def detect_label_mask(gray, dish_mask, detect_light=False):
        dark = ImageProcessor.detect_label_mask_dark(gray, dish_mask)
        if detect_light:
            light = ImageProcessor.detect_label_mask_light(gray, dish_mask)
            return cv2.bitwise_or(dark, light)
        return dark

    @staticmethod
    def estimate_single_colony_area(areas):
        if len(areas) < 3: return float(np.median(areas))
        log_a = np.log(areas[areas > 0])
        n_bins = min(C.LOG_HIST_MAX_BINS, max(C.LOG_HIST_MIN_BINS, len(log_a)//3))
        counts, edges = np.histogram(log_a, bins=n_bins)
        pk = int(np.argmax(counts))
        lo, hi = edges[max(0, pk-1)], edges[min(len(edges)-1, pk+2)]
        pa = areas[(log_a >= lo) & (log_a <= hi)]
        if len(pa) == 0: return float(np.exp((edges[pk]+edges[pk+1])/2))
        return float(np.median(pa))

    @staticmethod
    def watershed_per_component(cnt, avg_area, h, w):
        """Watershed via peak_local_max — same as v1.1 (proven stable)."""
        mask = np.zeros((h, w), dtype=np.uint8)
        cv2.drawContours(mask, [cnt], -1, 255, -1)
        dist = cv2.distanceTransform(mask, cv2.DIST_L2, 5)
        if dist.max() < 2: return 1
        er = max(3, int(np.sqrt(avg_area / np.pi)))
        md = max(3, int(er * C.WS_MIN_DIST_FACTOR))
        ta = max(2.0, float(dist.max()) * C.WS_THRESH_FACTOR)
        if HAS_SKIMAGE:
            coords = peak_local_max(dist, min_distance=md, threshold_abs=ta, labels=mask)
            return max(1, len(coords))
        return max(1, round(cv2.contourArea(cnt)/avg_area))

    @staticmethod
    def contour_features(cnt):
        area = cv2.contourArea(cnt)
        per = cv2.arcLength(cnt, True)
        circ = (4*np.pi*area/(per*per)) if per > 0 else 0.0
        x, y, bw, bh = cv2.boundingRect(cnt)
        ar = bw/bh if bh > 0 else 1.0
        hull_a = cv2.contourArea(cv2.convexHull(cnt))
        sol = area/hull_a if hull_a > 0 else 1.0
        M = cv2.moments(cnt)
        cx = int(M['m10']/M['m00']) if M['m00'] > 0 else x+bw//2
        cy = int(M['m01']/M['m00']) if M['m00'] > 0 else y+bh//2
        return dict(area=area, circularity=circ, aspect_ratio=ar, solidity=sol, cx=cx, cy=cy)

    @staticmethod
    def fill_contour_with_circles(cnt, n, col_radius):
        if n <= 0: return []
        x, y, bw, bh = cv2.boundingRect(cnt)
        if n == 1:
            M = cv2.moments(cnt)
            if M['m00'] > 0: return [(int(M['m10']/M['m00']), int(M['m01']/M['m00']))]
            return [(x+bw//2, y+bh//2)]
        r = max(2, col_radius)
        dy, dx = max(1, int(r*1.732)), max(1, int(r*2.0))
        m = np.zeros((bh+2, bw+2), dtype=np.uint8)
        sh = cnt.copy(); sh[:,:,0] -= x; sh[:,:,1] -= y
        cv2.drawContours(m, [sh], -1, 255, -1)
        cands = []
        for row, yi in enumerate(range(r, bh, dy)):
            ox = r if row % 2 else 0
            for xi in range(ox, bw, dx):
                if yi < m.shape[0] and xi < m.shape[1] and m[yi, xi] > 0:
                    cands.append((xi+x, yi+y))
        if not cands:
            M = cv2.moments(cnt)
            if M['m00'] > 0: return [(int(M['m10']/M['m00']), int(M['m01']/M['m00']))]
            return [(x+bw//2, y+bh//2)]
        if len(cands) <= n: return cands
        step = (len(cands) - 1) / max(1, n - 1)
        return [cands[min(int(i * step), len(cands) - 1)] for i in range(n)]

    @staticmethod
    def color_filter_mask(img_bgr, work_mask):
        hsv = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2HSV)
        gray = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2GRAY)
        wp = gray[work_mask > 0]
        if len(wp) == 0: return work_mask
        bg_med = np.median(wp)
        dark = (gray < bg_med * 0.7).astype(np.uint8) * 255
        sat = (hsv[:,:,1] > C.HSV_S_LO).astype(np.uint8) * 255
        combined = cv2.bitwise_or(dark, sat)
        return cv2.bitwise_and(combined, combined, mask=work_mask)

    def _process_single_dish(self, img, gray, cx, cy, r, params):
        h, w = img.shape[:2]
        dish_mask = np.zeros((h, w), dtype=np.uint8)
        cv2.circle(dish_mask, (cx, cy), max(1, int(r*C.DISH_MASK_RATIO)), 255, -1)
        has_label, label_mask, hidden_est = False, None, 0
        if bool(params.get('detect_label', True)):
            detect_light = bool(params.get('detect_light_label', False))
            label_mask = self.detect_label_mask(gray, dish_mask, detect_light=detect_light)
            has_label = cv2.countNonZero(label_mask) > 0
        work_mask = dish_mask.copy()
        if has_label and label_mask is not None:
            work_mask = cv2.bitwise_and(work_mask, cv2.bitwise_not(label_mask))
        normalized = self.normalize_background(gray, work_mask)
        clahe = cv2.createCLAHE(clipLimit=C.CLAHE_CLIP, tileGridSize=C.CLAHE_TILE)
        enhanced = clahe.apply(normalized)
        if bool(params.get('use_otsu', False)):
            wp = enhanced[work_mask > 0]
            if len(wp) > 0:
                ov, _ = cv2.threshold(wp, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
                thresh_val = max(5, int(ov * 0.48))
            else: thresh_val = int(params['threshold'])
        else: thresh_val = int(params['threshold'])
        _, binary = cv2.threshold(enhanced, thresh_val, 255, cv2.THRESH_BINARY)
        binary = cv2.bitwise_and(binary, binary, mask=work_mask)
        if bool(params.get('use_color_filter', False)):
            cm = self.color_filter_mask(img, work_mask)
            binary = cv2.bitwise_and(binary, cv2.dilate(cm, cv2.getStructuringElement(cv2.MORPH_ELLIPSE,(5,5)), iterations=1))
        km = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (C.MORPH_KERNEL, C.MORPH_KERNEL))
        binary = cv2.morphologyEx(binary, cv2.MORPH_OPEN, km, iterations=1)
        binary = cv2.morphologyEx(binary, cv2.MORPH_CLOSE, km, iterations=1)
        contours, _ = cv2.findContours(binary, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        min_a, max_a = max(1, int(params['min_area'])), max(2, int(params['max_area']))
        fb, uw, fe, fn = bool(params['filter_bubbles']), bool(params['use_watershed']), bool(params['filter_elongated']), bool(params['filter_nonconvex'])
        raw = []
        for cnt in contours:
            feat = self.contour_features(cnt)
            a = feat['area']
            if a < min_a: continue
            if fb and feat['circularity'] > C.BUBBLE_CIRC and a > max_a * C.BUBBLE_AREA_MULT: continue
            ar = feat['aspect_ratio']
            if ar > C.MAX_ASPECT or ar < C.MIN_ASPECT: continue
            if fe and a > min_a*3:
                rect = cv2.minAreaRect(cnt); rw, rh = rect[1]
                if min(rw, rh) > 0 and max(rw, rh)/min(rw, rh) > C.ELONGATION_THRESH: continue
            if fn and a > min_a*5 and feat['solidity'] < C.SOLIDITY_THRESH: continue
            raw.append(dict(contour=cnt, feat=feat))
        all_areas = np.array([o['feat']['area'] for o in raw if o['feat']['area'] <= max_a], dtype=float)
        if len(all_areas) >= 3: avg_area = self.estimate_single_colony_area(all_areas)
        elif len(all_areas) > 0: avg_area = float(np.median(all_areas))
        else: avg_area = float(min_a * 5)
        avg_area = max(avg_area, float(min_a))
        ct = avg_area * C.CLUSTER_AREA_MULT
        col_r = max(4, int(np.sqrt(avg_area / np.pi)))
        colonies, total = [], 0
        for obj in raw:
            cnt, feat, a = obj['contour'], obj['feat'], obj['feat']['area']
            if a <= ct:
                est, wsc = 1, [(feat['cx'], feat['cy'])]
            else:
                ae = max(1, round(a/avg_area))
                if uw:
                    wc = self.watershed_per_component(cnt, avg_area, h, w)
                    est = wc if (wc >= 2 and wc <= ae*C.WS_SANITY_HI and wc >= ae*C.WS_SANITY_LO) else ae
                else: est = ae
                wsc = self.fill_contour_with_circles(cnt, est, col_r)
            if not wsc: wsc = [(feat['cx'], feat['cy'])]
            total += est
            colonies.append(dict(contour=cnt, feat=feat, center=(feat['cx'], feat['cy']),
                                 ws_centers=wsc, estimated=est, is_cluster=(est > 1)))
        if has_label:
            da, wa = cv2.countNonZero(dish_mask), cv2.countNonZero(work_mask)
            if wa > 0 and wa < da*0.95: hidden_est = int((total/wa)*(da-wa))
        return dict(total=total, colony_count=len(colonies), cluster_count=sum(1 for c in colonies if c['is_cluster']),
                    avg_colony_area=avg_area, col_radius=col_r, dish=(cx, cy, r), colonies=colonies,
                    binary=binary, enhanced=enhanced, has_label=has_label, hidden_estimate=hidden_est,
                    label_mask=label_mask, work_mask=work_mask, threshold_used=thresh_val)

    def process(self, image_path, params, progress_cb=None):
        if progress_cb: progress_cb(0.05, "Загрузка...")
        img = cv_imread(image_path)
        if img is None: raise ValueError(f"Не удалось открыть: {image_path}")
        h0, w0 = img.shape[:2]; scale = 1.0
        if max(h0, w0) > C.MAX_IMAGE_DIM:
            scale = C.MAX_IMAGE_DIM / max(h0, w0)
            img = cv2.resize(img, (int(w0*scale), int(h0*scale)), interpolation=cv2.INTER_AREA)
        h, w = img.shape[:2]
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        if progress_cb: progress_cb(0.15, "Чашки...")
        ov = params.get('dish_overrides')
        dishes = [(int(x), int(y), int(r)) for x, y, r in ov] if ov else self.detect_dishes(gray)
        if progress_cb: progress_cb(0.25, "Обработка...")
        annotated, all_col, dish_res = img.copy(), [], []
        sn = bool(params.get('show_numbers', True))
        for di, (dcx, dcy, dr) in enumerate(dishes):
            if progress_cb: progress_cb(0.25 + 0.55*(di/max(1, len(dishes))), f"Чашка {di+1}...")
            dd = self._process_single_dish(img, gray, dcx, dcy, dr, params)
            for c in dd['colonies']: c['dish_idx'] = di
            all_col.extend(dd['colonies']); dish_res.append(dd)
            cv2.circle(annotated, (dcx, dcy), dr, (80, 80, 220), 2)
            if dd['has_label'] and dd['label_mask'] is not None:
                lc = np.zeros_like(annotated); lc[dd['label_mask'] > 0] = (0, 60, 180)
                cv2.addWeighted(annotated, 1.0, lc, 0.35, 0, annotated)
                lcs, _ = cv2.findContours(dd['label_mask'], cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
                cv2.drawContours(annotated, lcs, -1, (0, 100, 255), 2)
            cr = dd['col_radius']
            for col in dd['colonies']:
                if col['is_cluster']:
                    for cx_c, cy_c in col['ws_centers']:
                        cv2.circle(annotated, (cx_c, cy_c), cr, (20, 160, 255), 2)
                    if sn:
                        cc = col['center']
                        cv2.putText(annotated, str(col['estimated']), (cc[0]-8, cc[1]-cr-4),
                                    cv2.FONT_HERSHEY_SIMPLEX, 0.55, (0, 100, 220), 2, cv2.LINE_AA)
                else:
                    cv2.drawContours(annotated, [col['contour']], -1, (30, 200, 30), 2)
                    if sn: cv2.circle(annotated, col['center'], 3, (30, 200, 30), -1)
            if len(dishes) > 1:
                cv2.putText(annotated, f"#{di+1}: {dd['total']}", (dcx-30, max(15, dcy-dr-6)),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.7, (80, 180, 255), 2, cv2.LINE_AA)
        if progress_cb: progress_cb(0.85, "Аннотация...")
        gt = sum(d['total'] for d in dish_res)
        lbl = f"Kolonii: {gt}"
        (lw, lh), _ = cv2.getTextSize(lbl, cv2.FONT_HERSHEY_SIMPLEX, 1.1, 2)
        cv2.rectangle(annotated, (6, 6), (lw+18, lh+18), (30, 30, 30), -1)
        cv2.putText(annotated, lbl, (12, lh+10), cv2.FONT_HERSHEY_SIMPLEX, 1.1, (60, 220, 60), 2, cv2.LINE_AA)
        avg_a = float(np.mean([d['avg_colony_area'] for d in dish_res])) if dish_res else float(params['min_area']*5)
        if progress_cb: progress_cb(1.0, "Готово")
        return dict(total=gt, colony_count=len(all_col), cluster_count=sum(1 for c in all_col if c['is_cluster']),
                    avg_colony_area=avg_a, col_radius=max(4, int(np.sqrt(avg_a/np.pi))),
                    dish=dishes[0] if dishes else (w//2, h//2, int(min(h, w)*C.HOUGH_FALLBACK_R_RATIO)),
                    dishes=dishes, dish_results=dish_res, colonies=all_col, annotated=annotated, img_clean=img.copy(),
                    binary=dish_res[0]['binary'] if dish_res else np.zeros((h, w), np.uint8),
                    enhanced=dish_res[0]['enhanced'] if dish_res else np.zeros((h, w), np.uint8),
                    scale=scale, has_label=any(d['has_label'] for d in dish_res),
                    hidden_estimate=sum(d['hidden_estimate'] for d in dish_res),
                    label_mask=dish_res[0]['label_mask'] if dish_res else None)


# ═══════════════════════════ LAZY CACHE ══════════════════════════════════════

class LazyImageCache:
    def __init__(self):
        self._tmpdir = tempfile.mkdtemp(prefix="colony_")
        self._n = 0; self._paths = {}
    def store(self, key, img):
        self._n += 1
        p = os.path.join(self._tmpdir, f"i{self._n}.png")
        cv_imwrite(p, img, [cv2.IMWRITE_PNG_COMPRESSION, 1])
        old = self._paths.pop(key, None)
        if old and os.path.exists(old):
            try: os.remove(old)
            except OSError: pass
        self._paths[key] = p
    def load(self, key):
        p = self._paths.get(key)
        return cv_imread(p) if p and os.path.exists(p) else None
    def remove(self, key):
        p = self._paths.pop(key, None)
        if p and os.path.exists(p):
            try: os.remove(p)
            except OSError: pass
    def cleanup(self):
        for p in self._paths.values():
            try: os.remove(p)
            except OSError: pass
        self._paths.clear()
        try: os.rmdir(self._tmpdir)
        except OSError: pass


# ═══════════════════════════ LEARNING ENGINE ═════════════════════════════════

class LearningEngine:
    def __init__(self):
        d = Path(os.environ.get('APPDATA', str(Path.home()))) / 'ColonyCounter'
        d.mkdir(parents=True, exist_ok=True)
        self._path = d / 'learned_params.json'; self._state = self._load()
    def _load(self):
        try:
            with open(self._path, 'r', encoding='utf-8') as f: return json.load(f)
        except: return {'threshold_ema': None, 'samples': 0}
    def _save(self):
        try:
            with open(self._path, 'w', encoding='utf-8') as f: json.dump(self._state, f, indent=2)
        except: pass
    def update(self, auto_count, excluded, added, cur_threshold):
        if auto_count < C.LEARN_MIN_AUTO: return None
        net = excluded/auto_count - added/auto_count
        if abs(net) < C.LEARN_MIN_RATIO: return None
        delta = max(-C.LEARN_MAX_DELTA, min(C.LEARN_MAX_DELTA, round(net * C.LEARN_DELTA_K)))
        sug = max(5, min(100, cur_threshold + delta))
        ema = self._state.get('threshold_ema')
        ne = sug if ema is None else (C.LEARN_ALPHA*sug + (1-C.LEARN_ALPHA)*ema)
        self._state['threshold_ema'] = ne
        self._state['samples'] = self._state.get('samples', 0) + 1
        self._save()
        r = round(ne)
        return r if r != cur_threshold else None
    @property
    def suggestion(self):
        e = self._state.get('threshold_ema')
        return round(e) if e is not None and self._state.get('samples', 0) >= 2 else None
    @property
    def samples(self): return self._state.get('samples', 0)
    def reset(self): self._state = {'threshold_ema': None, 'samples': 0}; self._save()


# ═══════════════════════════ GUI APPLICATION ═════════════════════════════════

class App:
    def __init__(self, root):
        self.root = root
        self.root.title(f"Colony Counter v{C.VERSION}")
        self.root.geometry("1440x900")
        self.root.minsize(1100, 700)
        self.root.config(bg=T.BG)
        self.root.protocol("WM_DELETE_WINDOW", self._on_close)

        # Load saved theme preference
        T.apply(App._load_theme_pref())

        self.processor = ImageProcessor()
        self.learner = LearningEngine()
        self._cache = LazyImageCache()
        self._lock = threading.Lock()  # protects image_data, manual_marks, excluded_auto
        self.image_paths, self.image_data = [], {}
        self.current_path = None
        self._prev_path = None
        self._pil_orig = None
        self._pil_proc = None
        self._proc_cache_key = None  # (path, marks_hash, excl_hash, ann_hash) for _make_proc cache
        self._proc_cache_img = None
        self.manual_marks, self.excluded_auto = {}, {}
        self.display_names, self.dish_overrides = {}, {}
        self._proc_transform = None
        self._dish_edit, self._dish_drag = False, None
        self._zoom, self._pan = 1.0, [0.0, 0.0]
        self._pan_drag = None
        self._processing = False
        self._compare_pos = 0.5
        self._undo_stack = []
        self._redo_stack = []
        self._annotations = {}
        import atexit
        atexit.register(self._cache.cleanup)

        self.p = dict(
            min_area=tk.IntVar(value=80), max_area=tk.IntVar(value=3000),
            threshold=tk.IntVar(value=25),
            filter_bubbles=tk.BooleanVar(value=True), use_watershed=tk.BooleanVar(value=True),
            show_numbers=tk.BooleanVar(value=True), filter_elongated=tk.BooleanVar(value=True),
            filter_nonconvex=tk.BooleanVar(value=True), auto_learn=tk.BooleanVar(value=True),
            detect_label=tk.BooleanVar(value=True), detect_light_label=tk.BooleanVar(value=False),
            use_otsu=tk.BooleanVar(value=False), use_color_filter=tk.BooleanVar(value=False),
            # Calibration
            dish_diameter_mm=tk.DoubleVar(value=90.0),
            # CFU
            plating_volume_ml=tk.DoubleVar(value=0.1),
            dilution_factor=tk.DoubleVar(value=1.0),
            # Serial dilution group
            dilution_group=tk.StringVar(value=""),
        )
        s = self.learner.suggestion
        if s: self.p['threshold'].set(s)

        self._build_ui()
        self._setup_keys()

    def _on_close(self):
        self._cache.cleanup(); self.root.destroy()

    def _toggle_theme(self):
        """Switch dark <-> light and rebuild all UI widgets."""
        T.toggle()
        self.root.config(bg=T.BG)
        # Save current listbox selection
        sel_idx = None
        sel = self.listbox.curselection()
        if sel: sel_idx = sel[0]
        # Destroy all children and rebuild
        for w in self.root.winfo_children():
            w.destroy()
        self._build_ui()
        # Re-populate listbox
        for path in self.image_paths:
            self.listbox.insert(tk.END, self.display_names.get(path, Path(path).name))
        if sel_idx is not None and sel_idx < len(self.image_paths):
            self.listbox.selection_set(sel_idx)
            self.listbox.see(sel_idx)
        # Refresh panels
        self._refresh_results()
        if self.current_path and self.image_data.get(self.current_path):
            self._refresh_stats(self.image_data[self.current_path])
            self._switch_tab('result')
            self.root.after(80, self._refresh_proc_canvas)
        self._refresh_learn_label()
        # Save preference
        self._save_theme_pref()

    def _save_theme_pref(self):
        cfg_dir = Path(os.environ.get('APPDATA', str(Path.home()))) / 'ColonyCounter'
        cfg_dir.mkdir(parents=True, exist_ok=True)
        try:
            with open(cfg_dir / 'theme.json', 'w') as f:
                json.dump({'theme': T._current}, f)
        except: pass

    @staticmethod
    def _load_theme_pref():
        cfg_dir = Path(os.environ.get('APPDATA', str(Path.home()))) / 'ColonyCounter'
        try:
            with open(cfg_dir / 'theme.json', 'r') as f:
                return json.load(f).get('theme', 'dark')
        except: return 'dark'

    def _setup_keys(self):
        self.root.bind('<Control-o>', lambda e: self._add_images())
        self.root.bind('<Control-s>', lambda e: self._export_excel())
        self.root.bind('<Control-e>', lambda e: self._export_image())
        self.root.bind('<Control-S>', lambda e: self._save_session())
        self.root.bind('<Control-O>', lambda e: self._load_session())
        self.root.bind('<Left>', lambda e: self._navigate(-1))
        self.root.bind('<Right>', lambda e: self._navigate(1))
        self.root.bind('<space>', lambda e: self._process_current())
        self.root.bind('<Delete>', lambda e: self._remove_image())
        self.root.bind('<Control-z>', lambda e: self._undo())
        self.root.bind('<Control-y>', lambda e: self._redo())
        # F1-F4: parameter presets
        self.root.bind('<F1>', lambda e: self._apply_preset('default'))
        self.root.bind('<F2>', lambda e: self._apply_preset('sensitive'))
        self.root.bind('<F3>', lambda e: self._apply_preset('strict'))
        self.root.bind('<F4>', lambda e: self._apply_preset('large'))

    # ── BUILD UI ─────────────────────────────────────────────────────────────

    def _build_ui(self):
        # ── HEADER BAR ───────────────────────────────────────────────────────
        hdr = tk.Frame(self.root, bg=T.BG, height=40)
        hdr.pack(fill=tk.X)
        hdr.pack_propagate(False)

        # Logo
        tk.Label(hdr, text="\u25c9", bg=T.BG, fg=T.ACCENT,
                 font=('Consolas', 16)).pack(side=tk.LEFT, padx=(12, 4))
        tk.Label(hdr, text="ColonyCounter", bg=T.BG, fg=T.ACCENT,
                 font=T.FONT_TITLE).pack(side=tk.LEFT)
        tk.Label(hdr, text=f"v{C.VERSION}", bg=T.BG, fg=T.FG4,
                 font=T.FONT_XS).pack(side=tk.LEFT, padx=(6, 0), pady=(4, 0))

        # Header buttons — right side
        hdr_r = tk.Frame(hdr, bg=T.BG)
        hdr_r.pack(side=tk.RIGHT, padx=12)

        # Theme toggle
        theme_icon = "\u263e" if T.is_dark() else "\u2600"   # ☾ / ☀
        DarkButton(hdr_r, theme_icon, self._toggle_theme, variant='ghost', small=True).pack(side=tk.LEFT, padx=2)

        tk.Frame(hdr_r, bg=T.BORDER, width=1).pack(side=tk.LEFT, fill=tk.Y, padx=4, pady=6)

        for txt, cmd in [("Excel", self._export_excel),
                         ("CSV", self._export_csv),
                         ("PDF", self._export_pdf if HAS_MATPLOTLIB else None),
                         ("Сессия", self._save_session),
                         ("Открыть", self._load_session)]:
            if cmd:
                DarkButton(hdr_r, txt, cmd, variant='ghost', small=True).pack(side=tk.LEFT, padx=2)

        # Separator
        tk.Frame(self.root, bg=T.BORDER, height=1).pack(fill=tk.X)

        # ── PROGRESS BAR ─────────────────────────────────────────────────────
        self._prog_frame = tk.Frame(self.root, bg=T.BG, height=3)
        self._prog_frame.pack(fill=tk.X)
        self._prog_frame.pack_propagate(False)
        self._prog_bar = tk.Frame(self._prog_frame, bg=T.ACCENT, height=3, width=0)
        self._prog_bar.place(x=0, y=0, height=3, relwidth=0)

        # ── MAIN BODY ────────────────────────────────────────────────────────
        body = tk.Frame(self.root, bg=T.BG)
        body.pack(fill=tk.BOTH, expand=True, padx=6, pady=6)

        # Left panel
        left = tk.Frame(body, bg=T.BG1, width=210, highlightthickness=1,
                        highlightbackground=T.BORDER)
        left.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 6))
        left.pack_propagate(False)
        self._build_left(left)

        # Right panel
        right = tk.Frame(body, bg=T.BG1, width=260, highlightthickness=1,
                         highlightbackground=T.BORDER)
        right.pack(side=tk.RIGHT, fill=tk.Y, padx=(6, 0))
        right.pack_propagate(False)
        self._build_right(right)

        # Center panel
        center = tk.Frame(body, bg=T.BG)
        center.pack(fill=tk.BOTH, expand=True)
        self._build_center(center)

        # ── STATUS BAR ───────────────────────────────────────────────────────
        tk.Frame(self.root, bg=T.BORDER, height=1).pack(fill=tk.X)
        sf = tk.Frame(self.root, bg=T.BG, height=24)
        sf.pack(fill=tk.X)
        sf.pack_propagate(False)
        self._status = tk.Label(sf, text="Готово. Ctrl+O — файлы, Ctrl+S — Excel.",
                                bg=T.BG, fg=T.FG4, font=T.FONT_XS, anchor='w')
        self._status.pack(fill=tk.X, padx=12)

    # ── LEFT PANEL ───────────────────────────────────────────────────────────

    def _build_left(self, parent):
        tk.Label(parent, text="ИЗОБРАЖЕНИЯ", bg=T.BG1, fg=T.FG4,
                 font=T.FONT_XS).pack(anchor=tk.W, padx=8, pady=(8, 4))

        # Listbox
        lf = tk.Frame(parent, bg=T.BG)
        lf.pack(fill=tk.BOTH, expand=True, padx=6, pady=(0, 4))
        sb = tk.Scrollbar(lf, troughcolor=T.BG2, bg=T.BG3,
                          activebackground=T.FG4, highlightthickness=0, bd=0)
        sb.pack(side=tk.RIGHT, fill=tk.Y)
        self.listbox = tk.Listbox(lf, yscrollcommand=sb.set,
                                  bg=T.BG, fg=T.FG, font=T.FONT_SM,
                                  selectmode=tk.SINGLE,
                                  selectbackground=T.ACCENT_DIM,
                                  selectforeground=T.ACCENT,
                                  highlightthickness=0, bd=0,
                                  activestyle='none', relief='flat')
        self.listbox.pack(fill=tk.BOTH, expand=True)
        sb.config(command=self.listbox.yview)
        self.listbox.bind('<<ListboxSelect>>', self._on_select)
        self.listbox.bind('<Double-Button-1>', lambda e: self._rename_image())

        # Buttons
        bf = tk.Frame(parent, bg=T.BG1)
        bf.pack(fill=tk.X, padx=6, pady=2)
        DarkButton(bf, "+ Файлы", self._add_images, 'secondary', small=True).pack(side=tk.LEFT, padx=(0,2), fill=tk.X, expand=True)
        DarkButton(bf, "+ Папка", self._add_folder, 'secondary', small=True).pack(side=tk.LEFT, padx=(0,2), fill=tk.X, expand=True)
        DarkButton(bf, "- Удалить", self._remove_image, 'ghost', small=True).pack(side=tk.LEFT, fill=tk.X, expand=True)

        # Separator
        tk.Frame(parent, bg=T.BORDER, height=1).pack(fill=tk.X, padx=6, pady=6)

        # Action buttons
        DarkButton(parent, "ОБРАБОТАТЬ ВСЕ", self._process_all, 'primary').pack(fill=tk.X, padx=6, pady=2)
        DarkButton(parent, "Excel", self._export_excel, 'secondary', small=True).pack(fill=tk.X, padx=6, pady=1)
        if HAS_MATPLOTLIB:
            DarkButton(parent, "PDF", self._export_pdf, 'secondary', small=True).pack(fill=tk.X, padx=6, pady=1)

        tk.Frame(parent, bg=T.BORDER, height=1).pack(fill=tk.X, padx=6, pady=6)

        # Results
        tk.Label(parent, text="РЕЗУЛЬТАТЫ", bg=T.BG1, fg=T.FG4,
                 font=T.FONT_XS).pack(anchor=tk.W, padx=8, pady=(0, 2))
        self._res_text = tk.Text(parent, bg=T.BG, fg=T.FG3, font=T.FONT_SM,
                                 height=10, state=tk.DISABLED,
                                 highlightthickness=0, bd=0, relief='flat',
                                 insertbackground=T.FG)
        self._res_text.pack(fill=tk.BOTH, expand=True, padx=6, pady=(0, 6))

    # ── CENTER PANEL ─────────────────────────────────────────────────────────

    def _build_center(self, parent):
        # Tab bar
        tab_bar = tk.Frame(parent, bg=T.BG)
        tab_bar.pack(fill=tk.X, pady=(0, 4))

        self._tabs = {}
        self._tab_var = tk.StringVar(value='result')
        for tid, label in [('original', 'Оригинал'), ('result', 'Результат'), ('compare', 'Сравнение')]:
            b = tk.Label(tab_bar, text=label, bg=T.BG, fg=T.FG4,
                         font=T.FONT_SM, padx=12, pady=4, cursor='hand2')
            b.pack(side=tk.LEFT)
            b.bind('<Button-1>', lambda e, t=tid: self._switch_tab(t))
            self._tabs[tid] = b

        # Toolbar — right of tabs
        tb = tk.Frame(tab_bar, bg=T.BG)
        tb.pack(side=tk.RIGHT)
        for txt, cmd in [("Зум x1", self._reset_zoom),
                         ("Отменить", self._undo_manual),
                         ("Очистить", self._clear_manual),
                         ("Восстан.", self._restore_auto)]:
            DarkButton(tb, txt, cmd, 'ghost', small=True).pack(side=tk.LEFT, padx=1)

        tk.Frame(tb, bg=T.BORDER, width=1).pack(side=tk.LEFT, fill=tk.Y, padx=4, pady=2)
        self._dish_btn = DarkButton(tb, "Границы", self._toggle_dish_edit, 'ghost', small=True)
        self._dish_btn.pack(side=tk.LEFT, padx=1)
        DarkButton(tb, "Сброс", self._reset_dish_overrides, 'ghost', small=True).pack(side=tk.LEFT, padx=1)

        # Canvas container
        self._canvas_frame = tk.Frame(parent, bg=T.CANVAS_BG, highlightthickness=1,
                                      highlightbackground=T.BORDER)
        self._canvas_frame.pack(fill=tk.BOTH, expand=True)

        self.canvas_orig = tk.Canvas(self._canvas_frame, bg=T.CANVAS_BG,
                                     highlightthickness=0, cursor='crosshair')
        self.canvas_proc = tk.Canvas(self._canvas_frame, bg=T.CANVAS_BG,
                                     highlightthickness=0, cursor='tcross')
        self.canvas_compare = tk.Canvas(self._canvas_frame, bg=T.CANVAS_BG,
                                        highlightthickness=0, cursor='sb_h_double_arrow')

        # Show default tab
        self._switch_tab('result')

        # Bindings
        self.canvas_orig.bind("<Configure>", lambda e: self._redraw(self.canvas_orig, self._pil_orig))
        self.canvas_proc.bind("<Configure>", lambda e: self._refresh_proc_canvas())
        self.canvas_proc.bind("<Button-1>", self._on_lmb_down)
        self.canvas_proc.bind("<B1-Motion>", self._on_lmb_motion)
        self.canvas_proc.bind("<ButtonRelease-1>", self._on_lmb_up)
        self.canvas_proc.bind("<Button-3>", self._on_proc_rclick)
        self.canvas_proc.bind("<MouseWheel>", self._on_proc_wheel)
        self.canvas_proc.bind("<ButtonPress-2>", self._on_proc_pan_start)
        self.canvas_proc.bind("<B2-Motion>", self._on_proc_pan_move)
        self.canvas_proc.bind("<ButtonRelease-2>", self._on_proc_pan_end)
        self.canvas_compare.bind("<Button-1>", self._on_compare_click)
        self.canvas_compare.bind("<B1-Motion>", self._on_compare_drag)
        self.canvas_compare.bind("<Configure>", lambda e: self._draw_comparison())

        # Hint
        tk.Label(parent, text="ЛКМ-метка  ПКМ-убрать  Колесо-зум  СКМ-пан  Стрелки-навигация  Space-обработка",
                 bg=T.BG, fg=T.FG4, font=T.FONT_XS).pack(anchor=tk.W, pady=(4, 0))

    def _switch_tab(self, tid):
        self._tab_var.set(tid)
        for t, w in self._tabs.items():
            if t == tid:
                w.config(bg=T.ACCENT_DIM, fg=T.ACCENT)
            else:
                w.config(bg=T.BG, fg=T.FG4)
        # Show correct canvas
        for c in (self.canvas_orig, self.canvas_proc, self.canvas_compare):
            c.pack_forget()
        if tid == 'original':
            self.canvas_orig.pack(fill=tk.BOTH, expand=True)
            if self._pil_orig:
                self.root.after(30, lambda: self._blit(self._pil_orig, self.canvas_orig))
        elif tid == 'result':
            self.canvas_proc.pack(fill=tk.BOTH, expand=True)
            self.root.after(30, self._refresh_proc_canvas)
        else:
            self.canvas_compare.pack(fill=tk.BOTH, expand=True)
            self.root.after(30, self._draw_comparison)

    # ── RIGHT PANEL ──────────────────────────────────────────────────────────

    def _build_right(self, parent):
        # Scrollable
        canvas = tk.Canvas(parent, bg=T.BG1, highlightthickness=0, bd=0)
        vsb = tk.Scrollbar(parent, command=canvas.yview,
                           troughcolor=T.BG2, bg=T.BG3, highlightthickness=0, bd=0)
        inner = tk.Frame(canvas, bg=T.BG1)
        inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=inner, anchor=tk.NW, width=245)
        canvas.configure(yscrollcommand=vsb.set)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)

        def _wheel(e): canvas.yview_scroll(int(-1*(e.delta/120)), "units")
        canvas.bind("<Enter>", lambda e: canvas.bind_all("<MouseWheel>", _wheel))
        canvas.bind("<Leave>", lambda e: canvas.unbind_all("<MouseWheel>"))

        self._build_params(inner)

    def _build_params(self, parent):
        # ── Detection ────────────────────────────────────────────────────────
        s1 = DarkSection(parent, "Обнаружение")
        s1.pack(fill=tk.X, padx=6, pady=(8, 4))
        DarkSlider(s1.body, "Мин. площадь (пикс.)", self.p['min_area'], 10, 500, 10).pack(fill=tk.X, pady=2)
        DarkSlider(s1.body, "Макс. площадь (пикс.)", self.p['max_area'], 100, 15000, 50).pack(fill=tk.X, pady=2)
        DarkSlider(s1.body, "Порог чувствительности", self.p['threshold'], 5, 100, 1).pack(fill=tk.X, pady=2)

        # ── Filters ──────────────────────────────────────────────────────────
        s2 = DarkSection(parent, "Фильтры")
        s2.pack(fill=tk.X, padx=6, pady=4)
        for txt, var in [("Пузыри воздуха", self.p['filter_bubbles']),
                         ("Watershed кластеры", self.p['use_watershed']),
                         ("Числа на картинке", self.p['show_numbers']),
                         ("Вытянутые объекты", self.p['filter_elongated']),
                         ("Невыпуклые объекты", self.p['filter_nonconvex']),
                         ("Маскировать этикетку", self.p['detect_label'])]:
            DarkCheck(s2.body, txt, var).pack(fill=tk.X, pady=1)

        # ── v2.0 features ───────────────────────────────────────────────────
        s3 = DarkSection(parent, "v2.0")
        s3.pack(fill=tk.X, padx=6, pady=4)
        DarkCheck(s3.body, "Светлые этикетки", self.p['detect_light_label']).pack(fill=tk.X, pady=1)
        DarkCheck(s3.body, "Otsu per-image", self.p['use_otsu']).pack(fill=tk.X, pady=1)
        DarkCheck(s3.body, "Цветовой фильтр HSV", self.p['use_color_filter']).pack(fill=tk.X, pady=1)

        # ── Learning ─────────────────────────────────────────────────────────
        s4 = DarkSection(parent, "Авто-обучение")
        s4.pack(fill=tk.X, padx=6, pady=4)
        DarkCheck(s4.body, "Обучаться по правкам", self.p['auto_learn']).pack(fill=tk.X, pady=1)
        self._learn_lbl = tk.Label(s4.body, text="", bg=T.BG1, fg=T.FG4, font=T.FONT_XS, anchor='w')
        self._learn_lbl.pack(fill=tk.X, pady=(2, 0))
        lr = tk.Frame(s4.body, bg=T.BG1)
        lr.pack(fill=tk.X, pady=(4, 0))
        DarkButton(lr, "Применить", self._apply_learned, 'secondary', small=True).pack(side=tk.LEFT, expand=True, fill=tk.X, padx=(0,2))
        DarkButton(lr, "Сброс", self._reset_learning, 'danger', small=True).pack(side=tk.LEFT, expand=True, fill=tk.X)
        self._refresh_learn_label()

        # ── Calibration ──────────────────────────────────────────────────────
        sc = DarkSection(parent, "Калибровка")
        sc.pack(fill=tk.X, padx=6, pady=4)
        DarkSlider(sc.body, "Диаметр чашки (мм)", self.p['dish_diameter_mm'], 30, 150, 1).pack(fill=tk.X, pady=2)

        # ── CFU ──────────────────────────────────────────────────────────────
        sf = DarkSection(parent, "CFU/мл")
        sf.pack(fill=tk.X, padx=6, pady=4)
        DarkSlider(sf.body, "Объём посева (мл)", self.p['plating_volume_ml'], 0.01, 1.0, 0.01).pack(fill=tk.X, pady=2)
        DarkSlider(sf.body, "Разведение (1:N)", self.p['dilution_factor'], 1, 1000000, 1).pack(fill=tk.X, pady=2)

        # ── Dilution group ───────────────────────────────────────────────────
        sdg = DarkSection(parent, "Серия разведений")
        sdg.pack(fill=tk.X, padx=6, pady=4)
        gr = tk.Frame(sdg.body, bg=T.BG1)
        gr.pack(fill=tk.X)
        tk.Label(gr, text="Группа:", bg=T.BG1, fg=T.FG3, font=T.FONT_XS).pack(side=tk.LEFT)
        tk.Entry(gr, textvariable=self.p['dilution_group'], bg=T.BG2, fg=T.FG,
                 font=T.FONT_SM, insertbackground=T.FG, highlightthickness=1,
                 highlightbackground=T.BORDER, bd=0, width=12).pack(side=tk.LEFT, padx=4)
        DarkButton(sdg.body, "Кривая разведений", self._show_dilution_curve, 'ghost', small=True).pack(fill=tk.X, pady=2)

        # ── Presets ──────────────────────────────────────────────────────────
        sp = DarkSection(parent, "Пресеты (F1-F4)")
        sp.pack(fill=tk.X, padx=6, pady=4)
        pr = tk.Frame(sp.body, bg=T.BG1)
        pr.pack(fill=tk.X)
        for key, label in [('default', 'F1 Стд'), ('sensitive', 'F2 Чувст'),
                           ('strict', 'F3 Строг'), ('large', 'F4 Крупн')]:
            DarkButton(pr, label, lambda k=key: self._apply_preset(k), 'ghost', small=True).pack(
                side=tk.LEFT, expand=True, fill=tk.X, padx=1)

        # ── Actions ──────────────────────────────────────────────────────────
        af = tk.Frame(parent, bg=T.BG1)
        af.pack(fill=tk.X, padx=6, pady=8)
        DarkButton(af, "ОБРАБОТАТЬ ТЕКУЩЕЕ", self._process_current, 'primary').pack(fill=tk.X, pady=2)
        DarkButton(af, "Авто-порог (Otsu)", self._auto_threshold, 'secondary', small=True).pack(fill=tk.X, pady=1)
        ar = tk.Frame(af, bg=T.BG1)
        ar.pack(fill=tk.X, pady=1)
        if HAS_MATPLOTLIB:
            DarkButton(ar, "Статистика", self._show_statistics, 'ghost', small=True).pack(side=tk.LEFT, expand=True, fill=tk.X, padx=(0,1))
            DarkButton(ar, "Heatmap", self._show_heatmap, 'ghost', small=True).pack(side=tk.LEFT, expand=True, fill=tk.X, padx=(1,1))
            DarkButton(ar, "Бок-о-бок", self._show_side_by_side, 'ghost', small=True).pack(side=tk.LEFT, expand=True, fill=tk.X, padx=(1,0))
        DarkButton(af, "Аннотация", self._add_annotation, 'ghost', small=True).pack(fill=tk.X, pady=1)
        if HAS_MATPLOTLIB:
            DarkButton(af, "Воспроизвод.", self._show_reproducibility, 'ghost', small=True).pack(fill=tk.X, pady=1)

        # ── Stats ────────────────────────────────────────────────────────────
        s5 = DarkSection(parent, "Статистика")
        s5.pack(fill=tk.X, padx=6, pady=4)
        self._stats_text = tk.Text(s5.body, bg=T.BG, fg=T.FG3, font=T.FONT_SM,
                                   height=9, state=tk.DISABLED,
                                   highlightthickness=0, bd=0, relief='flat',
                                   insertbackground=T.FG)
        self._stats_text.pack(fill=tk.X)

        # ── Legend ───────────────────────────────────────────────────────────
        s6 = DarkSection(parent, "Легенда")
        s6.pack(fill=tk.X, padx=6, pady=(4, 8))
        for color, label in [(T.ACCENT, "одиночная колония"),
                             ('#3b82f6', "кластер"),
                             (T.YELLOW, "ручная метка"),
                             (T.RED, "исключённая")]:
            r = tk.Frame(s6.body, bg=T.BG1)
            r.pack(fill=tk.X, pady=1)
            tk.Canvas(r, width=10, height=10, bg=color, highlightthickness=0).pack(side=tk.LEFT, padx=(0, 6))
            tk.Label(r, text=label, bg=T.BG1, fg=T.FG4, font=T.FONT_XS).pack(side=tk.LEFT)

    # ═══════════════════════ FILE OPS ════════════════════════════════════════

    def _add_images(self):
        paths = filedialog.askopenfilenames(
            title="Изображения",
            filetypes=[("Изображения", "*.jpg *.jpeg *.png *.bmp *.tiff *.tif"), ("Все", "*.*")])
        self._add_paths(list(paths))

    def _add_folder(self):
        folder = filedialog.askdirectory(title="Папка с изображениями")
        if not folder: return
        exts = {'.jpg', '.jpeg', '.png', '.bmp', '.tiff', '.tif'}
        paths = [str(e) for e in sorted(Path(folder).iterdir()) if e.is_file() and e.suffix.lower() in exts]
        if not paths: messagebox.showinfo("Пусто", "Нет изображений."); return
        self._add_paths(paths)

    def _add_paths(self, paths):
        added = 0
        for path in paths:
            try:
                with Image.open(path) as pil:
                    nf = getattr(pil, 'n_frames', 1)
            except: nf = 1
            if nf > 1 and path.lower().endswith(('.tiff', '.tif')):
                for fi in range(nf):
                    vp = f"{path}::frame{fi}"
                    if vp not in self.image_paths:
                        self.image_paths.append(vp); self.image_data[vp] = None
                        nm = f"{Path(path).name} [#{fi+1}]"
                        self.display_names[vp] = nm; self.listbox.insert(tk.END, nm); added += 1
            else:
                if path not in self.image_paths:
                    self.image_paths.append(path); self.image_data[path] = None
                    self.display_names[path] = Path(path).name
                    self.listbox.insert(tk.END, Path(path).name); added += 1
        if added: self._set_status(f"Добавлено {added} изображений")

    def _load_tiff_frame(self, vp):
        if '::frame' not in vp: return None
        rp, fi = vp.rsplit('::frame', 1)
        try:
            with Image.open(rp) as p:
                p.seek(int(fi))
                return cv2.cvtColor(np.array(p.convert('RGB')), cv2.COLOR_RGB2BGR)
        except: return None

    def _process_image(self, path, params, progress_cb=None):
        """Unified processing: handles TIFF frames, temp files, and cleanup."""
        if '::frame' in path:
            fi = self._load_tiff_frame(path)
            if fi is None:
                raise ValueError(f"Не удалось загрузить кадр: {path}")
            tmp = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
            try:
                cv_imwrite(tmp.name, fi)
                return self.processor.process(tmp.name, params, progress_cb)
            finally:
                try: os.unlink(tmp.name)
                except OSError: pass
        else:
            return self.processor.process(path, params, progress_cb)

    def _remove_image(self):
        sel = self.listbox.curselection()
        if not sel: return
        idx = sel[0]; path = self.image_paths.pop(idx)
        self.image_data.pop(path, None); self.display_names.pop(path, None)
        self._cache.remove(f"{path}_a"); self._cache.remove(f"{path}_c")
        self.listbox.delete(idx)
        if self.current_path == path:
            self.current_path = None; self._pil_orig = None; self._pil_proc = None
            self.canvas_orig.delete("all"); self.canvas_proc.delete("all")
        self._refresh_results()

    def _rename_image(self):
        sel = self.listbox.curselection()
        if not sel: return
        idx = sel[0]; path = self.image_paths[idx]
        nm = simpledialog.askstring("Переименовать", "Имя:", initialvalue=self.display_names.get(path, ''), parent=self.root)
        if not nm or not nm.strip(): return
        self.display_names[path] = nm.strip()
        self.listbox.delete(idx); self.listbox.insert(idx, nm.strip()); self.listbox.selection_set(idx)
        self._refresh_results()

    def _navigate(self, d):
        if not self.image_paths: return
        sel = self.listbox.curselection()
        idx = sel[0] if sel else 0
        ni = max(0, min(len(self.image_paths)-1, idx+d))
        self.listbox.selection_clear(0, tk.END); self.listbox.selection_set(ni)
        self.listbox.see(ni); self._on_select()

    # ═══════════════════════ SELECTION ════════════════════════════════════════

    def _on_select(self, _e=None):
        sel = self.listbox.curselection()
        if not sel: return
        path = self.image_paths[sel[0]]
        if self._prev_path and self._prev_path != path: self._do_learning(self._prev_path)
        self._prev_path = path; self.current_path = path
        self._zoom, self._pan = 1.0, [0.0, 0.0]
        result = self.image_data.get(path)
        if result:
            self._refresh_stats(result)
            self._switch_tab('result')
            self.root.after(60, self._refresh_proc_canvas)
        else:
            self._pil_proc = None; self.canvas_proc.delete("all")
            self._switch_tab('original')
        self.root.after(60, lambda p=path: self._show_file(p, self.canvas_orig))

    # ═══════════════════════ PROCESSING ═══════════════════════════════════════

    def _get_params(self): return {k: v.get() for k, v in self.p.items()}

    def _process_current(self):
        if not self.current_path: return
        if self._processing: return
        self._run_threaded([self.current_path])

    def _process_all(self):
        if not self.image_paths or self._processing: return
        self._run_threaded(list(self.image_paths))

    def _run_threaded(self, paths):
        self._processing = True
        def worker():
            n = len(paths)
            for i, path in enumerate(paths):
                try:
                    params = self._get_params()
                    ov = self.dish_overrides.get(path)
                    if ov: params['dish_overrides'] = ov
                    _path = path  # capture for closure
                    def pcb(frac, msg, _i=i, _n=n, _p=_path):
                        overall = (_i+frac)/_n
                        self.root.after(0, lambda v=overall, m=msg, pp=_p: (
                            self._prog_bar.place_configure(relwidth=v),
                            self._set_status(f"[{_i+1}/{_n}] {Path(pp.split('::')[0]).name}: {m}")))
                    result = self._process_image(path, params, pcb)
                    self._cache.store(f"{path}_a", result['annotated'])
                    self._cache.store(f"{path}_c", result['img_clean'])
                    rl = {k: v for k, v in result.items() if k not in ('annotated', 'img_clean')}
                    rl['_cached'] = True
                    with self._lock:
                        self.root.after(0, lambda p=path, r=rl: self._on_done(p, r))
                except Exception as exc:
                    self.root.after(0, lambda e=exc: self._set_status(f"Ошибка: {e}"))
            self.root.after(0, self._on_all_done)
        threading.Thread(target=worker, daemon=True).start()

    def _on_done(self, path, result):
        with self._lock:
            self.image_data[path] = result
        if self.current_path == path:
            self._refresh_stats(result)
            self._switch_tab('result')
            self.root.after(60, self._refresh_proc_canvas)

    def _on_all_done(self):
        self._processing = False
        self._prog_bar.place_configure(relwidth=1.0)
        self.root.after(1500, lambda: self._prog_bar.place_configure(relwidth=0))
        self._refresh_results()
        n = sum(1 for r in self.image_data.values() if r)
        self._set_status(f"Готово. Обработано: {n}")

    def _run_sync(self, path, silent=False):
        try:
            params = self._get_params()
            ov = self.dish_overrides.get(path)
            if ov: params['dish_overrides'] = ov
            result = self._process_image(path, params)
            self._cache.store(f"{path}_a", result['annotated'])
            self._cache.store(f"{path}_c", result['img_clean'])
            rl = {k: v for k, v in result.items() if k not in ('annotated', 'img_clean')}
            rl['_cached'] = True
            with self._lock:
                self.image_data[path] = rl
            if self.current_path == path:
                self._refresh_stats(rl); self.root.after(60, self._refresh_proc_canvas)
        except Exception as exc:
            if not silent: messagebox.showerror("Ошибка", str(exc))

    def _auto_threshold(self):
        if not self.current_path: return
        try:
            path = self.current_path
            if '::frame' in path:
                img = self._load_tiff_frame(path)
            else:
                img = cv_imread(path)
            if img is None: return
            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
            cx, cy, r = self.processor.detect_dish(gray)
            h, w = gray.shape; mask = np.zeros((h, w), np.uint8)
            cv2.circle(mask, (cx, cy), int(r*C.DISH_MASK_RATIO), 255, -1)
            norm = self.processor.normalize_background(gray, mask)
            enh = cv2.createCLAHE(clipLimit=C.CLAHE_CLIP, tileGridSize=C.CLAHE_TILE).apply(norm)
            ov, _ = cv2.threshold(enh, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
            s = max(5, int(ov * 0.48)); self.p['threshold'].set(s)
            self._set_status(f"Otsu: {s} (raw={int(ov)})")
        except Exception as e: messagebox.showerror("Ошибка", str(e))

    # ═══════════════════════ DISPLAY ═════════════════════════════════════════

    def _set_status(self, text): self._status.config(text=text)

    def _show_file(self, path, canvas):
        try:
            if '::frame' in path:
                ci = self._load_tiff_frame(path)
                if ci is None: return
                pil = Image.fromarray(cv2.cvtColor(ci, cv2.COLOR_BGR2RGB))
            else: pil = Image.open(path)
            if canvas is self.canvas_orig: self._pil_orig = pil
            else: self._pil_proc = pil
            self._blit(pil, canvas)
        except Exception as e: self._set_status(f"Ошибка: {e}")

    def _redraw(self, canvas, pil):
        if pil: self._blit(pil, canvas)

    def _blit(self, pil, canvas):
        canvas.update_idletasks()
        cw, ch = canvas.winfo_width(), canvas.winfo_height()
        if cw <= 1: cw = 800
        if ch <= 1: ch = 600
        iw, ih = pil.size; fit = min(cw/iw, ch/ih)
        if canvas is self.canvas_proc:
            ds = fit * self._zoom; px, py = self._pan
        else: ds, px, py = fit, 0.0, 0.0
        ox = cw/2 + px - iw*ds/2; oy = ch/2 + py - ih*ds/2
        if canvas is self.canvas_proc: self._proc_transform = (ds, ox, oy, iw, ih)
        sx0, sy0 = max(0, int(-ox/ds)), max(0, int(-oy/ds))
        sx1, sy1 = min(iw, int((cw-ox)/ds)+1), min(ih, int((ch-oy)/ds)+1)
        if sx1 <= sx0 or sy1 <= sy0: canvas.delete("all"); return
        crop = pil.crop((sx0, sy0, sx1, sy1))
        ow, oh = max(1, round((sx1-sx0)*ds)), max(1, round((sy1-sy0)*ds))
        resized = crop.resize((ow, oh), Image.BILINEAR)
        photo = ImageTk.PhotoImage(resized)
        canvas.delete("all")
        canvas.create_image(round(ox+sx0*ds), round(oy+sy0*ds), anchor=tk.NW, image=photo)
        canvas.image = photo

    # ── Comparison ───────────────────────────────────────────────────────────

    def _on_compare_click(self, e):
        cw = self.canvas_compare.winfo_width()
        if cw > 0: self._compare_pos = e.x / cw
        self._draw_comparison()

    def _on_compare_drag(self, e):
        cw = self.canvas_compare.winfo_width()
        if cw > 0: self._compare_pos = max(0, min(1, e.x/cw))
        self._draw_comparison()

    def _draw_comparison(self):
        if not self.current_path or not self._pil_orig: return
        result = self.image_data.get(self.current_path)
        if not result: return
        canvas = self.canvas_compare; canvas.update_idletasks()
        cw, ch = canvas.winfo_width(), canvas.winfo_height()
        if cw <= 1 or ch <= 1: return
        ann = self._get_annotated(self.current_path)
        if ann is None: return
        pp = Image.fromarray(cv2.cvtColor(ann, cv2.COLOR_BGR2RGB))
        iw, ih = self._pil_orig.size; fit = min(cw/iw, ch/ih)
        nw, nh = int(iw*fit), int(ih*fit)
        orig_r = self._pil_orig.resize((nw, nh), Image.LANCZOS)
        proc_r = pp.resize((nw, nh), Image.LANCZOS)
        comp = orig_r.copy()
        sx = int(nw * self._compare_pos)
        if sx < nw: comp.paste(proc_r.crop((sx, 0, nw, nh)), (sx, 0))
        ox, oy = (cw-nw)//2, (ch-nh)//2
        photo = ImageTk.PhotoImage(comp)
        canvas.delete("all")
        canvas.create_image(ox, oy, anchor=tk.NW, image=photo)
        lx = ox + sx
        canvas.create_line(lx, oy, lx, oy+nh, fill=T.YELLOW, width=2, dash=(6, 3))
        canvas.create_text(ox+8, oy+8, text="Оригинал", anchor=tk.NW, fill=T.YELLOW, font=T.FONT_SM)
        canvas.create_text(ox+nw-8, oy+8, text="Результат", anchor=tk.NE, fill=T.YELLOW, font=T.FONT_SM)
        canvas.image = photo

    # ═══════════════════════ MANUAL MODE ═════════════════════════════════════

    def _c2i(self, cx, cy):
        if not self._proc_transform: return None
        s, ox, oy, iw, ih = self._proc_transform
        ix, iy = int((cx-ox)/s), int((cy-oy)/s)
        return (ix, iy) if 0 <= ix < iw and 0 <= iy < ih else None

    def _ensure_transform(self):
        if not self._pil_proc: return
        cw, ch = self.canvas_proc.winfo_width(), self.canvas_proc.winfo_height()
        if cw <= 1 or ch <= 1: return
        iw, ih = self._pil_proc.size; fit = min(cw/iw, ch/ih)
        ds = fit * self._zoom; px, py = self._pan
        self._proc_transform = (ds, cw/2+px-iw*ds/2, ch/2+py-ih*ds/2, iw, ih)

    def _on_lmb_down(self, e):
        if self._dish_edit: self.canvas_proc.config(cursor='fleur'); self._dish_drag_start(e)
        else: self._on_proc_click(e)

    def _on_lmb_motion(self, e):
        if self._dish_edit and self._dish_drag: self._dish_drag_move(e)

    def _on_lmb_up(self, e):
        if self._dish_edit and self._dish_drag:
            self._dish_drag_end(e)
            self.canvas_proc.config(cursor='hand2' if self._dish_edit else 'tcross')

    def _push_undo(self):
        """Snapshot current marks/excl state for undo."""
        p = self.current_path
        if not p: return
        self._undo_stack.append((p,
            [tuple(m) for m in self.manual_marks.get(p, [])],
            [tuple(e) for e in self.excluded_auto.get(p, [])]))
        self._redo_stack.clear()
        if len(self._undo_stack) > 100: self._undo_stack.pop(0)

    def _undo(self):
        if not self._undo_stack: return
        p, marks, excl = self._undo_stack.pop()
        # Save current state to redo
        self._redo_stack.append((p,
            [tuple(m) for m in self.manual_marks.get(p, [])],
            [tuple(e) for e in self.excluded_auto.get(p, [])]))
        self.manual_marks[p] = list(marks)
        self.excluded_auto[p] = list(excl)
        self._refresh_proc_canvas(); self._refresh_stats_cur(); self._refresh_results()

    def _redo(self):
        if not self._redo_stack: return
        p, marks, excl = self._redo_stack.pop()
        self._undo_stack.append((p,
            [tuple(m) for m in self.manual_marks.get(p, [])],
            [tuple(e) for e in self.excluded_auto.get(p, [])]))
        self.manual_marks[p] = list(marks)
        self.excluded_auto[p] = list(excl)
        self._refresh_proc_canvas(); self._refresh_stats_cur(); self._refresh_results()

    def _on_proc_click(self, e):
        if not self.current_path: return
        self._ensure_transform()
        pt = self._c2i(e.x, e.y)
        if pt is None: return
        self._push_undo()
        self.manual_marks.setdefault(self.current_path, []).append(pt)
        self._refresh_proc_canvas(); self._refresh_stats_cur(); self._refresh_results()

    def _on_proc_rclick(self, e):
        if not self.current_path or self._dish_edit: return
        self._ensure_transform()
        pt = self._c2i(e.x, e.y)
        if pt is None: return
        path = self.current_path; result = self.image_data.get(path)
        cands = []
        for i, m in enumerate(self.manual_marks.get(path, [])): cands.append(('m', i, m))
        if result:
            excl = set(self.excluded_auto.get(path, []))
            for col in result['colonies']:
                for cx_c, cy_c in col['ws_centers']:
                    if (cx_c, cy_c) not in excl: cands.append(('a', (cx_c, cy_c), (cx_c, cy_c)))
        if not cands: return
        dists = [(abs(c[2][0]-pt[0])+abs(c[2][1]-pt[1]), i) for i, c in enumerate(cands)]
        _, ni = min(dists); kind, ref, _ = cands[ni]
        self._push_undo()
        if kind == 'm': self.manual_marks[path].pop(ref)
        else: self.excluded_auto.setdefault(path, []).append(ref)
        self._refresh_proc_canvas(); self._refresh_stats_cur(); self._refresh_results()

    def _undo_manual(self):
        self._undo()  # Now uses full undo system

    def _clear_manual(self):
        if not self.current_path: return
        if self.manual_marks.get(self.current_path):
            self._push_undo()
            self.manual_marks[self.current_path] = []
            self._refresh_proc_canvas(); self._refresh_stats_cur(); self._refresh_results()

    def _restore_auto(self):
        if not self.current_path: return
        if self.excluded_auto.get(self.current_path):
            self._push_undo()
            self.excluded_auto[self.current_path] = []
            self._refresh_proc_canvas(); self._refresh_stats_cur(); self._refresh_results()

    # ═══════════════════════ LEARNING ════════════════════════════════════════

    def _do_learning(self, path):
        if not self.p['auto_learn'].get(): return
        result = self.image_data.get(path)
        if not result: return
        excl, added = len(self.excluded_auto.get(path, [])), len(self.manual_marks.get(path, []))
        if excl == 0 and added == 0: return
        nt = self.learner.update(result['total'], excl, added, self.p['threshold'].get())
        if nt is not None:
            self.p['threshold'].set(nt); self._set_status(f"Обучение: порог -> {nt}")
            self._refresh_learn_label()

    def _apply_learned(self, silent=False):
        s = self.learner.suggestion
        if s is not None:
            self.p['threshold'].set(s)
            if not silent: self._set_status(f"Порог = {s}")
        elif not silent:
            self._set_status("Нет данных обучения")

    def _refresh_learn_label(self):
        s, n = self.learner.suggestion, self.learner.samples
        self._learn_lbl.config(text=f"Модель: {n} сессий | ~{s}" if s else "Нет данных")

    def _reset_learning(self):
        self.learner.reset(); self._refresh_learn_label(); self._set_status("Модель сброшена")

    # ═══════════════════════ ZOOM / PAN ══════════════════════════════════════

    def _on_proc_wheel(self, e):
        if not self._pil_proc: return "break"
        f = C.ZOOM_FACTOR if e.delta > 0 else 1/C.ZOOM_FACTOR
        oz = self._zoom; nz = max(C.ZOOM_MIN, min(C.ZOOM_MAX, oz*f))
        if nz == oz: return "break"
        cw, ch = self.canvas_proc.winfo_width(), self.canvas_proc.winfo_height()
        iw, ih = self._pil_proc.size; fit = min(cw/iw, ch/ih)
        oox = cw/2+self._pan[0]-iw*fit*oz/2; ooy = ch/2+self._pan[1]-ih*fit*oz/2
        ix, iy = (e.x-oox)/(fit*oz), (e.y-ooy)/(fit*oz)
        nox, noy = e.x-ix*fit*nz, e.y-iy*fit*nz
        self._pan[0] = nox-(cw/2-iw*fit*nz/2); self._pan[1] = noy-(ch/2-ih*fit*nz/2)
        self._zoom = nz; self._refresh_proc_canvas(); return "break"

    def _on_proc_pan_start(self, e): self._pan_drag = (e.x, e.y, self._pan[0], self._pan[1])
    def _on_proc_pan_move(self, e):
        if not self._pan_drag: return
        sx, sy, px, py = self._pan_drag
        self._pan[0] = px+(e.x-sx); self._pan[1] = py+(e.y-sy); self._refresh_proc_canvas()
    def _on_proc_pan_end(self, _e=None): self._pan_drag = None
    def _reset_zoom(self): self._zoom = 1.0; self._pan = [0.0, 0.0]; self._refresh_proc_canvas()

    # ═══════════════════════ IMAGE HELPERS ════════════════════════════════════

    def _get_annotated(self, path):
        r = self.image_data.get(path)
        if not r: return None
        return self._cache.load(f"{path}_a") if r.get('_cached') else r.get('annotated')

    def _get_clean(self, path):
        r = self.image_data.get(path)
        if not r: return None
        return self._cache.load(f"{path}_c") if r.get('_cached') else r.get('img_clean')

    def _refresh_proc_canvas(self):
        if not self.current_path: return
        cv_img = self._make_proc(self.current_path)
        if cv_img is not None:
            self._pil_proc = Image.fromarray(cv2.cvtColor(cv_img, cv2.COLOR_BGR2RGB))
            self._blit(self._pil_proc, self.canvas_proc)
        elif self._pil_proc: self._blit(self._pil_proc, self.canvas_proc)
        if self._dish_edit: self._draw_dish_handles()

    # ═══════════════════════ DISH EDIT ═══════════════════════════════════════

    def _toggle_dish_edit(self):
        self._dish_edit = not self._dish_edit
        self._dish_btn.set_text("Границы [ON]" if self._dish_edit else "Границы")
        self.canvas_proc.config(cursor='hand2' if self._dish_edit else 'tcross')
        self._refresh_proc_canvas()

    def _reset_dish_overrides(self):
        if not self.current_path: return
        if self.current_path in self.dish_overrides:
            del self.dish_overrides[self.current_path]; self._run_sync(self.current_path); self._refresh_results()

    def _get_dishes(self):
        if not self.current_path: return []
        ov = self.dish_overrides.get(self.current_path)
        if ov: return [tuple(d) for d in ov]
        r = self.image_data.get(self.current_path)
        return r.get('dishes', [r['dish']]) if r else []

    def _i2c(self, ix, iy):
        if not self._proc_transform: return None
        s, ox, oy, _, _ = self._proc_transform
        return (ox+ix*s, oy+iy*s)

    def _draw_dish_handles(self):
        self.canvas_proc.delete('dh'); self._ensure_transform()
        if not self._proc_transform: return
        s = self._proc_transform[0]
        for di, (dcx, dcy, dr) in enumerate(self._get_dishes()):
            cc = self._i2c(dcx, dcy)
            if not cc: continue
            ccx, ccy = cc; cr = dr*s
            self.canvas_proc.create_oval(ccx-cr, ccy-cr, ccx+cr, ccy+cr,
                                         outline=T.YELLOW, width=2, dash=(8, 4), tags='dh')
            hw = 7
            self.canvas_proc.create_rectangle(ccx-hw, ccy-hw, ccx+hw, ccy+hw,
                                              fill=T.YELLOW, outline=T.BG, tags='dh')
            for ex, ey in [(dcx, dcy-dr), (dcx+dr, dcy), (dcx, dcy+dr), (dcx-dr, dcy)]:
                ec = self._i2c(ex, ey)
                if ec: self.canvas_proc.create_rectangle(ec[0]-5, ec[1]-5, ec[0]+5, ec[1]+5,
                                                         fill='cyan', outline=T.BG, tags='dh')

    def _dish_drag_start(self, e):
        self._ensure_transform(); pt = self._c2i(e.x, e.y)
        if not pt: return
        px, py = pt
        for di, (dcx, dcy, dr) in enumerate(self._get_dishes()):
            dc = ((px-dcx)**2+(py-dcy)**2)**0.5
            if dc <= max(15, dr*0.08):
                self._dish_drag = dict(type='move', di=di, s=(px, py), o=(dcx, dcy, dr)); return
            if abs(dc-dr) <= max(14, dr*0.07):
                self._dish_drag = dict(type='resize', di=di, s=(px, py), o=(dcx, dcy, dr)); return

    def _dish_drag_move(self, e):
        self._ensure_transform(); pt = self._c2i(e.x, e.y)
        if not pt or not self._dish_drag: return
        d = self._dish_drag; px, py = pt; dcx, dcy, dr = d['o']; sx, sy = d['s']
        if d['type'] == 'move': self._set_dish(d['di'], int(dcx+px-sx), int(dcy+py-sy), dr)
        else: self._set_dish(d['di'], dcx, dcy, max(20, int(((px-dcx)**2+(py-dcy)**2)**0.5)))
        self._refresh_proc_canvas()

    def _dish_drag_end(self, _e=None):
        if not self._dish_drag: return
        self._dish_drag = None
        if self.current_path: self._run_sync(self.current_path)

    def _set_dish(self, di, cx, cy, r):
        if not self.current_path: return
        dishes = [[int(d[0]), int(d[1]), int(d[2])] for d in self._get_dishes()]
        while len(dishes) <= di: dishes.append(dishes[-1][:] if dishes else [cx, cy, r])
        dishes[di] = [int(cx), int(cy), int(r)]
        self.dish_overrides[self.current_path] = dishes

    # ═══════════════════════ STATS PANELS ════════════════════════════════════

    def _refresh_stats_cur(self):
        if self.current_path and self.image_data.get(self.current_path):
            self._refresh_stats(self.image_data[self.current_path])

    def _grand_total(self, path):
        r = self.image_data.get(path)
        if not r: return (0, 0, 0)
        excl = set(self.excluded_auto.get(path, []))
        aa = sum(len([c for c in col['ws_centers'] if c not in excl]) for col in r['colonies'])
        return (aa, len(self.manual_marks.get(path, [])), len(excl))

    def _px_per_mm(self, result):
        """Pixels per mm based on dish radius and user-set diameter."""
        if not result: return None
        dish = result.get('dish')
        if not dish: return None
        r_px = dish[2]
        d_mm = self.p['dish_diameter_mm'].get()
        if d_mm > 0 and r_px > 0:
            return (r_px * 2) / d_mm
        return None

    def _calc_cfu_ml(self, colony_count):
        vol = self.p['plating_volume_ml'].get()
        dil = self.p['dilution_factor'].get()
        if vol > 0 and dil > 0:
            return colony_count / (vol * (1.0 / dil))
        return None

    def _classify_morphology(self, result):
        """Classify colonies into size/shape categories."""
        if not result or not result['colonies']: return {}
        areas = [c['feat']['area'] for c in result['colonies']]
        circs = [c['feat']['circularity'] for c in result['colonies']]
        med_a = np.median(areas) if areas else 1
        small = sum(1 for a in areas if a < med_a * 0.5)
        medium = sum(1 for a in areas if med_a * 0.5 <= a <= med_a * 2.0)
        large = sum(1 for a in areas if a > med_a * 2.0)
        rnd = sum(1 for c in circs if c > 0.7)
        irreg = sum(1 for c in circs if c <= 0.7)
        return dict(small=small, medium=medium, large=large, round=rnd, irregular=irreg)

    def _refresh_stats(self, result):
        path = self.current_path
        aa, mn, en = self._grand_total(path) if path else (result['total'], 0, 0)
        singles = result['colony_count'] - result['cluster_count']
        grand = aa + mn; hid = result.get('hidden_estimate', 0)
        dr = result.get('dish_results', [])
        dl = ""
        if len(dr) > 1:
            for i, d in enumerate(dr, 1): dl += f"  #{i}:       {d['total']:>6}\n"
        # Calibration
        ppm = self._px_per_mm(result)
        area_str = f"{result['avg_colony_area']:>5.0f} px"
        if ppm and ppm > 0:
            area_mm2 = result['avg_colony_area'] / (ppm * ppm)
            area_str += f" ({area_mm2:.3f} мм\u00b2)"
        # CFU
        cfu = self._calc_cfu_ml(grand)
        cfu_str = f"\nCFU/мл:     {cfu:>.0f}\n" if cfu is not None and cfu > 0 else ""
        # Morphology
        morph = self._classify_morphology(result)
        morph_str = ""
        if morph:
            morph_str = (f"{'─'*22}\n"
                         f"Мелкие:     {morph['small']:>6}\n"
                         f"Средние:    {morph['medium']:>6}\n"
                         f"Крупные:    {morph['large']:>6}\n"
                         f"Круглые:    {morph['round']:>6}\n"
                         f"Неправ.:    {morph['irregular']:>6}\n")

        text = (f"Авто:       {result['total']:>6}\n" + dl +
                f"  одиноч.:  {singles:>6}\n  кластер.: {result['cluster_count']:>6}\n"
                f"  исключ.:  {en:>6}\nРучных:     {mn:>6}\n"
                f"{'─'*22}\nИТОГО:      {grand:>6}\n"
                + cfu_str
                + (f"+скрытых:   {hid:>6}\n" if hid else "")
                + f"{'─'*22}\nСр.площ.:   {area_str}\n"
                + (f"Этикетка:   да\n" if result.get('has_label') else "")
                + morph_str)
        self._stats_text.config(state=tk.NORMAL)
        self._stats_text.delete("1.0", tk.END)
        self._stats_text.insert(tk.END, text)
        self._stats_text.config(state=tk.DISABLED)

    def _refresh_results(self):
        lines, ts = [], 0
        for path in self.image_paths:
            nm = self.display_names.get(path, Path(path).name)
            short = (nm[:18]+"...") if len(nm) > 19 else nm
            r = self.image_data.get(path)
            if r:
                a, m, _ = self._grand_total(path); c = a+m; ts += c
                lines.append(f"{short:<19} {c:>4}")
            else: lines.append(f"{short:<19}  ---")
        lines.append(f"{'─'*24}")
        lines.append(f"{'ИТОГО':<19} {ts:>4}")
        self._res_text.config(state=tk.NORMAL)
        self._res_text.delete("1.0", tk.END)
        self._res_text.insert(tk.END, "\n".join(lines))
        self._res_text.config(state=tk.DISABLED)

    # ═══════════════════════ EXPORT ══════════════════════════════════════════

    def _export_excel(self):
        processed = [(p, r) for p in self.image_paths if (r := self.image_data.get(p))]
        if not processed: messagebox.showwarning("Нет данных", "Обработайте изображения."); return
        sp = filedialog.asksaveasfilename(defaultextension=".xlsx",
            initialfile=f"colonies_{datetime.date.today()}.xlsx", filetypes=[("Excel", "*.xlsx")])
        if not sp: return
        try:
            wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Результаты"
            hf = PatternFill("solid", fgColor="2E4DA0")
            hfn = XlFont(color="FFFFFF", bold=True, size=11)
            tf = PatternFill("solid", fgColor="D6E4BC")
            af = PatternFill("solid", fgColor="EEF2FB")
            thin = Side(style='thin', color="AAAAAA")
            brd = Border(left=thin, right=thin, top=thin, bottom=thin)
            ca, la = Alignment(horizontal='center', vertical='center'), Alignment(horizontal='left', vertical='center')
            ws.merge_cells("A1:H1")
            ws["A1"].value = f"Colony Counter v{C.VERSION}"; ws["A1"].font = XlFont(bold=True, size=14, color="1A1A6E"); ws["A1"].alignment = ca
            ws["A2"] = f"Дата: {datetime.datetime.now().strftime('%d.%m.%Y %H:%M')}"
            ws["A2"].font = XlFont(italic=True, size=9, color="555555")
            headers = ["#","Файл","Авто","Искл.","Ручн.","ИТОГО","Одиноч.","Кластер.","Ср.площ.","Путь"]
            for ci, h in enumerate(headers, 1):
                c = ws.cell(row=4, column=ci, value=h); c.fill = hf; c.font = hfn; c.alignment = ca; c.border = brd
            for i, w in enumerate([4,28,10,10,10,12,14,12,16,50], 1): ws.column_dimensions[get_column_letter(i)].width = w
            row, ts = 5, 0
            for idx, (path, result) in enumerate(processed, 1):
                aa, mn, en = self._grand_total(path); g = aa+mn
                singles = result['colony_count'] - result['cluster_count']
                dp = path.split('::frame')[0] if '::frame' in path else path
                fl = af if idx % 2 == 0 else PatternFill()
                vals = [idx, self.display_names.get(path, Path(path).name), result['total'],
                        f"-{en}" if en else 0, mn, g, singles, result['cluster_count'],
                        round(result['avg_colony_area'], 1), dp]
                for ci, v in enumerate(vals, 1):
                    c = ws.cell(row=row, column=ci, value=v)
                    c.border = brd; c.fill = fl; c.alignment = ca if ci not in (2, 10) else la
                    if ci == 6: c.font = XlFont(bold=True)
                ts += g; row += 1
            ws.cell(row=row, column=2, value="ИТОГО").font = XlFont(bold=True)
            ws.cell(row=row, column=6, value=ts).font = XlFont(bold=True, size=12)
            for ci in range(1, 11):
                c = ws.cell(row=row, column=ci); c.fill = tf; c.border = brd; c.alignment = ca
            ws.freeze_panes = "A5"; ws.auto_filter.ref = f"A4:J{row-1}"
            wb.save(sp); messagebox.showinfo("OK", f"Сохранено:\n{sp}")
            self._set_status(f"Excel: {sp}")
        except Exception as e: messagebox.showerror("Ошибка", str(e))

    def _export_pdf(self):
        if not HAS_MATPLOTLIB: return
        processed = [(p, r) for p in self.image_paths if (r := self.image_data.get(p))]
        if not processed: messagebox.showwarning("Нет данных", "Обработайте изображения."); return
        sp = filedialog.asksaveasfilename(defaultextension=".pdf",
            initialfile=f"colonies_{datetime.date.today()}.pdf", filetypes=[("PDF", "*.pdf")])
        if not sp: return
        try:
            from matplotlib.backends.backend_pdf import PdfPages
            with PdfPages(sp) as pdf:
                fig, ax = plt.subplots(figsize=(11, 8)); ax.axis('off')
                ax.set_title(f"Colony Counter v{C.VERSION}\n{datetime.datetime.now().strftime('%d.%m.%Y %H:%M')}", fontsize=16, fontweight='bold')
                td = []
                for i, (p, r) in enumerate(processed, 1):
                    aa, mn, _ = self._grand_total(p)
                    nm = self.display_names.get(p, Path(p).name)
                    if len(nm) > 28: nm = nm[:25]+"..."
                    td.append([i, nm, r['total'], mn, aa+mn, round(r['avg_colony_area'], 1)])
                t = ax.table(cellText=td, colLabels=["#","Файл","Авто","Ручн.","ИТОГО","Ср.пл."], cellLoc='center', loc='center')
                t.auto_set_font_size(False); t.set_fontsize(9); t.scale(1, 1.5)
                for k, cell in t.get_celld().items():
                    if k[0] == 0: cell.set_facecolor('#2E4DA0'); cell.set_text_props(color='white', fontweight='bold')
                pdf.savefig(fig); plt.close(fig)
                for p, r in processed:
                    ann = self._get_annotated(p)
                    if ann is None: continue
                    fig, axes = plt.subplots(1, 2, figsize=(14, 7))
                    fig.suptitle(self.display_names.get(p, Path(p).name), fontsize=14, fontweight='bold')
                    axes[0].imshow(cv2.cvtColor(ann, cv2.COLOR_BGR2RGB)); axes[0].set_title("Результат"); axes[0].axis('off')
                    areas = [c['feat']['area'] for c in r['colonies']]
                    if areas:
                        axes[1].hist(areas, bins=min(30, max(5, len(areas)//3)), color='#2E4DA0', edgecolor='white', alpha=0.8)
                        axes[1].set_title("Площади"); axes[1].set_xlabel("пикс."); axes[1].set_ylabel("кол-во")
                        axes[1].axvline(r['avg_colony_area'], color='red', linestyle='--', label=f"Ср.={r['avg_colony_area']:.0f}")
                        axes[1].legend()
                    else: axes[1].text(0.5, 0.5, "Нет", ha='center'); axes[1].axis('off')
                    fig.tight_layout(); pdf.savefig(fig); plt.close(fig)
            messagebox.showinfo("OK", f"PDF:\n{sp}"); self._set_status(f"PDF: {sp}")
        except Exception as e: messagebox.showerror("Ошибка", str(e))

    def _export_image(self):
        if not self.current_path: return
        ci = self._make_proc(self.current_path)
        if ci is None: messagebox.showwarning("Нет", "Обработайте."); return
        sp = filedialog.asksaveasfilename(defaultextension=".png",
            initialfile=f"{Path(self.current_path).stem}_result.png",
            filetypes=[("PNG", "*.png"), ("JPEG", "*.jpg"), ("BMP", "*.bmp")])
        if sp: cv_imwrite(sp, ci); self._set_status(f"Сохранено: {Path(sp).name}")

    # ═══════════════════════ STATISTICS WINDOW ═══════════════════════════════

    def _show_statistics(self):
        if not HAS_MATPLOTLIB: return
        processed = [(p, r) for p in self.image_paths if (r := self.image_data.get(p))]
        if not processed: messagebox.showwarning("Нет", "Обработайте."); return
        win = tk.Toplevel(self.root); win.title("Статистика"); win.geometry("1000x700")
        win.config(bg=T.BG)
        fig = Figure(figsize=(12, 8), dpi=100, facecolor=T.BG1)
        names = [self.display_names.get(p, Path(p).name)[:14] for p, _ in processed]
        counts = [sum(self._grand_total(p)[:2]) for p, _ in processed]

        ax1 = fig.add_subplot(2, 2, 1, facecolor=T.BG)
        ax1.bar(range(len(names)), counts, color=T.ACCENT); ax1.set_xticks(range(len(names)))
        ax1.set_xticklabels(names, rotation=45, ha='right', fontsize=7, color=T.FG3)
        ax1.set_title("Колоний", color=T.FG, fontsize=10); ax1.tick_params(colors=T.FG3)

        ax2 = fig.add_subplot(2, 2, 2, facecolor=T.BG)
        all_a = [c['feat']['area'] for _, r in processed for c in r['colonies']]
        if all_a:
            ax2.hist(all_a, bins=min(50, max(10, len(all_a)//5)), color=T.ACCENT, edgecolor=T.BG, alpha=0.8)
            ax2.axvline(np.median(all_a), color=T.RED, linestyle='--')
        ax2.set_title("Площади", color=T.FG, fontsize=10); ax2.tick_params(colors=T.FG3)

        ax3 = fig.add_subplot(2, 2, 3, facecolor=T.BG)
        sl = [r['colony_count']-r['cluster_count'] for _, r in processed]
        cl = [r['cluster_count'] for _, r in processed]
        ax3.bar(range(len(names)), sl, label='Одиноч.', color=T.ACCENT)
        ax3.bar(range(len(names)), cl, bottom=sl, label='Кластер.', color=T.ORANGE)
        ax3.set_xticks(range(len(names))); ax3.set_xticklabels(names, rotation=45, ha='right', fontsize=7, color=T.FG3)
        ax3.legend(fontsize=8); ax3.set_title("Типы", color=T.FG, fontsize=10); ax3.tick_params(colors=T.FG3)

        ax4 = fig.add_subplot(2, 2, 4, facecolor=T.BG)
        ax4.bar(range(len(names)), [r['avg_colony_area'] for _, r in processed], color='#8b5cf6')
        ax4.set_xticks(range(len(names))); ax4.set_xticklabels(names, rotation=45, ha='right', fontsize=7, color=T.FG3)
        ax4.set_title("Ср. площадь", color=T.FG, fontsize=10); ax4.tick_params(colors=T.FG3)

        fig.tight_layout()
        c = FigureCanvasTkAgg(fig, master=win); c.draw()
        c.get_tk_widget().pack(fill=tk.BOTH, expand=True)

    # ═══════════════════════ SESSION ═════════════════════════════════════════

    def _save_session(self):
        sp = filedialog.asksaveasfilename(defaultextension=".colsession",
            initialfile=f"session_{datetime.date.today()}.colsession",
            filetypes=[("Session", "*.colsession"), ("JSON", "*.json")])
        if not sp: return
        try:
            sd = str(Path(sp).parent)
            imgs = []
            for path in self.image_paths:
                rp = path.split('::frame')[0] if '::frame' in path else path
                try: rel = os.path.relpath(rp, sd)
                except: rel = rp
                fs = path[len(rp):] if '::frame' in path else ''
                imgs.append({'path': path, 'rel': rel+fs, 'name': self.display_names.get(path, ''),
                             'marks': [list(m) for m in self.manual_marks.get(path, [])],
                             'excl': [list(c) for c in self.excluded_auto.get(path, [])],
                             'dish_ov': self.dish_overrides.get(path, [])})
            session = {'v': 2, 'cur': self.current_path, 'params': {k: v.get() for k, v in self.p.items()}, 'imgs': imgs}
            with open(sp, 'w', encoding='utf-8') as f: json.dump(session, f, indent=2, ensure_ascii=False)
            self._set_status(f"Сессия: {Path(sp).name}")
        except Exception as e: messagebox.showerror("Ошибка", str(e))

    def _load_session(self):
        lp = filedialog.askopenfilename(filetypes=[("Session", "*.colsession"), ("JSON", "*.json")])
        if not lp: return
        try:
            with open(lp, 'r', encoding='utf-8') as f: s = json.load(f)
            sd = str(Path(lp).parent)
            self.image_paths.clear(); self.image_data.clear(); self.manual_marks.clear()
            self.excluded_auto.clear(); self.display_names.clear(); self.dish_overrides.clear()
            self._cache.cleanup(); self.listbox.delete(0, tk.END)
            self.current_path = self._prev_path = self._pil_orig = self._pil_proc = None
            self.canvas_orig.delete("all"); self.canvas_proc.delete("all")
            for k, v in s.get('params', {}).items():
                if k in self.p: self.p[k].set(v)
            missing = []
            for d in s.get('imgs', []):
                path = d['path']; rp = path.split('::frame')[0] if '::frame' in path else path
                if not Path(rp).exists():
                    rel = d.get('rel', ''); rr = rel.split('::frame')[0] if '::frame' in rel else rel
                    afr = os.path.normpath(os.path.join(sd, rr))
                    if Path(afr).exists(): path = afr + (path[len(rp):] if '::frame' in path else '')
                    else: missing.append(rp); continue
                self.image_paths.append(path); self.image_data[path] = None
                nm = d.get('name') or Path(path).name
                self.display_names[path] = nm; self.listbox.insert(tk.END, nm)
                marks = [tuple(m) for m in d.get('marks', [])]
                if marks: self.manual_marks[path] = marks
                excl = [tuple(c) for c in d.get('excl', [])]
                if excl: self.excluded_auto[path] = excl
                ov = d.get('dish_ov', [])
                if ov: self.dish_overrides[path] = ov
            cur = s.get('cur')
            if cur and cur in self.image_paths:
                idx = self.image_paths.index(cur); self.listbox.selection_set(idx); self.listbox.see(idx)
            msg = f"Сессия: {len(self.image_paths)} изобр."
            if missing: msg += f" Не найдено: {len(missing)}"
            self._set_status(msg + "  Нажмите ОБРАБОТАТЬ ВСЕ")
        except Exception as e: messagebox.showerror("Ошибка", str(e))


    # ═══════════════════════ CSV EXPORT ═════════════════════════════════════

    def _export_csv(self):
        processed = [(p, r) for p in self.image_paths if (r := self.image_data.get(p))]
        if not processed: messagebox.showwarning("Нет данных", "Обработайте."); return
        sp = filedialog.asksaveasfilename(defaultextension=".csv",
            initialfile=f"colonies_{datetime.date.today()}.csv", filetypes=[("CSV", "*.csv")])
        if not sp: return
        try:
            import csv
            with open(sp, 'w', newline='', encoding='utf-8-sig') as f:
                w = csv.writer(f, delimiter=';')
                w.writerow(["#", "Файл", "Авто", "Исключено", "Ручных", "ИТОГО",
                            "Одиночных", "Кластеров", "Ср.площадь_px",
                            "Ср.площадь_мм2", "CFU/мл", "Группа", "Разведение"])
                for idx, (path, result) in enumerate(processed, 1):
                    aa, mn, en = self._grand_total(path); g = aa + mn
                    singles = result['colony_count'] - result['cluster_count']
                    ppm = self._px_per_mm(result)
                    area_mm2 = (result['avg_colony_area'] / (ppm*ppm)) if ppm and ppm > 0 else ""
                    cfu = self._calc_cfu_ml(g)
                    cfu_str = f"{cfu:.0f}" if cfu and cfu > 0 else ""
                    grp = self.p['dilution_group'].get()
                    dil = self.p['dilution_factor'].get()
                    w.writerow([idx, self.display_names.get(path, Path(path).name),
                                result['total'], en, mn, g, singles, result['cluster_count'],
                                round(result['avg_colony_area'], 1),
                                round(area_mm2, 4) if area_mm2 else "",
                                cfu_str, grp, dil])
            self._set_status(f"CSV: {sp}")
        except Exception as e: messagebox.showerror("Ошибка", str(e))

    # ═══════════════════════ PRESETS ═════════════════════════════════════════

    def _apply_preset(self, name):
        presets = {
            'default':   dict(min_area=80, max_area=3000, threshold=25),
            'sensitive': dict(min_area=40, max_area=5000, threshold=15),
            'strict':    dict(min_area=120, max_area=2000, threshold=40),
            'large':     dict(min_area=200, max_area=10000, threshold=30),
        }
        p = presets.get(name)
        if not p: return
        for k, v in p.items():
            if k in self.p: self.p[k].set(v)
        self._set_status(f"Пресет: {name} (F1-F4)")

    # ═══════════════════════ HEATMAP ════════════════════════════════════════

    def _show_heatmap(self):
        if not HAS_MATPLOTLIB: return
        if not self.current_path: return
        result = self.image_data.get(self.current_path)
        if not result or not result['colonies']:
            messagebox.showwarning("Нет", "Обработайте."); return
        win = tk.Toplevel(self.root); win.title("Тепловая карта плотности")
        win.geometry("700x700"); win.config(bg=T.BG)
        fig = Figure(figsize=(7, 7), dpi=100, facecolor=T.BG1)
        ax = fig.add_subplot(111, facecolor=T.BG)
        # Get colony positions
        xs = [c['center'][0] for c in result['colonies']]
        ys = [c['center'][1] for c in result['colonies']]
        # Show original image as background
        ann = self._get_annotated(self.current_path)
        if ann is not None:
            ax.imshow(cv2.cvtColor(ann, cv2.COLOR_BGR2RGB), alpha=0.4)
        # Heatmap overlay
        if xs and ys:
            from matplotlib.colors import LinearSegmentedColormap
            ax.hexbin(xs, ys, gridsize=20, cmap='YlOrRd', alpha=0.6, mincnt=1)
            ax.scatter(xs, ys, c=T.ACCENT, s=8, alpha=0.5, edgecolors='none')
        ax.set_title(f"Плотность колоний — {self.display_names.get(self.current_path, '')}",
                     color=T.FG, fontsize=11)
        ax.set_aspect('equal')
        ax.invert_yaxis()
        ax.tick_params(colors=T.FG3)
        fig.tight_layout()
        c = FigureCanvasTkAgg(fig, master=win); c.draw()
        c.get_tk_widget().pack(fill=tk.BOTH, expand=True)

    # ═══════════════════════ SIDE BY SIDE ════════════════════════════════════

    def _show_side_by_side(self):
        if not HAS_MATPLOTLIB: return
        processed = [(p, r) for p in self.image_paths if (r := self.image_data.get(p))]
        if len(processed) < 2:
            messagebox.showwarning("Нужно 2+", "Обработайте минимум 2 изображения."); return
        win = tk.Toplevel(self.root); win.title("Сравнение бок-о-бок")
        win.geometry("1000x600"); win.config(bg=T.BG)
        # Selection frame
        sel_f = tk.Frame(win, bg=T.BG)
        sel_f.pack(fill=tk.X, padx=10, pady=6)
        names = [self.display_names.get(p, Path(p).name) for p, _ in processed]
        tk.Label(sel_f, text="Изображение 1:", bg=T.BG, fg=T.FG3, font=T.FONT_SM).pack(side=tk.LEFT)
        var1 = tk.StringVar(value=names[0])
        tk.OptionMenu(sel_f, var1, *names).pack(side=tk.LEFT, padx=4)
        tk.Label(sel_f, text="Изображение 2:", bg=T.BG, fg=T.FG3, font=T.FONT_SM).pack(side=tk.LEFT, padx=(20, 0))
        var2 = tk.StringVar(value=names[min(1, len(names)-1)])
        tk.OptionMenu(sel_f, var2, *names).pack(side=tk.LEFT, padx=4)

        canvas_frame = tk.Frame(win, bg=T.BG)
        canvas_frame.pack(fill=tk.BOTH, expand=True)

        def draw():
            for w in canvas_frame.winfo_children(): w.destroy()
            n1, n2 = var1.get(), var2.get()
            p1 = next((p for p, _ in processed if self.display_names.get(p, Path(p).name) == n1), None)
            p2 = next((p for p, _ in processed if self.display_names.get(p, Path(p).name) == n2), None)
            if not p1 or not p2: return
            fig = Figure(figsize=(12, 5), dpi=100, facecolor=T.BG1)
            for i, (p, title) in enumerate([(p1, n1), (p2, n2)]):
                ax = fig.add_subplot(1, 2, i+1, facecolor=T.BG)
                ann = self._get_annotated(p)
                if ann is not None:
                    ax.imshow(cv2.cvtColor(ann, cv2.COLOR_BGR2RGB))
                r = self.image_data.get(p)
                aa, mn, _ = self._grand_total(p)
                ax.set_title(f"{title}\n{aa+mn} колоний", color=T.FG, fontsize=10)
                ax.axis('off')
            fig.tight_layout()
            c = FigureCanvasTkAgg(fig, master=canvas_frame); c.draw()
            c.get_tk_widget().pack(fill=tk.BOTH, expand=True)

        DarkButton(sel_f, "Показать", draw, 'primary', small=True).pack(side=tk.LEFT, padx=10)
        draw()

    # ═══════════════════════ ANNOTATIONS ═════════════════════════════════════

    def _add_annotation(self):
        if not self.current_path: return
        text = simpledialog.askstring("Аннотация", "Текст для изображения:", parent=self.root)
        if not text or not text.strip(): return
        self._annotations.setdefault(self.current_path, []).append(text.strip())
        self._refresh_proc_canvas()
        self._set_status(f"Аннотация добавлена: {text.strip()[:30]}")

    def _make_proc(self, path):
        """Override to include annotations."""
        result = self.image_data.get(path)
        if not result: return None
        excl = set(self.excluded_auto.get(path, [])); marks = self.manual_marks.get(path, [])
        if not excl and not marks and path not in self._annotations:
            return self._get_annotated(path)
        base = self._get_clean(path)
        if base is None: return self._get_annotated(path)
        base = base.copy(); cr = result.get('col_radius', 10)
        for dcx, dcy, dr in (self.dish_overrides.get(path) or result.get('dishes', [result['dish']])):
            cv2.circle(base, (int(dcx), int(dcy)), int(dr), (80, 80, 220), 2)
        for dd in result.get('dish_results', []):
            if dd.get('has_label') and dd.get('label_mask') is not None:
                lc = np.zeros_like(base); lc[dd['label_mask'] > 0] = (0, 60, 180)
                cv2.addWeighted(base, 1.0, lc, 0.35, 0, base)
                lcs, _ = cv2.findContours(dd['label_mask'], cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
                cv2.drawContours(base, lcs, -1, (0, 100, 255), 2)
        at = 0
        for col in result['colonies']:
            ac = [c for c in col['ws_centers'] if c not in excl]
            ic = [c for c in col['ws_centers'] if c in excl]
            if ac:
                is_cl = len(ac) > 1 or col['is_cluster']
                color = (20, 160, 255) if is_cl else (30, 200, 30)
                if not col['is_cluster']:
                    cv2.drawContours(base, [col['contour']], -1, color, 2)
                    cv2.circle(base, ac[0], 3, color, -1)
                else:
                    for (ax, ay) in ac: cv2.circle(base, (ax, ay), cr, color, 2)
                at += len(ac)
            for (ex, ey) in ic:
                cv2.drawMarker(base, (ex, ey), (0, 0, 200), cv2.MARKER_TILTED_CROSS, cr*2, 2, cv2.LINE_AA)
        mr = max(2, cr//3)
        for (mx, my) in marks:
            cv2.circle(base, (mx, my), mr, (0, 215, 255), 1)
            cv2.drawMarker(base, (mx, my), (0, 215, 255), cv2.MARKER_CROSS, mr*2, 1, cv2.LINE_AA)
        g = at + len(marks); lbl = f"Kolonii: {g}"
        (lw, lh), _ = cv2.getTextSize(lbl, cv2.FONT_HERSHEY_SIMPLEX, 1.0, 2)
        cv2.rectangle(base, (6, 6), (lw+18, lh+18), (30, 30, 30), -1)
        cv2.putText(base, lbl, (12, lh+10), cv2.FONT_HERSHEY_SIMPLEX, 1.0, (60, 220, 60), 2, cv2.LINE_AA)
        # Draw text annotations
        annotations = self._annotations.get(path, [])
        h_img = base.shape[0]
        for ai, atxt in enumerate(annotations):
            y_pos = h_img - 30 - ai * 28
            cv2.putText(base, atxt, (12, y_pos), cv2.FONT_HERSHEY_SIMPLEX, 0.6,
                        (255, 255, 255), 1, cv2.LINE_AA)
        return base

    # ═══════════════════════ DILUTION CURVE ══════════════════════════════════

    def _show_dilution_curve(self):
        if not HAS_MATPLOTLIB:
            messagebox.showwarning("Нет matplotlib", "Установите matplotlib."); return
        # Group images by dilution_factor if they share the same group name
        processed = [(p, r) for p in self.image_paths if (r := self.image_data.get(p))]
        if len(processed) < 2:
            messagebox.showwarning("Нужно 2+", "Обработайте минимум 2 изображения с разными разведениями."); return
        # Ask user for dilution factors per image
        win = tk.Toplevel(self.root); win.title("Серийные разведения")
        win.geometry("600x500"); win.config(bg=T.BG)
        tk.Label(win, text="Укажите разведение (1:N) для каждого изображения:",
                 bg=T.BG, fg=T.FG, font=T.FONT_SM).pack(padx=10, pady=8)
        entries = []
        sf = tk.Frame(win, bg=T.BG)
        sf.pack(fill=tk.BOTH, expand=True, padx=10)
        for p, r in processed:
            row = tk.Frame(sf, bg=T.BG)
            row.pack(fill=tk.X, pady=2)
            nm = self.display_names.get(p, Path(p).name)
            aa, mn, _ = self._grand_total(p)
            tk.Label(row, text=f"{nm[:25]} ({aa+mn} кол.)", bg=T.BG, fg=T.FG3,
                     font=T.FONT_XS, width=35, anchor='w').pack(side=tk.LEFT)
            tk.Label(row, text="1:", bg=T.BG, fg=T.FG3, font=T.FONT_SM).pack(side=tk.LEFT)
            var = tk.StringVar(value="1")
            e = tk.Entry(row, textvariable=var, bg=T.BG2, fg=T.FG, font=T.FONT_SM,
                         width=10, insertbackground=T.FG, highlightthickness=1,
                         highlightbackground=T.BORDER, bd=0)
            e.pack(side=tk.LEFT, padx=4)
            entries.append((p, var))

        vol_var = tk.StringVar(value=str(self.p['plating_volume_ml'].get()))
        vr = tk.Frame(win, bg=T.BG)
        vr.pack(fill=tk.X, padx=10, pady=4)
        tk.Label(vr, text="Объём посева (мл):", bg=T.BG, fg=T.FG3, font=T.FONT_SM).pack(side=tk.LEFT)
        tk.Entry(vr, textvariable=vol_var, bg=T.BG2, fg=T.FG, font=T.FONT_SM,
                 width=8, insertbackground=T.FG, bd=0).pack(side=tk.LEFT, padx=4)

        def calc():
            try:
                vol = float(vol_var.get())
                points = []
                for p, var in entries:
                    dil = float(var.get())
                    aa, mn, _ = self._grand_total(p)
                    count = aa + mn
                    cfu = count / (vol * (1.0 / dil)) if vol > 0 and dil > 0 else 0
                    points.append((dil, count, cfu))
                points.sort(key=lambda x: x[0])
                # Plot
                fig_win = tk.Toplevel(win); fig_win.title("Кривая разведений")
                fig_win.geometry("800x500"); fig_win.config(bg=T.BG)
                fig = Figure(figsize=(8, 5), dpi=100, facecolor=T.BG1)
                ax1 = fig.add_subplot(1, 2, 1, facecolor=T.BG)
                dils = [p[0] for p in points]
                counts = [p[1] for p in points]
                ax1.plot(dils, counts, 'o-', color=T.ACCENT, markersize=8)
                ax1.set_xlabel("Разведение (1:N)", color=T.FG3)
                ax1.set_ylabel("Колоний", color=T.FG3)
                ax1.set_title("Колонии vs разведение", color=T.FG)
                ax1.set_xscale('log')
                ax1.tick_params(colors=T.FG3)
                ax1.grid(True, alpha=0.2)
                ax2 = fig.add_subplot(1, 2, 2, facecolor=T.BG)
                cfus = [p[2] for p in points]
                ax2.bar(range(len(dils)), cfus, color=T.ACCENT)
                ax2.set_xticks(range(len(dils)))
                ax2.set_xticklabels([f"1:{int(d)}" for d in dils], rotation=45, ha='right', fontsize=8, color=T.FG3)
                ax2.set_ylabel("CFU/мл", color=T.FG3)
                ax2.set_title(f"CFU/мл (среднее: {np.mean(cfus):.0f})", color=T.FG)
                ax2.tick_params(colors=T.FG3)
                fig.tight_layout()
                c = FigureCanvasTkAgg(fig, master=fig_win); c.draw()
                c.get_tk_widget().pack(fill=tk.BOTH, expand=True)
            except Exception as e:
                messagebox.showerror("Ошибка", str(e))

        DarkButton(win, "Рассчитать", calc, 'primary').pack(padx=10, pady=8)

    # ═══════════════════════ REPRODUCIBILITY ═════════════════════════════════

    def _show_reproducibility(self):
        """Placeholder for reproducibility comparison (run with different params)."""
        # This would re-run the same image with varied thresholds and compare
        if not self.current_path or not self.image_data.get(self.current_path):
            messagebox.showwarning("Нет", "Обработайте изображение."); return
        if not HAS_MATPLOTLIB: return
        path = self.current_path
        base_thresh = self.p['threshold'].get()
        results = []
        for delta in [-10, -5, 0, 5, 10]:
            t = max(5, min(100, base_thresh + delta))
            params = self._get_params()
            params['threshold'] = t
            try:
                r = self._process_image(path, params)
                results.append((t, r['total'], r['colony_count'], r['cluster_count']))
            except Exception:
                continue
        if not results: return
        win = tk.Toplevel(self.root); win.title("Воспроизводимость")
        win.geometry("700x400"); win.config(bg=T.BG)
        fig = Figure(figsize=(7, 4), dpi=100, facecolor=T.BG1)
        ax = fig.add_subplot(111, facecolor=T.BG)
        thresholds = [r[0] for r in results]
        totals = [r[1] for r in results]
        ax.plot(thresholds, totals, 'o-', color=T.ACCENT, markersize=10, linewidth=2)
        ax.axvline(base_thresh, color=T.RED, linestyle='--', label=f"Текущий: {base_thresh}")
        for t, tot in zip(thresholds, totals):
            ax.annotate(str(tot), (t, tot), textcoords="offset points",
                        xytext=(0, 12), ha='center', color=T.FG, fontsize=10)
        ax.set_xlabel("Порог", color=T.FG3, fontsize=11)
        ax.set_ylabel("Колоний", color=T.FG3, fontsize=11)
        ax.set_title("Чувствительность к порогу", color=T.FG, fontsize=13)
        ax.tick_params(colors=T.FG3)
        ax.legend(fontsize=9)
        ax.grid(True, alpha=0.2)
        fig.tight_layout()
        c = FigureCanvasTkAgg(fig, master=win); c.draw()
        c.get_tk_widget().pack(fill=tk.BOTH, expand=True)


# ═══════════════════════════ ENTRY POINT ═════════════════════════════════════

def main():
    root = tk.Tk()
    try: root.tk.call('tk', 'scaling', 1.25)
    except: pass
    App(root)
    root.mainloop()

if __name__ == "__main__":
    main()
