#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Colony Counter v1.1
Автоматический подсчет колоний бактерий на чашках Петри
Поддерживает: одиночные колонии и кластеры (через watershed + оценку по площади)
Новое в v1.1:
  • Обнаружение и маскирование этикетки/наклейки на чашке
  • Log-space гистограмма для точной оценки площади одиночной колонии
  • Per-component watershed через skimage.peak_local_max
  • Оценка скрытых под этикеткой колоний
  • Morphology CLOSE после OPEN (закрывает зазоры внутри колоний)
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog
import cv2
import numpy as np
from PIL import Image, ImageTk
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os
import json
import datetime
from pathlib import Path


# ─────────────────────────── IMAGE PROCESSING ────────────────────────────────

class ImageProcessor:
    """Вся логика обработки изображений."""

    # ── Dish detection ──────────────────────────────────────────────────────
    @staticmethod
    def detect_dish(gray):
        """Найти границу чашки Петри методом Хафа. Возвращает (cx, cy, r)."""
        h, w = gray.shape
        blurred = cv2.GaussianBlur(gray, (21, 21), 0)

        min_r = int(min(h, w) * 0.28)
        max_r = int(min(h, w) * 0.56)

        circles = cv2.HoughCircles(
            blurred, cv2.HOUGH_GRADIENT, dp=1,
            minDist=min(h, w) // 2,
            param1=60, param2=35,
            minRadius=min_r, maxRadius=max_r
        )

        if circles is not None:
            circles = np.round(circles[0]).astype(int)
            # Берём наибольший круг
            circles = sorted(circles, key=lambda c: c[2], reverse=True)
            return tuple(circles[0])
        else:
            # Запасной вариант: вписанная окружность
            cx, cy = w // 2, h // 2
            r = int(min(h, w) * 0.44)
            return (cx, cy, r)

    @staticmethod
    def detect_dishes(gray, max_dishes=2):
        """Найти до max_dishes чашек Петри. Возвращает список [(cx, cy, r), ...]."""
        h, w = gray.shape
        blurred = cv2.GaussianBlur(gray, (21, 21), 0)
        min_r = int(min(h, w) * 0.28)
        max_r = int(min(h, w) * 0.56)
        circles = cv2.HoughCircles(
            blurred, cv2.HOUGH_GRADIENT, dp=1,
            minDist=min_r,
            param1=60, param2=35,
            minRadius=min_r, maxRadius=max_r
        )
        if circles is None:
            return [(w // 2, h // 2, int(min(h, w) * 0.44))]
        circles = sorted(np.round(circles[0]).astype(int).tolist(),
                         key=lambda c: c[2], reverse=True)
        result = []
        for cx, cy, r in circles:
            if all(((cx - rx)**2 + (cy - ry)**2) ** 0.5 >= max(r, rr) * 0.7
                   for rx, ry, rr in result):
                result.append((int(cx), int(cy), int(r)))
            if len(result) >= max_dishes:
                break
        return result or [(w // 2, h // 2, int(min(h, w) * 0.44))]

    # ── Background normalisation ─────────────────────────────────────────────
    @staticmethod
    def normalize_background(gray, mask):
        """
        Вычитание фона: колонии темнее фона (проходящий свет).
        Возвращает изображение, где колонии — светлые объекты.
        """
        # Большое морфологическое открытие «стирает» мелкие колонии,
        # оставляя только плавный фон.
        k = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (71, 71))
        background = cv2.morphologyEx(gray, cv2.MORPH_DILATE, k)
        background = cv2.morphologyEx(background, cv2.MORPH_ERODE, k)

        # background ярче → вычитаем изображение из фона: колонии становятся яркими
        diff = cv2.subtract(background, gray)
        diff = cv2.bitwise_and(diff, diff, mask=mask)
        return diff

    # ── Label detection ──────────────────────────────────────────────────────
    @staticmethod
    def detect_label_mask(gray, dish_mask):
        """
        Обнаруживает прямоугольную этикетку/наклейку на чашке Петри.
        Возвращает маску области этикетки (или нулевую маску если не найдена).
        """
        masked = cv2.bitwise_and(gray, gray, mask=dish_mask)
        _, dark = cv2.threshold(masked, 60, 255, cv2.THRESH_BINARY_INV)
        dark = cv2.bitwise_and(dark, dark, mask=dish_mask)
        contours, _ = cv2.findContours(dark, cv2.RETR_EXTERNAL,
                                        cv2.CHAIN_APPROX_SIMPLE)
        label_mask = np.zeros(gray.shape, dtype=np.uint8)
        dish_area = cv2.countNonZero(dish_mask)

        for cnt in contours:
            area = cv2.contourArea(cnt)
            if area < dish_area * 0.02 or area > dish_area * 0.4:
                continue
            rect = cv2.minAreaRect(cnt)
            rw, rh = rect[1]
            if rw * rh == 0:
                continue
            aspect = max(rw, rh) / (min(rw, rh) + 1e-5)
            fill_ratio = area / (rw * rh)
            if aspect > 2.5 and fill_ratio > 0.4:
                box = np.intp(cv2.boxPoints(rect))
                cv2.fillConvexPoly(label_mask, box, 255)

        if cv2.countNonZero(label_mask) > 0:
            k = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (30, 30))
            label_mask = cv2.dilate(label_mask, k, iterations=2)
        return label_mask

    # ── Single-colony area estimation (log-space histogram) ──────────────────
    @staticmethod
    def estimate_single_colony_area(areas: np.ndarray) -> float:
        """
        Находит модальную площадь одиночной колонии через гистограмму в
        логарифмическом пространстве. Устойчиво к кластерам, которые бы
        сместили медиану вверх.
        """
        if len(areas) < 3:
            return float(np.median(areas))

        log_areas = np.log(areas[areas > 0])
        n_bins = min(30, max(10, len(log_areas) // 3))
        counts, edges = np.histogram(log_areas, bins=n_bins)

        peak_idx = int(np.argmax(counts))
        lo = edges[max(0, peak_idx - 1)]
        hi = edges[min(len(edges) - 1, peak_idx + 2)]
        peak_areas = areas[(log_areas >= lo) & (log_areas <= hi)]

        if len(peak_areas) == 0:
            peak_center = (edges[peak_idx] + edges[peak_idx + 1]) / 2
            return float(np.exp(peak_center))
        return float(np.median(peak_areas))

    # ── Per-component watershed ───────────────────────────────────────────────
    @staticmethod
    def watershed_per_component(cnt, avg_area: float, h: int, w: int) -> int:
        """
        Watershed для одного контура через skimage.peak_local_max.
        Возвращает количество обнаруженных пиков (≥ 1).
        """
        try:
            from skimage.feature import peak_local_max
        except ImportError:
            return max(1, round(cv2.contourArea(cnt) / avg_area))

        comp_mask = np.zeros((h, w), dtype=np.uint8)
        cv2.drawContours(comp_mask, [cnt], -1, 255, -1)

        dist = cv2.distanceTransform(comp_mask, cv2.DIST_L2, 5)
        if dist.max() < 2:
            return 1

        expected_r  = max(3, int(np.sqrt(avg_area / np.pi)))
        min_dist    = max(3, int(expected_r * 0.5))
        thresh_abs  = max(2.0, float(dist.max()) * 0.2)

        coords = peak_local_max(
            dist,
            min_distance=min_dist,
            threshold_abs=thresh_abs,
            labels=comp_mask,
        )
        return max(1, len(coords))

    # ── Contour features ─────────────────────────────────────────────────────
    @staticmethod
    def contour_features(cnt):
        area = cv2.contourArea(cnt)
        perimeter = cv2.arcLength(cnt, True)
        circularity = (4 * np.pi * area / (perimeter * perimeter)
                       if perimeter > 0 else 0.0)
        x, y, bw, bh = cv2.boundingRect(cnt)
        aspect_ratio = bw / bh if bh > 0 else 1.0
        hull = cv2.convexHull(cnt)
        hull_area = cv2.contourArea(hull)
        solidity = area / hull_area if hull_area > 0 else 1.0
        M = cv2.moments(cnt)
        if M['m00'] > 0:
            cx = int(M['m10'] / M['m00'])
            cy = int(M['m01'] / M['m00'])
        else:
            cx, cy = x + bw // 2, y + bh // 2
        return dict(area=area, circularity=circularity,
                    aspect_ratio=aspect_ratio, solidity=solidity,
                    cx=cx, cy=cy)

    # ── Grid fill for clusters ────────────────────────────────────────────────
    @staticmethod
    def fill_contour_with_circles(cnt, n, col_radius):
        """
        Возвращает n равномерно распределённых точек внутри контура
        на гексагональной решётке с шагом ≈ диаметр одной колонии.
        """
        if n <= 0:
            return []
        x, y, bw, bh = cv2.boundingRect(cnt)
        if n == 1:
            M = cv2.moments(cnt)
            if M['m00'] > 0:
                return [(int(M['m10'] / M['m00']), int(M['m01'] / M['m00']))]
            return [(x + bw // 2, y + bh // 2)]

        r = max(2, col_radius)
        dy = max(1, int(r * 1.732))   # true hex row height  (r * sqrt(3))
        dx = max(1, int(r * 2.0))     # true hex column step (2r = diameter)

        # Local mask (faster than pointPolygonTest per point)
        mask = np.zeros((bh + 2, bw + 2), dtype=np.uint8)
        shifted = cnt.copy()
        shifted[:, :, 0] -= x
        shifted[:, :, 1] -= y
        cv2.drawContours(mask, [shifted], -1, 255, -1)

        candidates = []
        for row, yi in enumerate(range(r, bh, dy)):
            ox = r if (row % 2) else 0
            for xi in range(ox, bw, dx):
                if yi < mask.shape[0] and xi < mask.shape[1] and mask[yi, xi] > 0:
                    candidates.append((xi + x, yi + y))

        if not candidates:
            M = cv2.moments(cnt)
            if M['m00'] > 0:
                return [(int(M['m10'] / M['m00']), int(M['m01'] / M['m00']))]
            return [(x + bw // 2, y + bh // 2)]

        if len(candidates) <= n:
            return candidates
        step = len(candidates) / n
        return [candidates[int(i * step)] for i in range(n)]

    # ── Per-dish pipeline ────────────────────────────────────────────────────
    def _process_single_dish(self, img, gray, cx, cy, r, params):
        """Полный пайплайн для одной чашки Петри. Возвращает dict с колониями."""
        h, w = img.shape[:2]
        dish_mask = np.zeros((h, w), dtype=np.uint8)
        cv2.circle(dish_mask, (cx, cy), max(1, int(r * 0.96)), 255, -1)

        has_label    = False
        label_mask   = None
        hidden_est   = 0
        detect_label = bool(params.get('detect_label', True))
        if detect_label:
            label_mask = self.detect_label_mask(gray, dish_mask)
            has_label  = cv2.countNonZero(label_mask) > 0

        work_mask = dish_mask.copy()
        if has_label and label_mask is not None:
            work_mask = cv2.bitwise_and(work_mask, cv2.bitwise_not(label_mask))

        normalized = self.normalize_background(gray, work_mask)
        clahe      = cv2.createCLAHE(clipLimit=3.0, tileGridSize=(8, 8))
        enhanced   = clahe.apply(normalized)

        thresh_val = int(params['threshold'])
        _, binary  = cv2.threshold(enhanced, thresh_val, 255, cv2.THRESH_BINARY)
        binary     = cv2.bitwise_and(binary, binary, mask=work_mask)

        k_morph = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (3, 3))
        binary  = cv2.morphologyEx(binary, cv2.MORPH_OPEN,  k_morph, iterations=1)
        binary  = cv2.morphologyEx(binary, cv2.MORPH_CLOSE, k_morph, iterations=1)

        contours, _ = cv2.findContours(binary, cv2.RETR_EXTERNAL,
                                        cv2.CHAIN_APPROX_SIMPLE)
        min_area         = max(1, int(params['min_area']))
        max_area         = max(min_area + 1, int(params['max_area']))
        filter_bubbles   = bool(params['filter_bubbles'])
        use_watershed    = bool(params['use_watershed'])
        filter_elongated = bool(params['filter_elongated'])
        filter_nonconvex = bool(params['filter_nonconvex'])

        raw = []
        for cnt in contours:
            feat = self.contour_features(cnt)
            area = feat['area']
            if area < min_area:
                continue
            if filter_bubbles and feat['circularity'] > 0.90 and area > max_area * 1.8:
                continue
            ar = feat['aspect_ratio']
            if ar > 6 or ar < 0.16:
                continue
            if filter_elongated and area > min_area * 3:
                rect = cv2.minAreaRect(cnt)
                rw, rh = rect[1]
                if min(rw, rh) > 0 and max(rw, rh) / min(rw, rh) > 3.5:
                    continue
            if filter_nonconvex and area > min_area * 5 and feat['solidity'] < 0.45:
                continue
            raw.append(dict(contour=cnt, feat=feat))

        all_areas = np.array([o['feat']['area'] for o in raw
                              if o['feat']['area'] <= max_area], dtype=float)
        if len(all_areas) >= 3:
            avg_area = self.estimate_single_colony_area(all_areas)
        elif len(all_areas) > 0:
            avg_area = float(np.median(all_areas))
        else:
            avg_area = float(min_area * 5)
        avg_area      = max(avg_area, float(min_area))
        cluster_thresh = avg_area * 1.8
        col_radius     = max(4, int(np.sqrt(avg_area / np.pi)))

        colonies = []
        total    = 0
        for obj in raw:
            cnt  = obj['contour']
            feat = obj['feat']
            area = feat['area']
            if area <= cluster_thresh:
                estimated  = 1
                ws_centers = [(feat['cx'], feat['cy'])]
            else:
                area_est = max(1, round(area / avg_area))
                if use_watershed:
                    ws_count = self.watershed_per_component(cnt, avg_area, h, w)
                    if (ws_count >= 2 and ws_count <= area_est * 1.5
                            and ws_count >= area_est * 0.5):
                        estimated = ws_count
                    else:
                        estimated = area_est
                else:
                    estimated = area_est
                ws_centers = self.fill_contour_with_circles(cnt, estimated, col_radius)
            if not ws_centers:
                ws_centers = [(feat['cx'], feat['cy'])]
            total += estimated
            colonies.append(dict(
                contour=cnt, feat=feat,
                center=(feat['cx'], feat['cy']),
                ws_centers=ws_centers,
                estimated=estimated,
                is_cluster=(estimated > 1),
            ))

        if has_label:
            dish_area = cv2.countNonZero(dish_mask)
            work_area = cv2.countNonZero(work_mask)
            if work_area > 0 and work_area < dish_area * 0.95:
                hidden_est = int((total / work_area) * (dish_area - work_area))

        return dict(
            total=total,
            colony_count=len(colonies),
            cluster_count=sum(1 for c in colonies if c['is_cluster']),
            avg_colony_area=avg_area,
            col_radius=col_radius,
            dish=(cx, cy, r),
            colonies=colonies,
            binary=binary,
            enhanced=enhanced,
            has_label=has_label,
            hidden_estimate=hidden_est,
            label_mask=label_mask,
            work_mask=work_mask,
        )

    # ── Main pipeline ────────────────────────────────────────────────────────
    def process(self, image_path: str, params: dict) -> dict:
        img = cv2.imread(image_path)
        if img is None:
            raise ValueError(f"Не удалось открыть: {image_path}")

        # ── Масштабирование ───────────────────────────────────────────────
        h0, w0 = img.shape[:2]
        max_dim = 2000
        scale   = 1.0
        if max(h0, w0) > max_dim:
            scale = max_dim / max(h0, w0)
            img   = cv2.resize(img, (int(w0 * scale), int(h0 * scale)),
                               interpolation=cv2.INTER_AREA)
        h, w = img.shape[:2]
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

        # ── Обнаружение чашек (или ручные переопределения) ────────────────
        overrides = params.get('dish_overrides', None)
        if overrides:
            dishes = [(int(cx), int(cy), int(r)) for cx, cy, r in overrides]
        else:
            dishes = self.detect_dishes(gray)

        # ── Обработка каждой чашки ────────────────────────────────────────
        annotated    = img.copy()
        all_colonies = []
        dish_results = []
        show_nums    = bool(params.get('show_numbers', True))

        for dish_idx, (dcx, dcy, dr) in enumerate(dishes):
            dr_dict = self._process_single_dish(img, gray, dcx, dcy, dr, params)
            for col in dr_dict['colonies']:
                col['dish_idx'] = dish_idx
            all_colonies.extend(dr_dict['colonies'])
            dish_results.append(dr_dict)

            # ── Аннотация чашки ───────────────────────────────────────────
            cv2.circle(annotated, (dcx, dcy), dr, (80, 80, 220), 2)
            if dr_dict['has_label'] and dr_dict['label_mask'] is not None:
                lbl_color = np.zeros_like(annotated)
                lbl_color[dr_dict['label_mask'] > 0] = (0, 60, 180)
                cv2.addWeighted(annotated, 1.0, lbl_color, 0.35, 0, annotated)
                lbl_cnts, _ = cv2.findContours(dr_dict['label_mask'],
                                               cv2.RETR_EXTERNAL,
                                               cv2.CHAIN_APPROX_SIMPLE)
                cv2.drawContours(annotated, lbl_cnts, -1, (0, 100, 255), 2)

            col_radius = dr_dict['col_radius']
            for col in dr_dict['colonies']:
                is_cl      = col['is_cluster']
                cnt        = col['contour']
                ws_centers = col['ws_centers']
                n          = col['estimated']
                if is_cl:
                    color = (20, 160, 255)
                    for cx_c, cy_c in ws_centers:
                        cv2.circle(annotated, (cx_c, cy_c), col_radius, color, 2)
                    if show_nums:
                        cc = col['center']
                        cv2.putText(annotated, str(n),
                                    (cc[0] - 8, cc[1] - col_radius - 4),
                                    cv2.FONT_HERSHEY_SIMPLEX, 0.55,
                                    (0, 100, 220), 2, cv2.LINE_AA)
                else:
                    color = (30, 200, 30)
                    cv2.drawContours(annotated, [cnt], -1, color, 2)
                    if show_nums:
                        cv2.circle(annotated, col['center'], 3, color, -1)

            # Надпись на каждой чашке при наличии нескольких
            if len(dishes) > 1:
                lbl_d = f"#{dish_idx + 1}: {dr_dict['total']}"
                cv2.putText(annotated, lbl_d,
                            (dcx - 30, max(15, dcy - dr - 6)),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.7,
                            (80, 180, 255), 2, cv2.LINE_AA)

        # ── Итоговая надпись ──────────────────────────────────────────────
        grand_total = sum(d['total'] for d in dish_results)
        label = f"Kolonii: {grand_total}"
        (lw, lh), _ = cv2.getTextSize(label, cv2.FONT_HERSHEY_SIMPLEX, 1.1, 2)
        cv2.rectangle(annotated, (6, 6), (lw + 18, lh + 18), (30, 30, 30), -1)
        cv2.putText(annotated, label, (12, lh + 10),
                    cv2.FONT_HERSHEY_SIMPLEX, 1.1, (60, 220, 60), 2, cv2.LINE_AA)

        avg_area   = (float(np.mean([d['avg_colony_area'] for d in dish_results]))
                      if dish_results else float(params['min_area'] * 5))
        col_radius = max(4, int(np.sqrt(avg_area / np.pi)))

        return dict(
            total=grand_total,
            colony_count=len(all_colonies),
            cluster_count=sum(1 for c in all_colonies if c['is_cluster']),
            avg_colony_area=avg_area,
            col_radius=col_radius,
            dish=dishes[0] if dishes else (w // 2, h // 2, int(min(h, w) * 0.44)),
            dishes=dishes,
            dish_results=dish_results,
            colonies=all_colonies,
            annotated=annotated,
            img_clean=img.copy(),
            binary=(dish_results[0]['binary'] if dish_results
                    else np.zeros((h, w), dtype=np.uint8)),
            enhanced=(dish_results[0]['enhanced'] if dish_results
                      else np.zeros((h, w), dtype=np.uint8)),
            scale=scale,
            has_label=any(d['has_label'] for d in dish_results),
            hidden_estimate=sum(d['hidden_estimate'] for d in dish_results),
            label_mask=(dish_results[0]['label_mask'] if dish_results else None),
        )


# ─────────────────────────── LEARNING ENGINE ─────────────────────────────────

class LearningEngine:
    """
    Адаптивная фоновая коррекция параметров на основе ручных правок.

    Логика:
      • пользователь убрал много авто-колоний (ПКМ)  → порог завышен  → поднять
      • пользователь добавил много вручную (ЛКМ)      → порог занижен  → опустить
    Используется EMA (экспоненциальное скользящее среднее) с alpha=0.30,
    чтобы модель накапливала опыт постепенно, а не прыгала от каждой сессии.
    """

    ALPHA       = 0.30   # скорость обучения (0 = не учиться, 1 = только последний)
    MIN_AUTO    = 5      # минимум авто-колоний на изображении для учёбы
    MIN_RATIO   = 0.04   # минимальная доля правок, чтобы считать их значимыми

    def __init__(self):
        cfg_dir = Path(os.environ.get('APPDATA', str(Path.home()))) / 'ColonyCounter'
        cfg_dir.mkdir(parents=True, exist_ok=True)
        self._path  = cfg_dir / 'learned_params.json'
        self._state = self._load()

    def _load(self) -> dict:
        try:
            with open(self._path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            return {'threshold_ema': None, 'samples': 0}

    def _save(self):
        try:
            with open(self._path, 'w', encoding='utf-8') as f:
                json.dump(self._state, f, indent=2)
        except Exception:
            pass

    def update(self, auto_count: int, excluded: int, added: int,
               cur_threshold: int) -> int | None:
        """
        Обновить модель на основе правок одного изображения.
        Возвращает новое значение порога если оно изменилось, иначе None.
        """
        if auto_count < self.MIN_AUTO:
            return None

        excl_ratio = excluded / auto_count
        add_ratio  = added  / auto_count
        net = excl_ratio - add_ratio  # > 0 → слишком много ложных; < 0 → пропущены

        if abs(net) < self.MIN_RATIO:
            return None

        # Коррекция: каждые 10% дельты ≈ 1 единица порога, ±5 за шаг максимум
        delta     = max(-5, min(5, round(net * 12)))
        suggested = max(5, min(100, cur_threshold + delta))

        ema = self._state.get('threshold_ema')
        new_ema = suggested if ema is None else (
            self.ALPHA * suggested + (1.0 - self.ALPHA) * ema
        )
        self._state['threshold_ema'] = new_ema
        self._state['samples']       = self._state.get('samples', 0) + 1
        self._save()

        new_thresh = round(new_ema)
        return new_thresh if new_thresh != cur_threshold else None

    @property
    def suggestion(self) -> int | None:
        """Текущая рекомендация модели по порогу (None если данных мало)."""
        ema = self._state.get('threshold_ema')
        if ema is None or self._state.get('samples', 0) < 2:
            return None
        return round(ema)

    @property
    def samples(self) -> int:
        return self._state.get('samples', 0)

    def reset(self):
        self._state = {'threshold_ema': None, 'samples': 0}
        self._save()


# ─────────────────────────── GUI APPLICATION ─────────────────────────────────

class App:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("Colony Counter v1.1 — Подсчёт колоний на чашках Петри")
        self.root.geometry("1360x860")
        self.root.minsize(1000, 680)

        self.processor = ImageProcessor()
        self.learner   = LearningEngine()
        self.image_paths: list[str] = []
        self.image_data: dict[str, dict | None] = {}
        self.current_path: str | None = None
        self._prev_path: str | None = None   # для триггера обучения при смене изображения
        # Хранилище текущих PIL-изображений для переотрисовки по resize
        self._pil_orig: Image.Image | None = None
        self._pil_proc: Image.Image | None = None
        # Ручные отметки: path -> [(img_x, img_y), ...]
        self.manual_marks: dict[str, list] = {}
        # Исключённые авто-колонии: path -> [(cx, cy), ...]
        self.excluded_auto: dict[str, list] = {}
        # Пользовательские отображаемые имена: path -> str
        self.display_names: dict[str, str] = {}
        # Трансформ для пересчёта координат canvas → изображение
        # (scale, offset_x, offset_y, img_w, img_h)
        self._proc_transform: tuple | None = None
        # Ручные переопределения границ чашек: path -> [[cx, cy, r], ...]
        self.dish_overrides: dict[str, list] = {}
        # Режим редактирования границ чашек
        self._dish_edit: bool = False
        self._dish_drag: dict | None = None
        # Zoom / pan state for proc canvas
        self._zoom_proc = 1.0
        self._pan_proc  = [0.0, 0.0]
        self._pan_drag  = None

        # ── Parameters ──────────────────────────────────────────────────
        self.p = dict(
            min_area    = tk.IntVar(value=80),
            max_area    = tk.IntVar(value=3000),
            threshold   = tk.IntVar(value=25),
            filter_bubbles   = tk.BooleanVar(value=True),
            use_watershed    = tk.BooleanVar(value=True),
            show_numbers     = tk.BooleanVar(value=True),
            filter_elongated = tk.BooleanVar(value=True),
            filter_nonconvex = tk.BooleanVar(value=True),
            auto_learn       = tk.BooleanVar(value=True),
            detect_label     = tk.BooleanVar(value=True),
        )
        # Если есть накопленные знания — применяем сразу
        self._apply_learned_params(silent=True)

        self._build_ui()
        self._setup_menu()

    # ── Menu ─────────────────────────────────────────────────────────────────
    def _setup_menu(self):
        mb = tk.Menu(self.root)
        self.root.config(menu=mb)

        fm = tk.Menu(mb, tearoff=0)
        mb.add_cascade(label="Файл", menu=fm)
        fm.add_command(label="Добавить изображения…",
                       command=self._add_images, accelerator="Ctrl+O")
        fm.add_command(label="Обработать все",
                       command=self._process_all)
        fm.add_command(label="Экспорт в Excel…",
                       command=self._export_excel, accelerator="Ctrl+S")
        fm.add_separator()
        fm.add_command(label="Сохранить сессию…",
                       command=self._save_session, accelerator="Ctrl+Shift+S")
        fm.add_command(label="Открыть сессию…",
                       command=self._load_session, accelerator="Ctrl+Shift+O")
        fm.add_separator()
        fm.add_command(label="Сохранить изображение результата…",
                       command=self._export_image_current, accelerator="Ctrl+E")
        fm.add_separator()
        fm.add_command(label="Выход", command=self.root.quit)

        hm = tk.Menu(mb, tearoff=0)
        mb.add_cascade(label="Справка", menu=hm)
        hm.add_command(label="О программе", command=self._about)

        self.root.bind('<Control-o>',       lambda e: self._add_images())
        self.root.bind('<Control-s>',       lambda e: self._export_excel())
        self.root.bind('<Control-S>',       lambda e: self._save_session())
        self.root.bind('<Control-O>',       lambda e: self._load_session())
        self.root.bind('<Control-e>',       lambda e: self._export_image_current())

    # ── UI construction ───────────────────────────────────────────────────────
    def _build_ui(self):
        # ── Status bar (bottom) ──────────────────────────────────────────
        self.status_var = tk.StringVar(value="Готово. Добавьте изображения (Ctrl+O).")
        ttk.Label(self.root, textvariable=self.status_var,
                  relief=tk.SUNKEN, anchor=tk.W,
                  padding=(6, 2)).pack(side=tk.BOTTOM, fill=tk.X)

        # ── Three-column layout ──────────────────────────────────────────
        paned = ttk.PanedWindow(self.root, orient=tk.HORIZONTAL)
        paned.pack(fill=tk.BOTH, expand=True, padx=4, pady=4)

        paned.add(self._build_left(),   weight=0)
        paned.add(self._build_center(), weight=4)
        paned.add(self._build_right(),  weight=0)

    # ── Left panel ───────────────────────────────────────────────────────────
    def _build_left(self):
        f = ttk.Frame(self.root, width=215)
        f.pack_propagate(False)

        ttk.Label(f, text="Список изображений",
                  font=('', 10, 'bold')).pack(pady=(6, 2))

        lf = ttk.Frame(f)
        lf.pack(fill=tk.BOTH, expand=True, padx=5)
        sb = ttk.Scrollbar(lf)
        sb.pack(side=tk.RIGHT, fill=tk.Y)
        self.listbox = tk.Listbox(lf, yscrollcommand=sb.set,
                                   selectmode=tk.SINGLE, font=('', 9),
                                   activestyle='dotbox')
        self.listbox.pack(fill=tk.BOTH, expand=True)
        sb.config(command=self.listbox.yview)
        self.listbox.bind('<<ListboxSelect>>', self._on_select)
        self.listbox.bind('<Double-Button-1>', lambda e: self._rename_image())

        bf = ttk.Frame(f)
        bf.pack(fill=tk.X, padx=5, pady=3)
        ttk.Button(bf, text="+ Добавить",
                   command=self._add_images).pack(side=tk.LEFT, expand=True, fill=tk.X, padx=1)
        ttk.Button(bf, text="− Удалить",
                   command=self._remove_image).pack(side=tk.LEFT, expand=True, fill=tk.X, padx=1)
        bf2 = ttk.Frame(f)
        bf2.pack(fill=tk.X, padx=5, pady=(0, 2))
        ttk.Button(bf2, text="✏ Переименовать",
                   command=self._rename_image).pack(fill=tk.X)

        ttk.Separator(f).pack(fill=tk.X, padx=5, pady=4)
        ttk.Button(f, text="▶  Обработать все",
                   command=self._process_all).pack(fill=tk.X, padx=5, pady=2)
        ttk.Button(f, text="💾  Экспорт в Excel",
                   command=self._export_excel).pack(fill=tk.X, padx=5, pady=2)

        ttk.Separator(f).pack(fill=tk.X, padx=5, pady=4)
        ttk.Label(f, text="Результаты:", font=('', 9, 'bold')).pack(anchor=tk.W, padx=6)

        rsb = ttk.Scrollbar(f)
        rsb.pack(side=tk.RIGHT, fill=tk.Y, padx=(0, 5))
        self.res_text = tk.Text(f, height=14, font=('Courier', 9),
                                 yscrollcommand=rsb.set, state=tk.DISABLED,
                                 bg='#f8f8f8')
        self.res_text.pack(fill=tk.BOTH, expand=True, padx=(5, 0))
        rsb.config(command=self.res_text.yview)

        return f

    # ── Center panel (image canvases) ─────────────────────────────────────────
    def _build_center(self):
        f = ttk.Frame(self.root)

        # ── Toolbar ──────────────────────────────────────────────────────
        tb = ttk.Frame(f)
        tb.pack(fill=tk.X, padx=4, pady=(4, 0))

        ttk.Button(tb, text="↩ Отменить",
                   command=self._undo_manual).pack(side=tk.LEFT, padx=2)
        ttk.Button(tb, text="🗑 Очистить ручные",
                   command=self._clear_manual).pack(side=tk.LEFT, padx=2)
        ttk.Button(tb, text="🔄 Восстановить авто",
                   command=self._restore_auto).pack(side=tk.LEFT, padx=2)
        ttk.Button(tb, text="🔍 Сбросить зум",
                   command=self._reset_zoom).pack(side=tk.LEFT, padx=2)
        ttk.Separator(tb, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y,
                                                    padx=4, pady=2)
        ttk.Button(tb, text="🖼 Сохранить изображение",
                   command=self._export_image_current).pack(side=tk.LEFT, padx=2)
        ttk.Separator(tb, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y,
                                                    padx=4, pady=2)
        self._dish_btn = ttk.Button(tb, text="✋ Границы чашек",
                                    command=self._toggle_dish_edit)
        self._dish_btn.pack(side=tk.LEFT, padx=2)
        ttk.Button(tb, text="↺ Сбросить границы",
                   command=self._reset_dish_overrides).pack(side=tk.LEFT, padx=2)

        ttk.Label(tb,
                  text="  ЛКМ — колония  |  ПКМ — убрать  |  Колесо — зум  |  СКМ — перемещение",
                  font=('', 8), foreground='#555').pack(side=tk.LEFT, padx=10)

        # ── Notebook с холстами ──────────────────────────────────────────
        self.nb = ttk.Notebook(f)
        self.nb.pack(fill=tk.BOTH, expand=True)

        self.canvas_orig = tk.Canvas(self.nb, bg='#1e1e1e', cursor='crosshair')
        self.canvas_proc = tk.Canvas(self.nb, bg='#1e1e1e', cursor='tcross')
        self.nb.add(self.canvas_orig, text="  Оригинал  ")
        self.nb.add(self.canvas_proc, text="  Результат  ")

        # Переотрисовка при изменении размера окна
        self.canvas_orig.bind("<Configure>",
            lambda e: self._redraw(self.canvas_orig, self._pil_orig))
        self.canvas_proc.bind("<Configure>",
            lambda e: self._refresh_proc_canvas())

        # Клики для ручного режима и редактирования границ
        self.canvas_proc.bind("<Button-1>",        self._on_lmb_down)
        self.canvas_proc.bind("<B1-Motion>",        self._on_lmb_motion)
        self.canvas_proc.bind("<ButtonRelease-1>",  self._on_lmb_up)
        self.canvas_proc.bind("<Button-3>",         self._on_proc_rclick)
        # Зум колесом и перемещение средней кнопкой
        self.canvas_proc.bind("<MouseWheel>",        self._on_proc_wheel)
        self.canvas_proc.bind("<ButtonPress-2>",     self._on_proc_pan_start)
        self.canvas_proc.bind("<B2-Motion>",         self._on_proc_pan_move)
        self.canvas_proc.bind("<ButtonRelease-2>",   self._on_proc_pan_end)
        return f

    # ── Right panel (parameters) ──────────────────────────────────────────────
    def _build_right(self):
        outer = ttk.Frame(self.root, width=270)
        outer.pack_propagate(False)

        ttk.Label(outer, text="Параметры",
                  font=('', 10, 'bold')).pack(pady=(6, 2))

        # Scrollable inner area
        canvas = tk.Canvas(outer, width=255, highlightthickness=0)
        vsb = ttk.Scrollbar(outer, orient=tk.VERTICAL, command=canvas.yview)
        inner = ttk.Frame(canvas)
        inner.bind("<Configure>",
                   lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=inner, anchor=tk.NW, width=255)
        canvas.configure(yscrollcommand=vsb.set)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)

        # Bind mouse wheel
        def _on_wheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        canvas.bind_all("<MouseWheel>", _on_wheel)

        self._build_params(inner)
        return outer

    def _build_params(self, parent):
        pad = dict(padx=8, pady=4)

        # ── Detection ────────────────────────────────────────────────────
        lf1 = ttk.LabelFrame(parent, text="Обнаружение колоний", padding=6)
        lf1.pack(fill=tk.X, **pad)
        self._slider(lf1, "Мин. площадь (пикс²):",
                     self.p['min_area'], 10, 500, 10)
        self._slider(lf1, "Макс. площадь одной колонии (пикс²):",
                     self.p['max_area'], 100, 15000, 50)
        self._slider(lf1, "Порог чувствительности (↓ = больше колоний):",
                     self.p['threshold'], 5, 100, 1)

        # ── Options ──────────────────────────────────────────────────────
        lf2 = ttk.LabelFrame(parent, text="Опции", padding=6)
        lf2.pack(fill=tk.X, **pad)
        ttk.Checkbutton(lf2, text="Фильтровать пузыри воздуха",
                         variable=self.p['filter_bubbles']).pack(anchor=tk.W)
        ttk.Checkbutton(lf2, text="Watershed: разделять кластеры",
                         variable=self.p['use_watershed']).pack(anchor=tk.W)
        ttk.Checkbutton(lf2, text="Показывать числа на картинке",
                         variable=self.p['show_numbers']).pack(anchor=tk.W)
        ttk.Checkbutton(lf2, text="Фильтр вытянутых объектов (ручки, царапины)",
                         variable=self.p['filter_elongated']).pack(anchor=tk.W)
        ttk.Checkbutton(lf2, text="Фильтр невыпуклых объектов (мусор, пятна)",
                         variable=self.p['filter_nonconvex']).pack(anchor=tk.W)
        ttk.Checkbutton(lf2, text="Обнаруживать и маскировать этикетку",
                         variable=self.p['detect_label']).pack(anchor=tk.W)

        # ── Auto-learning ─────────────────────────────────────────────────
        lf_learn = ttk.LabelFrame(parent, text="Авто-обучение", padding=6)
        lf_learn.pack(fill=tk.X, **pad)
        ttk.Checkbutton(lf_learn, text="Корректировать параметры по правкам",
                        variable=self.p['auto_learn']).pack(anchor=tk.W)
        self._learn_status_var = tk.StringVar()
        ttk.Label(lf_learn, textvariable=self._learn_status_var,
                  font=('Courier', 8), foreground='#444').pack(anchor=tk.W, pady=(2, 0))
        btn_row = ttk.Frame(lf_learn)
        btn_row.pack(fill=tk.X, pady=(4, 0))
        ttk.Button(btn_row, text="Применить",
                   command=self._apply_learned_params).pack(side=tk.LEFT, expand=True,
                                                             fill=tk.X, padx=(0, 2))
        ttk.Button(btn_row, text="Сброс модели",
                   command=self._reset_learning).pack(side=tk.LEFT, expand=True,
                                                       fill=tk.X, padx=(2, 0))
        self._refresh_learn_label()

        # ── Action buttons ───────────────────────────────────────────────
        ttk.Button(parent, text="▶  Обработать текущее",
                   command=self._process_current).pack(fill=tk.X, **pad)
        ttk.Button(parent, text="⚙  Авто-порог",
                   command=self._auto_threshold).pack(fill=tk.X, padx=8, pady=1)

        ttk.Separator(parent).pack(fill=tk.X, padx=8, pady=6)

        # ── Stats ────────────────────────────────────────────────────────
        lf3 = ttk.LabelFrame(parent, text="Статистика текущего", padding=6)
        lf3.pack(fill=tk.X, **pad)
        self.stats_text = tk.Text(lf3, height=10, font=('Courier', 9),
                                   state=tk.DISABLED, bg='#f8f8f8')
        self.stats_text.pack(fill=tk.BOTH)

        # ── Legend ───────────────────────────────────────────────────────
        lf4 = ttk.LabelFrame(parent, text="Легенда разметки", padding=6)
        lf4.pack(fill=tk.X, **pad)
        leg = tk.Canvas(lf4, height=50, bg='#2b2b2b', highlightthickness=0)
        leg.pack(fill=tk.X)
        leg.after(200, lambda: self._draw_legend(leg))

    @staticmethod
    def _draw_legend(canvas):
        canvas.update_idletasks()
        w = canvas.winfo_width() or 240
        canvas.create_rectangle(10, 10, 30, 30, outline='#1ec81e', width=2)
        canvas.create_text(36, 20, text="одиночная колония",
                           anchor=tk.W, fill='white', font=('', 8))
        canvas.create_rectangle(10, 32, 30, 48, outline='#20a0ff', width=2)
        canvas.create_text(36, 40, text="кластер (число = оценка)",
                           anchor=tk.W, fill='white', font=('', 8))

    def _slider(self, parent, label, var, from_, to, resolution):
        ttk.Label(parent, text=label, font=('', 8)).pack(anchor=tk.W)
        row = ttk.Frame(parent)
        row.pack(fill=tk.X)
        ttk.Scale(row, from_=from_, to=to, variable=var,
                  orient=tk.HORIZONTAL,
                  command=lambda v, vr=var: vr.set(
                      round(float(v) / resolution) * resolution)
                  ).pack(side=tk.LEFT, fill=tk.X, expand=True)
        lbl = ttk.Label(row, width=7)
        lbl.pack(side=tk.RIGHT)

        def _update(*_):
            lbl.config(text=str(int(var.get())))
        var.trace_add('write', _update)
        _update()

    # ── File operations ───────────────────────────────────────────────────────
    def _add_images(self):
        paths = filedialog.askopenfilenames(
            title="Выберите изображения чашек Петри",
            filetypes=[("Изображения", "*.jpg *.jpeg *.png *.bmp *.tiff *.tif"),
                       ("Все файлы", "*.*")]
        )
        added = 0
        for path in paths:
            if path not in self.image_paths:
                self.image_paths.append(path)
                self.image_data[path] = None
                self.display_names[path] = Path(path).name
                self.listbox.insert(tk.END, Path(path).name)
                added += 1
        if added:
            self.status_var.set(f"Добавлено {added} изображений. Нажмите «Обработать все».")

    def _remove_image(self):
        sel = self.listbox.curselection()
        if not sel:
            return
        idx = sel[0]
        path = self.image_paths.pop(idx)
        self.image_data.pop(path, None)
        self.display_names.pop(path, None)
        self.listbox.delete(idx)
        if self.current_path == path:
            self.current_path = None
            self.canvas_orig.delete("all")
            self.canvas_proc.delete("all")
        self._refresh_results_panel()

    def _rename_image(self):
        sel = self.listbox.curselection()
        if not sel:
            return
        idx = sel[0]
        path = self.image_paths[idx]
        current_name = self.display_names.get(path, Path(path).name)
        new_name = simpledialog.askstring(
            "Переименовать",
            "Введите новое отображаемое имя:",
            initialvalue=current_name,
            parent=self.root
        )
        if new_name is None or new_name.strip() == "":
            return
        new_name = new_name.strip()
        self.display_names[path] = new_name
        self.listbox.delete(idx)
        self.listbox.insert(idx, new_name)
        self.listbox.selection_set(idx)
        self._refresh_results_panel()

    # ── Selection ─────────────────────────────────────────────────────────────
    def _on_select(self, _event=None):
        sel = self.listbox.curselection()
        if not sel:
            return
        path = self.image_paths[sel[0]]
        # Обучаемся на предыдущем изображении перед переключением
        if self._prev_path and self._prev_path != path:
            self._do_learning(self._prev_path)
        self._prev_path   = path
        self.current_path = path
        self._zoom_proc = 1.0
        self._pan_proc  = [0.0, 0.0]
        result = self.image_data.get(path)
        if result:
            self._refresh_stats(result)
            self.nb.select(1)
            # После смены вкладки ждём пока canvas получит реальные размеры,
            # затем рисуем через _refresh_proc_canvas — он учитывает ручные метки
            self.root.after(60, self._refresh_proc_canvas)
        else:
            self._pil_proc = None
            self.canvas_proc.delete("all")
            self.nb.select(0)
        self.root.after(60, lambda p=path: self._show_file_on_canvas(p, self.canvas_orig))

    # ── Processing ────────────────────────────────────────────────────────────
    def _get_params(self):
        return {k: v.get() for k, v in self.p.items()}

    def _process_current(self):
        if not self.current_path:
            messagebox.showwarning("Внимание", "Сначала выберите изображение из списка.")
            return
        self._run_processing(self.current_path)
        self._refresh_results_panel()

    def _process_all(self):
        if not self.image_paths:
            messagebox.showwarning("Внимание", "Список изображений пуст.")
            return
        n = len(self.image_paths)
        for i, path in enumerate(self.image_paths):
            self.status_var.set(
                f"Обработка {i + 1}/{n}: {Path(path).name} …")
            self.root.update_idletasks()
            self._run_processing(path, silent=True)
        self._refresh_results_panel()
        if self.current_path and self.image_data.get(self.current_path):
            r = self.image_data[self.current_path]
            self._refresh_stats(r)
            self.nb.select(1)
            self.root.after(60, self._refresh_proc_canvas)
        self.status_var.set(f"Обработка завершена. Обработано {n} изображений.")

    def _run_processing(self, path: str, silent=False):
        try:
            params = self._get_params()
            overrides = self.dish_overrides.get(path)
            if overrides:
                params['dish_overrides'] = overrides
            self.status_var.set(f"Обработка: {Path(path).name} …")
            self.root.update_idletasks()

            result = self.processor.process(path, params)
            self.image_data[path] = result

            if self.current_path == path:
                self._refresh_stats(result)
                self.nb.select(1)
                self.root.after(60, self._refresh_proc_canvas)

            if not silent:
                self.status_var.set(
                    f"{Path(path).name}: найдено {result['total']} колоний "
                    f"({result['cluster_count']} кластеров).")
        except Exception as exc:
            msg = str(exc)
            if not silent:
                messagebox.showerror("Ошибка обработки", msg)
            self.status_var.set(f"Ошибка: {msg}")

    def _auto_threshold(self):
        """Определить пороговое значение методом Оцу по текущему изображению."""
        if not self.current_path:
            messagebox.showwarning("Внимание", "Выберите изображение.")
            return
        try:
            img  = cv2.imread(self.current_path)
            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
            h, w = gray.shape
            cx, cy, r = self.processor.detect_dish(gray)
            mask = np.zeros((h, w), dtype=np.uint8)
            cv2.circle(mask, (cx, cy), int(r * 0.96), 255, -1)
            normalized = self.processor.normalize_background(gray, mask)
            clahe = cv2.createCLAHE(clipLimit=3.0, tileGridSize=(8, 8))
            enhanced = clahe.apply(normalized)
            otsu_val, _ = cv2.threshold(enhanced, 0, 255,
                                        cv2.THRESH_BINARY + cv2.THRESH_OTSU)
            suggested = max(5, int(otsu_val * 0.48))
            self.p['threshold'].set(suggested)
            self.status_var.set(f"Авто-порог: {suggested} (Оцу = {int(otsu_val)}). "
                                "Нажмите «Обработать текущее».")
        except Exception as exc:
            messagebox.showerror("Ошибка", str(exc))

    # ── Display helpers ────────────────────────────────────────────────────────
    def _show_file_on_canvas(self, path: str, canvas: tk.Canvas):
        try:
            pil = Image.open(path)
            if canvas is self.canvas_orig:
                self._pil_orig = pil
            else:
                self._pil_proc = pil
            self._blit(pil, canvas)
        except Exception as exc:
            self.status_var.set(f"Ошибка загрузки: {exc}")

    def _show_cv_on_canvas(self, cv_img, canvas: tk.Canvas):
        rgb = cv2.cvtColor(cv_img, cv2.COLOR_BGR2RGB)
        pil = Image.fromarray(rgb)
        if canvas is self.canvas_orig:
            self._pil_orig = pil
        else:
            self._pil_proc = pil
        self._blit(pil, canvas)

    def _redraw(self, canvas: tk.Canvas, pil_img: "Image.Image | None"):
        if pil_img is not None:
            self._blit(pil_img, canvas)

    def _blit(self, pil_img: Image.Image, canvas: tk.Canvas):
        canvas.update_idletasks()
        cw = canvas.winfo_width()
        ch = canvas.winfo_height()
        if cw <= 1:
            cw = canvas.winfo_reqwidth()
        if ch <= 1:
            ch = canvas.winfo_reqheight()
        if cw <= 1:
            cw = 800
        if ch <= 1:
            ch = 600
        iw, ih = pil_img.size
        fit_scale = min(cw / iw, ch / ih)

        if canvas is self.canvas_proc:
            display_scale = fit_scale * self._zoom_proc
            pan_x, pan_y  = self._pan_proc
        else:
            display_scale = fit_scale
            pan_x, pan_y  = 0.0, 0.0

        # Top-left of the full image in canvas coordinates
        off_x = cw / 2 + pan_x - iw * display_scale / 2
        off_y = ch / 2 + pan_y - ih * display_scale / 2

        if canvas is self.canvas_proc:
            self._proc_transform = (display_scale, off_x, off_y, iw, ih)

        # Crop to only the visible region (performance when zoomed in)
        src_x0 = max(0, int(-off_x / display_scale))
        src_y0 = max(0, int(-off_y / display_scale))
        src_x1 = min(iw, int((cw - off_x) / display_scale) + 1)
        src_y1 = min(ih, int((ch - off_y) / display_scale) + 1)
        if src_x1 <= src_x0 or src_y1 <= src_y0:
            canvas.delete("all")
            return

        cropped  = pil_img.crop((src_x0, src_y0, src_x1, src_y1))
        out_w    = max(1, round((src_x1 - src_x0) * display_scale))
        out_h    = max(1, round((src_y1 - src_y0) * display_scale))
        resized  = cropped.resize((out_w, out_h), Image.LANCZOS)
        draw_x   = round(off_x + src_x0 * display_scale)
        draw_y   = round(off_y + src_y0 * display_scale)

        photo = ImageTk.PhotoImage(resized)
        canvas.delete("all")
        canvas.create_image(draw_x, draw_y, anchor=tk.NW, image=photo)
        canvas.image = photo

    # ── Manual mode ───────────────────────────────────────────────────────────

    def _canvas_to_img(self, cx: int, cy: int):
        """Перевод координат холста в координаты исходного изображения."""
        if self._proc_transform is None:
            return None
        scale, off_x, off_y, img_w, img_h = self._proc_transform
        ix = int((cx - off_x) / scale)
        iy = int((cy - off_y) / scale)
        if 0 <= ix < img_w and 0 <= iy < img_h:
            return (ix, iy)
        return None

    def _ensure_proc_transform(self):
        """Если трансформ устарел (canvas изменил размер), пересчитываем его."""
        if self._pil_proc is None:
            return
        cw = self.canvas_proc.winfo_width()
        ch = self.canvas_proc.winfo_height()
        if cw <= 1 or ch <= 1:
            return
        iw, ih = self._pil_proc.size
        fit_scale    = min(cw / iw, ch / ih)
        display_scale = fit_scale * self._zoom_proc
        pan_x, pan_y  = self._pan_proc
        off_x = cw / 2 + pan_x - iw * display_scale / 2
        off_y = ch / 2 + pan_y - ih * display_scale / 2
        self._proc_transform = (display_scale, off_x, off_y, iw, ih)

    # ── LMB routing (colony click vs dish drag) ───────────────────────────────

    def _on_lmb_down(self, event):
        if self._dish_edit:
            self._dish_drag_start(event)
        else:
            self._on_proc_click(event)

    def _on_lmb_motion(self, event):
        if self._dish_edit and self._dish_drag is not None:
            self._dish_drag_move(event)

    def _on_lmb_up(self, event):
        if self._dish_edit and self._dish_drag is not None:
            self._dish_drag_end(event)

    def _on_proc_click(self, event):
        """ЛКМ: добавить колонию."""
        if not self.current_path:
            return
        self._ensure_proc_transform()
        pt = self._canvas_to_img(event.x, event.y)
        if pt is None:
            return
        self.manual_marks.setdefault(self.current_path, []).append(pt)
        self._refresh_proc_canvas()
        self._refresh_stats_current()
        self._refresh_results_panel()

    def _on_proc_rclick(self, event):
        """ПКМ: удалить ближайшую ручную метку ИЛИ исключить авто-колонию."""
        if not self.current_path or self._dish_edit:
            return
        self._ensure_proc_transform()
        pt = self._canvas_to_img(event.x, event.y)
        if pt is None:
            return
        path = self.current_path
        result = self.image_data.get(path)

        # Собираем кандидатов: ('manual', idx, (x,y)) или ('auto', (cx,cy), (x,y))
        candidates = []
        for i, m in enumerate(self.manual_marks.get(path, [])):
            candidates.append(('manual', i, m))
        if result:
            excl = set(self.excluded_auto.get(path, []))
            for col in result['colonies']:
                for cx_c, cy_c in col['ws_centers']:
                    if (cx_c, cy_c) not in excl:
                        candidates.append(('auto', (cx_c, cy_c), (cx_c, cy_c)))

        if not candidates:
            return

        # Ближайшая точка (манхэттенское расстояние)
        dists = [(abs(c[2][0] - pt[0]) + abs(c[2][1] - pt[1]), i)
                 for i, c in enumerate(candidates)]
        _, nearest_i = min(dists)
        kind, ref, _ = candidates[nearest_i]

        if kind == 'manual':
            self.manual_marks[path].pop(ref)
        else:
            self.excluded_auto.setdefault(path, []).append(ref)

        self._refresh_proc_canvas()
        self._refresh_stats_current()
        self._refresh_results_panel()

    def _undo_manual(self):
        if not self.current_path:
            return
        marks = self.manual_marks.get(self.current_path, [])
        if marks:
            marks.pop()
            self._refresh_proc_canvas()
            self._refresh_stats_current()
            self._refresh_results_panel()

    def _clear_manual(self):
        if not self.current_path:
            return
        if self.manual_marks.get(self.current_path):
            self.manual_marks[self.current_path] = []
            self._refresh_proc_canvas()
            self._refresh_stats_current()
            self._refresh_results_panel()

    def _restore_auto(self):
        """Восстановить все исключённые авто-колонии текущего изображения."""
        if not self.current_path:
            return
        if self.excluded_auto.get(self.current_path):
            self.excluded_auto[self.current_path] = []
            self._refresh_proc_canvas()
            self._refresh_stats_current()
            self._refresh_results_panel()

    # ── Learning ──────────────────────────────────────────────────────────────

    def _do_learning(self, path: str):
        """Обновить модель на основе правок изображения path (вызывается фоново)."""
        if not self.p['auto_learn'].get():
            return
        result = self.image_data.get(path)
        if not result:
            return
        auto_count = result['total']
        excluded   = len(self.excluded_auto.get(path, []))
        added      = len(self.manual_marks.get(path, []))
        if excluded == 0 and added == 0:
            return  # пользователь ничего не менял — не учимся

        new_thresh = self.learner.update(
            auto_count, excluded, added,
            self.p['threshold'].get(),
        )
        if new_thresh is not None:
            self.p['threshold'].set(new_thresh)
            self.status_var.set(
                f"Авто-обучение: порог → {new_thresh}  "
                f"(сессий: {self.learner.samples})"
            )
            self._refresh_learn_label()

    def _apply_learned_params(self, silent=False):
        """Применить текущую рекомендацию модели к ползунку порога."""
        s = self.learner.suggestion
        if s is not None:
            self.p['threshold'].set(s)
            if not silent:
                self.status_var.set(
                    f"Авто-обучение применено: порог = {s}  "
                    f"(сессий: {self.learner.samples})"
                )

    def _refresh_learn_label(self):
        """Обновить строку статуса обучения в панели параметров (если она есть)."""
        s = self.learner.suggestion
        n = self.learner.samples
        if hasattr(self, '_learn_status_var'):
            if s is not None:
                self._learn_status_var.set(f"Модель: {n} сессий  |  порог ≈ {s}")
            else:
                self._learn_status_var.set("Модель: ещё нет данных")

    def _reset_learning(self):
        self.learner.reset()
        self._refresh_learn_label()
        self.status_var.set("Авто-обучение сброшено.")

    # ── Zoom / pan ────────────────────────────────────────────────────────────
    def _on_proc_wheel(self, event):
        if self._pil_proc is None:
            return "break"
        factor   = 1.15 if event.delta > 0 else 1.0 / 1.15
        old_zoom = self._zoom_proc
        new_zoom = max(0.25, min(10.0, old_zoom * factor))
        if new_zoom == old_zoom:
            return "break"
        cw = self.canvas_proc.winfo_width()
        ch = self.canvas_proc.winfo_height()
        iw, ih = self._pil_proc.size
        fit = min(cw / iw, ch / ih)
        old_off_x = cw / 2 + self._pan_proc[0] - iw * fit * old_zoom / 2
        old_off_y = ch / 2 + self._pan_proc[1] - ih * fit * old_zoom / 2
        img_x = (event.x - old_off_x) / (fit * old_zoom)
        img_y = (event.y - old_off_y) / (fit * old_zoom)
        new_off_x = event.x - img_x * fit * new_zoom
        new_off_y = event.y - img_y * fit * new_zoom
        self._pan_proc[0] = new_off_x - (cw / 2 - iw * fit * new_zoom / 2)
        self._pan_proc[1] = new_off_y - (ch / 2 - ih * fit * new_zoom / 2)
        self._zoom_proc = new_zoom
        self._refresh_proc_canvas()
        return "break"

    def _on_proc_pan_start(self, event):
        self._pan_drag = (event.x, event.y, self._pan_proc[0], self._pan_proc[1])

    def _on_proc_pan_move(self, event):
        if self._pan_drag is None:
            return
        sx, sy, px, py = self._pan_drag
        self._pan_proc[0] = px + (event.x - sx)
        self._pan_proc[1] = py + (event.y - sy)
        self._refresh_proc_canvas()

    def _on_proc_pan_end(self, _event=None):
        self._pan_drag = None

    def _reset_zoom(self, _event=None):
        self._zoom_proc = 1.0
        self._pan_proc  = [0.0, 0.0]
        self._refresh_proc_canvas()

    def _make_proc_image(self, path: str):
        """
        Перерисовывает изображение с нуля:
          - авто-колонии минус исключённые
          - ручные добавки (жёлтые)
          - исключённые авто-колонии (красный крест)
        """
        result = self.image_data.get(path)
        if result is None:
            return None

        excl_set  = set(self.excluded_auto.get(path, []))
        marks     = self.manual_marks.get(path, [])
        has_edits = bool(excl_set or marks)

        # Без правок — возвращаем исходную разметку
        if not has_edits:
            return result['annotated']

        base      = result['img_clean'].copy()
        col_r     = result.get('col_radius', 10)
        show_nums = True   # всегда показываем числа в ручном режиме

        # ── Границы чашек (поддержка нескольких + ручные переопределения) ──
        dishes_to_draw = (self.dish_overrides.get(path)
                          or result.get('dishes', [result['dish']]))
        for dcx, dcy, dr in dishes_to_draw:
            cv2.circle(base, (int(dcx), int(dcy)), int(dr), (80, 80, 220), 2)

        # ── Этикетки (перерисовываем из dish_results) ─────────────────────
        for dr_dict in result.get('dish_results', []):
            if dr_dict.get('has_label') and dr_dict.get('label_mask') is not None:
                lbl_color = np.zeros_like(base)
                lbl_color[dr_dict['label_mask'] > 0] = (0, 60, 180)
                cv2.addWeighted(base, 1.0, lbl_color, 0.35, 0, base)
                lbl_cnts, _ = cv2.findContours(dr_dict['label_mask'],
                                               cv2.RETR_EXTERNAL,
                                               cv2.CHAIN_APPROX_SIMPLE)
                cv2.drawContours(base, lbl_cnts, -1, (0, 100, 255), 2)

        # ── Авто-колонии ─────────────────────────────────────────────────
        auto_total = 0
        for col in result['colonies']:
            active_centers   = [c for c in col['ws_centers'] if c not in excl_set]
            inactive_centers = [c for c in col['ws_centers'] if c in excl_set]

            # Рисуем активные
            if active_centers:
                is_cl = len(active_centers) > 1 or col['is_cluster']
                color = (20, 160, 255) if is_cl else (30, 200, 30)
                if not col['is_cluster']:
                    # Одиночная колония — контур
                    cv2.drawContours(base, [col['contour']], -1, color, 2)
                    if show_nums:
                        cv2.circle(base, active_centers[0], 3, color, -1)
                else:
                    for (ax, ay) in active_centers:
                        cv2.circle(base, (ax, ay), col_r, color, 2)
                auto_total += len(active_centers)

            # Рисуем исключённые (красный крест + полупрозрачный)
            for (ex, ey) in inactive_centers:
                cv2.drawMarker(base, (ex, ey), (0, 0, 200),
                               cv2.MARKER_TILTED_CROSS, col_r * 2, 2, cv2.LINE_AA)

        # ── Ручные метки (жёлтые, компактные) ───────────────────────────
        mark_r = max(2, col_r // 3)   # компактная метка — 1/3 радиуса авто-колонии
        for (mx, my) in marks:
            cv2.circle(base, (mx, my), mark_r, (0, 215, 255), 1)
            cv2.drawMarker(base, (mx, my), (0, 215, 255),
                           cv2.MARKER_CROSS, mark_r * 2, 1, cv2.LINE_AA)

        # ── Итоговая надпись ─────────────────────────────────────────────
        grand = auto_total + len(marks)
        label = f"Kolonii: {grand}"
        (lw, lh), _ = cv2.getTextSize(label, cv2.FONT_HERSHEY_SIMPLEX, 1.0, 2)
        cv2.rectangle(base, (6, 6), (lw + 18, lh + 18), (30, 30, 30), -1)
        cv2.putText(base, label, (12, lh + 10),
                    cv2.FONT_HERSHEY_SIMPLEX, 1.0, (60, 220, 60), 2, cv2.LINE_AA)
        return base

    def _refresh_proc_canvas(self):
        """Перерисовать proc-холст с учётом ручных меток."""
        if not self.current_path:
            return
        cv_img = self._make_proc_image(self.current_path)
        if cv_img is not None:
            rgb = cv2.cvtColor(cv_img, cv2.COLOR_BGR2RGB)
            pil = Image.fromarray(rgb)
            self._pil_proc = pil
            self._blit(pil, self.canvas_proc)
        elif self._pil_proc is not None:
            self._blit(self._pil_proc, self.canvas_proc)
        # Рисуем ручки редактирования поверх изображения
        if self._dish_edit:
            self._draw_dish_handles()

    # ── Dish boundary editing ─────────────────────────────────────────────────

    def _toggle_dish_edit(self):
        """Включить/выключить режим редактирования границ чашек."""
        self._dish_edit = not self._dish_edit
        if hasattr(self, '_dish_btn'):
            self._dish_btn.config(
                text="✋ Границы [ВКЛ]" if self._dish_edit else "✋ Границы чашек"
            )
        self._refresh_proc_canvas()

    def _reset_dish_overrides(self):
        """Сбросить ручные границы чашек и переобработать."""
        if not self.current_path:
            return
        if self.current_path in self.dish_overrides:
            del self.dish_overrides[self.current_path]
            self._run_processing(self.current_path)
            self._refresh_results_panel()

    def _get_current_dishes(self):
        """Вернуть текущие границы чашек (переопределения или из результата)."""
        if not self.current_path:
            return []
        overrides = self.dish_overrides.get(self.current_path)
        if overrides:
            return [tuple(d) for d in overrides]
        result = self.image_data.get(self.current_path)
        if result:
            return result.get('dishes', [result['dish']])
        return []

    def _img_to_canvas(self, ix: int, iy: int):
        """Перевод координат изображения в координаты холста."""
        if self._proc_transform is None:
            return None
        scale, off_x, off_y, _iw, _ih = self._proc_transform
        return (off_x + ix * scale, off_y + iy * scale)

    def _draw_dish_handles(self):
        """Рисует интерактивные ручки редактирования границ чашек поверх холста."""
        self.canvas_proc.delete('dish_handle')
        self._ensure_proc_transform()
        if self._proc_transform is None:
            return
        dishes = self._get_current_dishes()
        scale  = self._proc_transform[0]
        for dish_idx, (dcx, dcy, dr) in enumerate(dishes):
            cc = self._img_to_canvas(dcx, dcy)
            if cc is None:
                continue
            ccx, ccy = cc
            cr = dr * scale
            # Пунктирная окружность
            self.canvas_proc.create_oval(
                ccx - cr, ccy - cr, ccx + cr, ccy + cr,
                outline='yellow', width=2, dash=(10, 5), tags='dish_handle'
            )
            # Центральная ручка (перемещение)
            hw = 8
            self.canvas_proc.create_rectangle(
                ccx - hw, ccy - hw, ccx + hw, ccy + hw,
                fill='yellow', outline='black', width=1, tags='dish_handle'
            )
            # Ручки по краям (изменение размера)
            for ex, ey in [(dcx, dcy - dr), (dcx + dr, dcy),
                           (dcx, dcy + dr), (dcx - dr, dcy)]:
                ec = self._img_to_canvas(ex, ey)
                if ec:
                    self.canvas_proc.create_rectangle(
                        ec[0] - 6, ec[1] - 6, ec[0] + 6, ec[1] + 6,
                        fill='cyan', outline='black', width=1, tags='dish_handle'
                    )
            if len(dishes) > 1:
                self.canvas_proc.create_text(
                    ccx, ccy - cr - 14,
                    text=f"Чашка {dish_idx + 1}",
                    fill='yellow', font=('', 10, 'bold'), tags='dish_handle'
                )

    def _dish_drag_start(self, event):
        """Нажатие в режиме редактирования: определяем захваченную ручку."""
        self._ensure_proc_transform()
        pt = self._canvas_to_img(event.x, event.y)
        if pt is None:
            return
        dishes = self._get_current_dishes()
        px, py = pt
        for dish_idx, (dcx, dcy, dr) in enumerate(dishes):
            dist_c = ((px - dcx) ** 2 + (py - dcy) ** 2) ** 0.5
            # Центр → режим перемещения
            if dist_c <= max(15, dr * 0.08):
                self._dish_drag = dict(
                    type='move', dish_idx=dish_idx,
                    start_img=(px, py), orig=(dcx, dcy, dr)
                )
                return
            # Край → режим изменения размера
            if abs(dist_c - dr) <= max(14, dr * 0.07):
                self._dish_drag = dict(
                    type='resize', dish_idx=dish_idx,
                    start_img=(px, py), orig=(dcx, dcy, dr)
                )
                return

    def _dish_drag_move(self, event):
        """Перетаскивание в режиме редактирования."""
        self._ensure_proc_transform()
        pt = self._canvas_to_img(event.x, event.y)
        if pt is None or self._dish_drag is None:
            return
        d        = self._dish_drag
        px, py   = pt
        dcx, dcy, dr = d['orig']
        sx, sy   = d['start_img']
        if d['type'] == 'move':
            self._update_dish_override(d['dish_idx'],
                                       int(dcx + px - sx),
                                       int(dcy + py - sy), dr)
        else:
            new_r = max(20, int(((px - dcx) ** 2 + (py - dcy) ** 2) ** 0.5))
            self._update_dish_override(d['dish_idx'], dcx, dcy, new_r)
        self._refresh_proc_canvas()

    def _dish_drag_end(self, _event=None):
        """Отпускание кнопки: сохранить и переобработать."""
        if self._dish_drag is None:
            return
        self._dish_drag = None
        if self.current_path:
            self._run_processing(self.current_path)

    def _update_dish_override(self, dish_idx, new_cx, new_cy, new_r):
        """Обновить одну запись в dish_overrides текущего изображения."""
        if not self.current_path:
            return
        dishes   = list(self._get_current_dishes())
        ov       = [[int(d[0]), int(d[1]), int(d[2])] for d in dishes]
        while len(ov) <= dish_idx:
            ov.append(ov[-1][:] if ov else [new_cx, new_cy, new_r])
        ov[dish_idx] = [int(new_cx), int(new_cy), int(new_r)]
        self.dish_overrides[self.current_path] = ov

    def _refresh_stats_current(self):
        if self.current_path and self.image_data.get(self.current_path):
            self._refresh_stats(self.image_data[self.current_path])

    def _grand_total(self, path: str) -> tuple[int, int, int]:
        """Возвращает (авто_активных, ручных, исключённых) для пути."""
        result = self.image_data.get(path)
        if not result:
            return (0, 0, 0)
        excl_set = set(self.excluded_auto.get(path, []))
        auto_active = sum(
            len([c for c in col['ws_centers'] if c not in excl_set])
            for col in result['colonies']
        )
        manual_n = len(self.manual_marks.get(path, []))
        excl_n   = len(excl_set)
        return (auto_active, manual_n, excl_n)

    # ── Stats / results panels ────────────────────────────────────────────────
    def _refresh_stats(self, result: dict):
        path = self.current_path
        auto_active, manual_n, excl_n = self._grand_total(path) if path else (result['total'], 0, 0)
        singles = result['colony_count'] - result['cluster_count']
        grand   = auto_active + manual_n
        hidden  = result.get('hidden_estimate', 0)
        has_lbl = result.get('has_label', False)
        dish_results = result.get('dish_results', [])
        dish_lines = ""
        if len(dish_results) > 1:
            for i, dr in enumerate(dish_results, 1):
                dish_lines += f"  Чашка {i}:         {dr['total']:>6}\n"
        text = (
            f"Авто колоний:     {result['total']:>6}\n"
            + dish_lines
            + f"  одиночных:      {singles:>6}\n"
            f"  кластеров:      {result['cluster_count']:>6}\n"
            f"  исключено:     {excl_n:>6}\n"
            f"Ручных меток:     {manual_n:>6}\n"
            f"─────────────────────────\n"
            f"ИТОГО:            {grand:>6}\n"
            + (f"+ скрытых:        {hidden:>6}\n" if hidden else "")
            + f"─────────────────────────\n"
            f"Ср. пл. колонии:  {result['avg_colony_area']:>5.0f} пикс²\n"
            + (f"Этикетка:         обнаружена\n" if has_lbl else "")
            + f"Масштаб загрузки: {result['scale']:.3f}\n"
        )
        self.stats_text.config(state=tk.NORMAL)
        self.stats_text.delete("1.0", tk.END)
        self.stats_text.insert(tk.END, text)
        self.stats_text.config(state=tk.DISABLED)

    def _refresh_results_panel(self):
        lines = []
        total_sum = 0
        for path in self.image_paths:
            name = self.display_names.get(path, Path(path).name)
            short = (name[:20] + "…") if len(name) > 21 else name
            result = self.image_data.get(path)
            if result:
                auto_a, manual_n, excl_n = self._grand_total(path)
                cnt = auto_a + manual_n
                total_sum += cnt
                lines.append(f"{short:<21} {cnt:>4}")
            else:
                lines.append(f"{short:<21}   ---")
        lines.append("─" * 27)
        lines.append(f"{'ИТОГО':<21} {total_sum:>5}")

        self.res_text.config(state=tk.NORMAL)
        self.res_text.delete("1.0", tk.END)
        self.res_text.insert(tk.END, "\n".join(lines))
        self.res_text.config(state=tk.DISABLED)

    # ── Excel export ──────────────────────────────────────────────────────────
    def _export_excel(self):
        processed = {p: r for p, r in self.image_data.items() if r is not None}
        if not processed:
            messagebox.showwarning("Нет данных",
                                   "Сначала обработайте хотя бы одно изображение.")
            return

        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile=f"colonies_{datetime.date.today()}.xlsx",
            filetypes=[("Excel файл", "*.xlsx")],
            title="Сохранить результаты в Excel"
        )
        if not save_path:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Результаты"

            # ── Styles ──────────────────────────────────────────────────
            hdr_fill  = PatternFill("solid", fgColor="2E4DA0")
            hdr_font  = Font(color="FFFFFF", bold=True, size=11)
            tot_fill  = PatternFill("solid", fgColor="D6E4BC")
            alt_fill  = PatternFill("solid", fgColor="EEF2FB")
            thin = Side(style='thin', color="AAAAAA")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            center = Alignment(horizontal='center', vertical='center')
            left   = Alignment(horizontal='left',   vertical='center')

            # ── Title ───────────────────────────────────────────────────
            ws.merge_cells("A1:H1")
            tc = ws["A1"]
            tc.value = "Результаты подсчёта колоний — Colony Counter v1.0"
            tc.font = Font(bold=True, size=14, color="1A1A6E")
            tc.alignment = center

            ws["A2"] = (f"Дата экспорта: "
                        f"{datetime.datetime.now().strftime('%d.%m.%Y  %H:%M')}")
            ws["A2"].font = Font(italic=True, size=9, color="555555")

            # ── Header row ──────────────────────────────────────────────
            headers = ["№", "Имя файла", "Авто (колоний)",
                       "Исключено", "Ручных меток", "ИТОГО колоний",
                       "Одиночных объектов", "Кластеров",
                       "Ср. площадь 1 колонии (пикс²)",
                       "Путь к файлу"]
            for col, h in enumerate(headers, 1):
                c = ws.cell(row=4, column=col, value=h)
                c.fill = hdr_fill
                c.font = hdr_font
                c.alignment = center
                c.border = border

            col_widths = [4, 30, 16, 12, 14, 16, 20, 14, 28, 50]
            for i, cw in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = cw

            ws.row_dimensions[4].height = 20

            # ── Data ────────────────────────────────────────────────────
            row = 5
            total_sum = 0
            for idx, (path, result) in enumerate(
                    [(p, self.image_data[p]) for p in self.image_paths
                     if self.image_data.get(p)], 1):
                singles          = result['colony_count'] - result['cluster_count']
                auto_a, manual_n, excl_n = self._grand_total(path)
                grand            = auto_a + manual_n
                fill = alt_fill if idx % 2 == 0 else PatternFill()
                vals = [idx, self.display_names.get(path, Path(path).name), result['total'],
                        f"-{excl_n}" if excl_n else 0, manual_n, grand,
                        singles, result['cluster_count'],
                        round(result['avg_colony_area'], 1),
                        path]
                for col, val in enumerate(vals, 1):
                    c = ws.cell(row=row, column=col, value=val)
                    c.border = border
                    c.fill = fill
                    c.alignment = center if col not in (2, 10) else left
                    if col == 6:
                        c.font = Font(bold=True)
                total_sum += grand
                row += 1

            # ── Total row ───────────────────────────────────────────────
            ws.cell(row=row, column=2, value="ИТОГО").font = Font(bold=True)
            total_cell = ws.cell(row=row, column=3, value=total_sum)
            total_cell.font = Font(bold=True, size=12)
            for col in range(1, 11):
                c = ws.cell(row=row, column=col)
                c.fill = tot_fill
                c.border = border
                c.alignment = center

            ws.freeze_panes = "A5"
            ws.auto_filter.ref = f"A4:J{row - 1}"

            # ── Detail sheets ────────────────────────────────────────────
            for path in self.image_paths:
                result = self.image_data.get(path)
                if not result:
                    continue
                sheet_name = Path(path).stem[:28]
                for ch in r'[]:*?/\\':
                    sheet_name = sheet_name.replace(ch, "_")
                ws2 = wb.create_sheet(title=sheet_name)

                manual_n = len(self.manual_marks.get(path, []))
                ws2["A1"] = f"Детали: {Path(path).name}"
                ws2["A1"].font = Font(bold=True, size=12)
                ws2["A2"] = (f"Авто: {result['total']}  |  "
                             f"Ручных: {manual_n}  |  "
                             f"ИТОГО: {result['total'] + manual_n}")
                ws2["A3"] = (f"Средняя площадь одиночной колонии: "
                             f"{result['avg_colony_area']:.1f} пикс²")

                det_headers = ["№ объекта", "X центра", "Y центра",
                               "Площадь (пикс²)", "Округлость",
                               "Число в кластере", "Тип"]
                for ci, h in enumerate(det_headers, 1):
                    c = ws2.cell(row=5, column=ci, value=h)
                    c.fill = hdr_fill
                    c.font = hdr_font
                    c.alignment = center
                    c.border = border
                    ws2.column_dimensions[get_column_letter(ci)].width = 16

                for ji, col_data in enumerate(result['colonies'], 1):
                    r2 = ji + 5
                    feat = col_data['feat']
                    ctype = "кластер" if col_data['is_cluster'] else "одиночная"
                    row_vals = [ji, feat['cx'], feat['cy'],
                                round(feat['area'], 1),
                                round(feat['circularity'], 3),
                                col_data['estimated'], ctype]
                    for ci, val in enumerate(row_vals, 1):
                        c = ws2.cell(row=r2, column=ci, value=val)
                        c.border = border
                        c.alignment = center

                # Ручные метки
                marks = self.manual_marks.get(path, [])
                if marks:
                    sep_row = len(result['colonies']) + 6
                    ws2.cell(row=sep_row, column=1,
                             value="── Ручные метки ──").font = Font(bold=True, italic=True)
                    for mi, (mx, my) in enumerate(marks, 1):
                        r3 = sep_row + mi
                        row_vals = [f"M{mi}", mx, my, "—", "—", 1, "ручная"]
                        for ci, val in enumerate(row_vals, 1):
                            c = ws2.cell(row=r3, column=ci, value=val)
                            c.border = border
                            c.alignment = center

                ws2.freeze_panes = "A6"

            wb.save(save_path)
            messagebox.showinfo("Экспорт завершён",
                                f"Файл сохранён:\n{save_path}")
            self.status_var.set(f"Сохранено: {save_path}")
        except Exception as exc:
            messagebox.showerror("Ошибка экспорта", str(exc))

    # ── Session save / load ───────────────────────────────────────────────────

    def _save_session(self):
        save_path = filedialog.asksaveasfilename(
            defaultextension=".colsession",
            initialfile=f"session_{datetime.date.today()}.colsession",
            filetypes=[("Colony Counter Session", "*.colsession"),
                       ("JSON", "*.json")],
            title="Сохранить сессию"
        )
        if not save_path:
            return
        try:
            params_dict = {k: v.get() for k, v in self.p.items()}
            images_data = []
            for path in self.image_paths:
                images_data.append({
                    'path': path,
                    'display_name': self.display_names.get(path, Path(path).name),
                    'manual_marks': [list(m) for m in
                                     self.manual_marks.get(path, [])],
                    'excluded_auto': [list(c) for c in
                                      self.excluded_auto.get(path, [])],
                    'dish_overrides': self.dish_overrides.get(path, []),
                })
            session = {
                'version': 1,
                'current_path': self.current_path,
                'params': params_dict,
                'images': images_data,
            }
            with open(save_path, 'w', encoding='utf-8') as f:
                json.dump(session, f, indent=2, ensure_ascii=False)
            self.status_var.set(f"Сессия сохранена: {Path(save_path).name}")
        except Exception as exc:
            messagebox.showerror("Ошибка сохранения сессии", str(exc))

    def _load_session(self):
        load_path = filedialog.askopenfilename(
            filetypes=[("Colony Counter Session", "*.colsession"),
                       ("JSON", "*.json")],
            title="Открыть сессию"
        )
        if not load_path:
            return
        try:
            with open(load_path, 'r', encoding='utf-8') as f:
                session = json.load(f)

            # ── Очищаем текущее состояние ────────────────────────────────
            self.image_paths.clear()
            self.image_data.clear()
            self.manual_marks.clear()
            self.excluded_auto.clear()
            self.display_names.clear()
            self.dish_overrides.clear()
            self.listbox.delete(0, tk.END)
            self.current_path = None
            self._prev_path   = None
            self._pil_orig    = None
            self._pil_proc    = None
            self.canvas_orig.delete("all")
            self.canvas_proc.delete("all")

            # ── Восстанавливаем параметры ────────────────────────────────
            for k, v in session.get('params', {}).items():
                if k in self.p:
                    self.p[k].set(v)

            # ── Восстанавливаем изображения и метки ──────────────────────
            missing = []
            for img_data in session.get('images', []):
                path = img_data['path']
                if not Path(path).exists():
                    missing.append(path)
                    continue
                self.image_paths.append(path)
                self.image_data[path] = None
                disp = img_data.get('display_name', Path(path).name)
                self.display_names[path] = disp
                self.listbox.insert(tk.END, disp)
                marks = [tuple(m) for m in img_data.get('manual_marks', [])]
                if marks:
                    self.manual_marks[path] = marks
                excl = [tuple(c) for c in img_data.get('excluded_auto', [])]
                if excl:
                    self.excluded_auto[path] = excl
                ov = img_data.get('dish_overrides', [])
                if ov:
                    self.dish_overrides[path] = ov

            # ── Выбираем текущий файл ────────────────────────────────────
            cur = session.get('current_path')
            if cur and cur in self.image_paths:
                idx = self.image_paths.index(cur)
                self.listbox.selection_set(idx)
                self.listbox.see(idx)

            msg = f"Сессия загружена: {len(self.image_paths)} изображений."
            if missing:
                msg += f" Не найдено: {len(missing)} файл(ов)."
                messagebox.showwarning(
                    "Файлы не найдены",
                    "Следующие файлы не найдены:\n" +
                    "\n".join(missing[:8]) +
                    ("\n…" if len(missing) > 8 else "")
                )
            self.status_var.set(msg +
                                " Нажмите «Обработать все» для восстановления.")
        except Exception as exc:
            messagebox.showerror("Ошибка загрузки сессии", str(exc))

    # ── Export result image ───────────────────────────────────────────────────

    def _export_image_current(self):
        if not self.current_path:
            messagebox.showwarning("Внимание", "Выберите изображение.")
            return
        cv_img = self._make_proc_image(self.current_path)
        if cv_img is None:
            messagebox.showwarning("Нет результата",
                                   "Сначала обработайте изображение.")
            return
        stem = Path(self.current_path).stem
        save_path = filedialog.asksaveasfilename(
            defaultextension=".png",
            initialfile=f"{stem}_result.png",
            filetypes=[("PNG", "*.png"), ("JPEG", "*.jpg *.jpeg"),
                       ("BMP", "*.bmp")],
            title="Сохранить результат как изображение"
        )
        if not save_path:
            return
        try:
            cv2.imwrite(save_path, cv_img)
            self.status_var.set(f"Изображение сохранено: {Path(save_path).name}")
        except Exception as exc:
            messagebox.showerror("Ошибка сохранения", str(exc))

    # ── About ─────────────────────────────────────────────────────────────────
    def _about(self):
        messagebox.showinfo(
            "О программе",
            "Colony Counter v1.1\n\n"
            "Автоматический подсчёт колоний бактерий\n"
            "на чашках Петри (проходящий свет).\n\n"
            "Алгоритм:\n"
            "  1. Обнаружение чашки (Hough Circles)\n"
            "  2. Обнаружение и маскирование этикетки\n"
            "  3. Вычитание фона (морф. открытие)\n"
            "  4. CLAHE — локальный контраст\n"
            "  5. Бинаризация по порогу\n"
            "  6. Морфологическая очистка (OPEN + CLOSE)\n"
            "  7. Per-component watershed (skimage)\n"
            "  8. Оценка числа колоний в кластере\n"
            "  9. Оценка скрытых колоний под этикеткой\n\n"
            "Зелёные контуры — одиночные колонии.\n"
            "Синие контуры   — кластеры (число = оценка).\n"
            "Синяя подсветка — область этикетки.\n\n"
            "Совет: используйте «Авто-порог» для\n"
            "первоначальной настройки параметров."
        )


# ─────────────────────────── ENTRY POINT ─────────────────────────────────────

def main():
    root = tk.Tk()
    try:
        root.tk.call('tk', 'scaling', 1.25)
    except Exception:
        pass

    # Apply a nicer theme if available
    style = ttk.Style(root)
    for theme in ('vista', 'winnative', 'clam', 'alt'):
        try:
            style.theme_use(theme)
            break
        except Exception:
            continue

    App(root)
    root.mainloop()


if __name__ == "__main__":
    main()
